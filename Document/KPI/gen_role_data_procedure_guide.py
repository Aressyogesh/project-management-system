"""
Role-Wise Data Entry Procedure Guide — v2
Detailed SOP with: QA role, granular sub-steps, Common Mistakes column.
Output: Document/KPI/Role-Wise Data Entry Procedure Guide.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

C = {
    "dark":     "1E293B", "slate":  "334155", "white": "FFFFFF",
    "border":   "CBD5E1", "thick":  "94A3B8",
    "row_a":    "F8FAFC", "row_b":  "FFFFFF",
    "warn":     "FFF0F0", "warn_t": "991B1B",
    "setup_bg": "F0F9FF", "setup_t":"0C4A6E", "setup_h":"0369A1",
    "daily_bg": "F0FDF4", "daily_t":"14532D", "daily_h":"16A34A",
    "week_bg":  "FFFBEB", "week_t": "78350F", "week_h": "D97706",
    "spr_bg":   "F5F3FF", "spr_t":  "3B0764", "spr_h":  "7C3AED",
    "mon_bg":   "FFF1F2", "mon_t":  "881337", "mon_h":  "E11D48",
    "klog_bg":  "F3E8FF", "klog_t": "4C1D95",
    "pm_t":  "5B21B6", "pm_b":  "F5F3FF",
    "tl_t":  "0F766E", "tl_b":  "F0FDFA",
    "dev_t": "1D4ED8", "dev_b": "EFF6FF",
    "qa_t":  "9A3412", "qa_b":  "FFF7ED",
    "des_t": "9F1239", "des_b": "FFF1F2",
    "dvo_t": "92400E", "dvo_b": "FFFBEB",
    "ts_b":  "DBEAFE", "ts_t":  "1E3A8A",
    "pva_b": "D1FAE5", "pva_t": "064E3B",
    "cap_b": "FCE7F3", "cap_t": "831843",
    "kpi_b": "F3E8FF", "kpi_t": "4C1D95",
    "mis_b": "FEF9C3", "mis_t": "713F12",
}

TIMING = {
    "PROJECT SETUP": (C["setup_bg"], C["setup_t"], C["setup_h"]),
    "DAILY":         (C["daily_bg"], C["daily_t"], C["daily_h"]),
    "WEEKLY":        (C["week_bg"],  C["week_t"],  C["week_h"]),
    "SPRINT/PHASE":  (C["spr_bg"],   C["spr_t"],   C["spr_h"]),
    "MONTHLY":       (C["mon_bg"],   C["mon_t"],   C["mon_h"]),
    "SELF-LOG":      (C["klog_bg"],  C["klog_t"],  "7C3AED"),
    "ADMIN TRIGGER": (C["warn"],     C["warn_t"],  "BE123C"),
}

NCOLS = 10

def fill(h): return PatternFill("solid", fgColor=h)
def fn(bold=False, color="1E293B", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def al(h="left", v="top", wrap=True, **_):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(c=None):
    s = Side(style="thin", color=c or C["border"])
    return Border(left=s, right=s, top=s, bottom=s)
def thick_b():
    s = Side(style="medium", color=C["thick"])
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, r, c, v="", bg=None, bold=False, color="1E293B",
        ha="left", va="top", size=9, italic=False, b=True):
    cl = ws.cell(row=r, column=c, value=v)
    if bg: cl.fill = fill(bg)
    cl.font  = fn(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = al(ha, va, wrap_text=True)
    cl.border = bdr() if b else Border()
    return cl

def page_hdr(ws, r, title, bg, tc="FFFFFF", size=13, h=36):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    cl = ws.cell(row=r, column=1, value=title)
    cl.fill = fill(bg); cl.font = fn(True, tc, size)
    cl.alignment = al("center", "center"); cl.border = thick_b()
    ws.row_dimensions[r].height = h

def sec_hdr(ws, r, label, bg, tc="FFFFFF"):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    cl = ws.cell(row=r, column=1, value=f"  {label}")
    cl.fill = fill(bg); cl.font = fn(True, tc, 9)
    cl.alignment = al("left", "center"); cl.border = thick_b()
    ws.row_dimensions[r].height = 18

def info(ws, r, text, bg, tc, h=44):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    cl = ws.cell(row=r, column=1, value=text)
    cl.fill = fill(bg); cl.font = fn(False, tc, 9)
    cl.alignment = al("left", "top"); cl.border = thick_b()
    ws.row_dimensions[r].height = h

COL_HDRS = [
    "#", "When", "Action — Sub-steps", "Where in PMS\n(Navigation Path)",
    "Fields / Exact Values & Examples",
    "⚠ Common Mistake\n(What Goes Wrong)",
    "Timesheet\nReport", "Planned\nvs Actual", "Capacity\nReport", "KPI\nImpact",
]
COL_W = [4, 13, 38, 26, 38, 32, 11, 11, 11, 11]

def col_hdr_row(ws, r):
    IMPACT = {6: (C["mis_b"], C["mis_t"]), 7: (C["ts_b"], C["ts_t"]),
              8: (C["pva_b"], C["pva_t"]), 9: (C["cap_b"], C["cap_t"]),
              10:(C["kpi_b"], C["kpi_t"])}
    for ci, h in enumerate(COL_HDRS, 1):
        bg, tc = IMPACT.get(ci, (C["dark"], C["white"]))
        cl = ws.cell(row=r, column=ci, value=h)
        cl.fill = fill(bg); cl.font = fn(True, tc, 8)
        cl.alignment = al("center", "center"); cl.border = bdr()
    ws.row_dimensions[r].height = 28

def set_widths(ws):
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def step_row(ws, r, num, timing, action, where, fields, mistake="",
             ts="", pva="", cap="", kpi="", row_h=70, skip=False):
    rb = C["warn"] if skip else (C["row_a"] if r % 2 == 0 else C["row_b"])
    tbg, ttc, _ = TIMING.get(timing, (rb, C["dark"], C["dark"]))
    cel(ws, r, 1,  num,    tbg,       True,  ttc,        "center")
    cel(ws, r, 2,  timing, tbg,       True,  ttc,        "center", size=8)
    cel(ws, r, 3,  action, rb,        False, C["warn_t"] if skip else C["dark"])
    cel(ws, r, 4,  where,  rb,        False, C["slate"],  size=8)
    cel(ws, r, 5,  fields, tbg,       False, ttc,         size=8)
    cel(ws, r, 6,  mistake,C["mis_b"],False, C["mis_t"],  size=8, italic=True)
    cel(ws, r, 7,  ts,     C["ts_b"], True,  C["ts_t"],  "center")
    cel(ws, r, 8,  pva,    C["pva_b"],True,  C["pva_t"], "center")
    cel(ws, r, 9,  cap,    C["cap_b"],True,  C["cap_t"], "center")
    cel(ws, r, 10, kpi,    C["kpi_b"],True,  C["kpi_t"], "center")
    ws.row_dimensions[r].height = row_h

def build_sheet(title, role_tc, role_bg, subtitle, intro, steps):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    page_hdr(ws, 1, f"{title.upper()}  —  {subtitle}", role_tc)
    info(ws, 2, intro, role_bg, role_tc, h=48)
    legend = ("  IMPACT KEY:   FEEDS = directly populates this report   "
              "PARTIAL = report shows incomplete/misleading data   "
              "NO EFFECT = not relevant for this report")
    info(ws, 3, legend, C["dark"], C["white"], h=15)
    col_hdr_row(ws, 4)
    set_widths(ws)
    ws.freeze_panes = "A5"
    r = 5
    for s in steps:
        if isinstance(s, str):
            _, _, hc = TIMING.get(s, (C["slate"], C["white"], C["slate"]))
            sec_hdr(ws, r, s, hc); r += 1
        else:
            cp = {k: v for k, v in s.items() if k != "row_h"}
            step_row(ws, r, row_h=s.get("row_h", 70), **cp); r += 1
    return ws


# ── SHEET 1: Overview ─────────────────────────────────────────────────────────
ws_ov = wb.active
ws_ov.title = "0. Overview"
ws_ov.sheet_view.showGridLines = False
page_hdr(ws_ov, 1,
    "ROLE-WISE DATA ENTRY PROCEDURE GUIDE  v2  —  "
    "Exact steps every role must follow for accurate Reports & KPI",
    C["dark"], size=12)
info(ws_ov, 2,
    "Share this file with ALL team members. Each coloured tab is that role's SOP. "
    "The Quick Checklist (last tab) can be printed and pinned at desks. "
    "Every missed step below directly causes 0 pts or wrong data in at least one report.",
    "EFF6FF", "1E3A8A", h=40)

ov_h = ["Role", "System Role in PMS", "Critical Daily Action",
        "Critical Sprint Action", "Monthly Must-Do", "Biggest KPI Risk if Skipped"]
for ci, h in enumerate(ov_h, 1):
    ws_ov.cell(row=3, column=ci, value=h).fill = fill(C["dark"])
    cl = ws_ov.cell(row=3, column=ci)
    cl.font = fn(True, C["white"], 9); cl.alignment = al("center", "center")
    cl.border = bdr()
ws_ov.row_dimensions[3].height = 22

ov_rows = [
    ("PROJECT MANAGER", "ADMIN or EMPLOYEE\n(project role = PM)", C["pm_t"], C["pm_b"],
     "Log own timesheet + approve team timesheets",
     "Create sprint with items (est. hrs, due dates, story points)",
     "Enter manual KPI scores for team by 5th of month",
     "Team loses up to 40 auto-KPI pts if sprint/items not set up correctly"),
    ("TEAM LEAD", "EMPLOYEE\n(project role = TL)", C["tl_t"], C["tl_b"],
     "Log own timesheet + verify team item statuses",
     "Verify all items have estimated hours and due dates before sprint starts",
     "Log Learning + Innovation self-logs",
     "Estimation Accuracy = 0 for whole team if est. hrs missing"),
    ("DEVELOPER", "EMPLOYEE\n(project role = Developer)", C["dev_t"], C["dev_b"],
     "Log timesheet against every work item worked on",
     "Move ALL completed items to QA_DONE before sprint closes",
     "Log Learning + Innovation self-logs; submit leave requests",
     "Sprint Reliability + Delivery Timeliness + Throughput all = 0 if items not in QA_DONE"),
    ("QA ENGINEER", "EMPLOYEE\n(project role = QA)", C["qa_t"], C["qa_b"],
     "Log timesheet against testing tasks + update work item status",
     "Log all bugs with Severity, Classification, and Responsible User filled",
     "Log Learning (testing tools) + Innovation (test automation)",
     "QA contribution invisible in all reports if timesheet not logged"),
    ("DESIGNER", "EMPLOYEE\n(project role = Designer)", C["des_t"], C["des_b"],
     "Log timesheet against design tasks",
     "Move approved design deliverables to QA_DONE SAME DAY as approval",
     "Log Learning (design tools) + Innovation (AI design tools)",
     "Delivery Timeliness = 0 if QA_DONE status is set late after approval"),
    ("DEVOPS", "EMPLOYEE\n(project role = DevOps)", C["dvo_t"], C["dvo_b"],
     "Log timesheet against infra/pipeline work items",
     "Create work items for ALL DevOps tasks before starting work",
     "Log Innovation self-log (biggest KPI opportunity for DevOps)",
     "DevOps contribution invisible in all reports if no work items created"),
]

for ri, row in enumerate(ov_rows, 4):
    rb = C["row_a"] if ri % 2 == 0 else C["row_b"]
    cel(ws_ov, ri, 1, row[0], row[3], True,  row[2], "center", size=10)
    cel(ws_ov, ri, 2, row[1], rb,     False, C["slate"], size=9)
    cel(ws_ov, ri, 3, row[4], C["daily_bg"], False, C["daily_t"], size=9)
    cel(ws_ov, ri, 4, row[5], C["spr_bg"],   False, C["spr_t"],   size=9)
    cel(ws_ov, ri, 5, row[6], C["mon_bg"],   False, C["mon_t"],   size=9)
    cel(ws_ov, ri, 6, row[7], C["warn"],     False, C["warn_t"],  size=9, italic=True)
    ws_ov.row_dimensions[ri].height = 60

for ci, w in enumerate([20, 22, 32, 36, 32, 42], 1):
    ws_ov.column_dimensions[get_column_letter(ci)].width = w
ws_ov.freeze_panes = "A4"


# ── SHEET 2: Project Manager ──────────────────────────────────────────────────
pm_steps = [
    "PROJECT SETUP",
    dict(num="P1", timing="PROJECT SETUP",
         action="Create the project\n1. Left sidebar → Projects → + New Project\n2. Fill Project Name (e.g. 'PMS v2 Development')\n3. Type: SCRUM (sprints) or KANBAN (continuous)\n4. Status: ACTIVE\n5. Start Date + End Date\n6. Save",
         where="Left sidebar\n→ Projects\n→ + New Project",
         fields="• Project Name: descriptive (e.g. 'PMS Backend v2')\n• Type: SCRUM if you work in sprints, KANBAN for support/ops work\n• Status: ACTIVE (required — ARCHIVE/ON_HOLD projects don't appear on board)\n• Start Date: project kick-off date\n• End Date: expected delivery date",
         mistake="Setting Status = ON_HOLD or ARCHIVE on creation — the project won't appear for team members to find and log work against.",
         ts="PARTIAL", pva="PARTIAL", cap="FEEDS", kpi="FEEDS", row_h=100),

    dict(num="P2", timing="PROJECT SETUP",
         action="Add all team members with correct project roles\n1. Open the project → Members tab\n2. Click + Add Member\n3. Select user from dropdown\n4. Set their Project Role\n5. Repeat for each team member\n6. Save",
         where="Projects → [Project Name]\n→ Members tab\n→ + Add Member",
         fields="Project Roles to assign:\n• PROJECT_MANAGER — can approve timesheets, see all reports\n• TEAM_LEAD — can view team reports (new RBAC)\n• DEVELOPER — codes, logs timesheet, moves items\n• QA — tests, logs bugs, moves items to QA_DONE\n• DESIGNER — designs, logs timesheet, moves design items\n• DEVOPS — infra, pipelines, deployment tasks",
         mistake="Assigning wrong project roles (e.g. Developer instead of QA) — RBAC controls what they see in Reports. TLs won't see team reports if not assigned as TEAM_LEAD.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=110),

    dict(num="P3", timing="PROJECT SETUP",
         action="Configure Holidays (one time per year)\n1. Left sidebar → Settings → Holidays\n2. Click + Add Holiday\n3. Enter Holiday Name + Date\n4. Toggle 'Recurring' ON for annual holidays\n5. Save\n6. Repeat for all public holidays this year",
         where="Left sidebar → Settings\n→ Holidays\n→ + Add Holiday",
         fields="Example entries:\n• Name: Republic Day | Date: 2026-01-26 | Recurring: ON\n• Name: Independence Day | Date: 2026-08-15 | Recurring: ON\n• Name: Diwali | Date: 2026-10-20 | Recurring: OFF (date changes each year)\n\nAdd ALL company holidays at the start of each year.",
         mistake="Skipping this step — Capacity Report shows ALL days as 'available' even on national holidays. PM plans meetings on holidays. Capacity data is completely wrong.",
         ts="NO", pva="NO", cap="FEEDS", kpi="NO", row_h=110),

    dict(num="P4", timing="PROJECT SETUP",
         action="Verify Portal Configuration\n1. Settings → Portal Config\n2. Check Working Days: Mon=ON, Tue=ON, Wed=ON, Thu=ON, Fri=ON, Sat=OFF, Sun=OFF\n3. Business Hours: 09:00 – 18:00\n4. Task Duration Unit: Hours\n5. Save if any changes made",
         where="Left sidebar → Settings\n→ Portal Config",
         fields="• Working Days: Mon–Fri = ON (adjust for your company schedule)\n• Business Start Time: 09:00 | End Time: 18:00\n• Date Format: DD-MM-YYYY\n• Task Duration: Hours (CRITICAL — if set to Days, Estimation Accuracy breaks)\n• Timezone: match your office timezone",
         mistake="Task Duration left as 'Days' — when you enter Estimated Hours = 8 it's treated as 8 DAYS not 8 hours. Planned vs Actual and Estimation Accuracy scores will all be wrong.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=100),

    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Create a new Sprint before the sprint starts\n1. Board → Sprint Manager (top right button)\n2. Click + New Sprint\n3. Enter Sprint Name (e.g. 'Sprint 5 – June 2026')\n4. Set Goal (e.g. 'Ship dashboard filters + fix auth bugs')\n5. Start Date: first day of sprint\n6. End Date: last day of sprint\n7. Save\n8. On sprint start day: click Activate",
         where="Board → Sprint Manager\n→ + New Sprint\n\nThen on day 1:\nBoard → Sprint Manager\n→ Activate Sprint",
         fields="• Name: 'Sprint [N] – [Month] [Year]' (consistent naming helps reporting)\n• Goal: short 1-line objective (e.g. 'Complete login v2 + API rate limiting')\n• Start Date: e.g. 2026-06-01 (Monday)\n• End Date: e.g. 2026-06-30 (last working day)\n\nACTIVATE the sprint on start day — items in inactive sprints don't count for Sprint Reliability KPI",
         mistake="Creating the sprint but forgetting to Activate it — ALL items in the sprint show 0 on Sprint Reliability because the sprint is in DRAFT status. Sprint Reliability = 0 for entire team.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=110),

    dict(num="S2", timing="SPRINT/PHASE",
         action="Create work items — fill EVERY KPI-critical field\n1. Board → + New Work Item\n2. Set Type (TASK / STORY / BUG)\n3. Enter Title\n4. Assign to team member\n5. Set Sprint (select active sprint)\n6. Set Story Points (1–5)\n7. Set Estimated Hours (CRITICAL)\n8. Set Due Date (CRITICAL)\n9. Set Start Date\n10. Save — repeat for all items",
         where="Board → + New Work Item\n(or Board → List View\n→ + Add Item)",
         fields="KPI-CRITICAL fields (never leave blank):\n• Type: TASK (feature work), STORY (user story), BUG (defect), SUB_TASK\n• Assignee: specific person (not blank)\n• Sprint: active sprint name\n• Story Points: 1=trivial, 2=simple, 3=medium, 4=complex, 5=very complex\n• Estimated Hours: 4=half day, 8=1 day, 16=2 days, 24=3 days, 40=1 week\n• Due Date: realistic date within sprint (e.g. Jun 15 for mid-sprint delivery)\n• Start Date: when work begins (e.g. Jun 02)\n\nMissing est. hrs → Estimation Accuracy = 0\nMissing due date → Delivery Timeliness = 0\nMissing sprint → Sprint Reliability = 0",
         mistake="Leaving Estimated Hours blank or entering 0 — the Planned vs Actual report shows 0 planned for that person. Estimation Accuracy KPI score = 0 pts. One missing field = 0 pts for that entire metric.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=130),

    dict(num="S3", timing="SPRINT/PHASE",
         action="Close the sprint on the LAST day of the sprint period\n1. Board → Sprint Manager\n2. Find the active sprint\n3. Click Close Sprint\n4. Confirm — unfinished items move to backlog\n5. Sprint Reliability is now finalised",
         where="Board → Sprint Manager\n→ Close Sprint",
         fields="• Close on the actual last day of the sprint\n• BEFORE closing, ensure all finished items are in QA_DONE status\n• After closing: unfinished items automatically move to backlog\n• Sprint Reliability score is now locked for this period",
         mistake="Closing sprint BEFORE ensuring all finished items are in QA_DONE — completed work is not counted. Alternatively: NOT closing the sprint at all — items from the next sprint bleed into KPI calculations for the wrong period.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log your own timesheet (every working day)\n1. Left sidebar → Timesheet\n2. Click + Add Entry (or + Log Hours)\n3. Select Work Item from dropdown\n4. Date: today\n5. Hours: actual hours worked on this item\n6. Description: 1-line note\n7. Save\n8. Repeat for each additional item worked on today",
         where="Left sidebar → Timesheet\n→ + Add Entry",
         fields="• Work Item: select from dropdown (must be assigned to you)\n• Date: today's date (do NOT backdate more than 2 days)\n• Hours: actual time (e.g. 3.5 for 3 hrs 30 mins)\n• Description: 'Reviewed PRs, updated API design doc'\n\nIf you worked on 3 items today:\n→ Create 3 separate entries (e.g. 3 hrs + 4 hrs + 1 hr = 8 hrs total)\n→ Do NOT log all 8 hrs against 1 item",
         mistake="Logging all 8 hours against a single item when you worked on multiple tasks — Planned vs Actual report shows 8 hrs actual vs 0 planned for items you actually touched but didn't log against.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=120),

    dict(num="D2", timing="DAILY",
         action="Monitor board — check team member status updates\n1. Board → Kanban view\n2. Scan each column for issues:\n   • Items in TODO with no activity for >2 days\n   • Items in BLOCKED (follow up immediately)\n   • Items in IN_PROGRESS for >3 days with no timesheet\n3. Message the team member to update",
         where="Board → Kanban View\n(switch using the icon top right)",
         fields="Expected item flow:\nTODO → IN_PROGRESS → IN_REVIEW → TESTING → QA_DONE\n\nRed flags to look for daily:\n• Item stuck in TODO when sprint started 3+ days ago\n• Item in IN_PROGRESS for >5 days (may be blocked)\n• Item in BLOCKED — action required from PM\n• Item in TESTING for >2 days with no QA timesheet",
         mistake="Ignoring BLOCKED items — if an item stays BLOCKED for >1 day without PM action, the team member's Dependency & Agile KPI is impacted AND the item won't be delivered on time.",
         ts="PARTIAL", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=110),

    "WEEKLY",
    dict(num="W1", timing="WEEKLY",
         action="Approve team timesheet entries (every Friday)\n1. Left sidebar → Timesheet\n2. Click Approval Queue tab\n3. Review each pending entry\n4. Click Approve (green) or Reject (red)\n5. If rejecting: add a reason in the rejection note\n6. Notify the team member to resubmit if rejected",
         where="Left sidebar → Timesheet\n→ Approval Queue tab",
         fields="What to check before approving:\n• Hours are reasonable for the work item description\n• Date is within the correct period\n• Work item is actually assigned to this person\n• Not duplicated (same item, same date, same hours as another entry)\n\nApprove: hours are correct → click Approve\nReject: wrong hours or duplicate → click Reject → write reason",
         mistake="Approving all entries without checking — if someone logs 16 hrs in a single day, approving it makes Capacity Report show them as 'over-occupied' and inflates actual hours in Planned vs Actual.",
         ts="FEEDS", pva="NO", cap="NO", kpi="NO", row_h=110),

    dict(num="W2", timing="WEEKLY",
         action="Approve or reject team leave requests within 24 hours\n1. Left sidebar → Leave Module\n2. Click Pending Requests\n3. Review each request (type, dates, reason)\n4. Approve or Reject\n5. If rejecting: write a brief reason",
         where="Left sidebar → Leave Module\n→ Pending Requests",
         fields="Leave types:\n• Casual Leave: pre-planned personal leave\n• Sick Leave: medical absence (may submit same day)\n• Annual Leave: planned vacation\n• Unpaid Leave: no paid leave balance remaining\n\nAfter Approving → person's leave shows as blue in Capacity Report\nAfter Rejecting → person's absence = unapproved = Attendance KPI = 0",
         mistake="Not approving within 24 hours — the employee's absence is treated as unapproved even if they submitted a request. Their Attendance KPI score = 0 pts, and Capacity shows them as 'available' on a day they're absent.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=100),

    "MONTHLY",
    dict(num="M1", timing="MONTHLY",
         action="Enter manual KPI scores for ALL team members\n(deadline: 5th of the NEXT month)\n1. Left sidebar → KPI Appraisal\n2. Click 'Enter Monthly Scores' button\n3. Select month\n4. For each team member, enter 3 scores:\n   a. Engineering Hygiene (0 / 3 / 5)\n   b. Reporting & Documentation (0 / 3 / 5)\n   c. Positive Behaviour (0 / 3 / 5)\n5. Add a note for any 0 score\n6. Save",
         where="Left sidebar → KPI Appraisal\n→ Enter Monthly Scores",
         fields="Score guide:\nEngineering Hygiene:\n  5 = clean code, no lint errors, security best practices always followed\n  3 = minor lint warnings, occasional issues flagged in review\n  0 = repeated violations, insecure code patterns, ignored feedback\n\nReporting & Documentation:\n  5 = all status updates on time, clear docs, KT notes when needed\n  3 = some delays in updates, formatting issues\n  0 = missed critical updates or no documentation provided\n\nPositive Behaviour:\n  5 = punctual, professional, zero late comings this month\n  3 = minor issues (1–2 late comings <10 mins)\n  0 = unprofessional conduct or >3 late comings",
         mistake="Missing the 5th-of-month deadline — default score for missing entries = 0. Each person loses up to 15 KPI points (3 metrics × 5 pts each). For a 5-person team that's up to 75 pts unaccounted for.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=140),

    dict(num="M2", timing="MONTHLY",
         action="Log your own Learning self-log\n1. Left sidebar → KPI (or Self-Log)\n2. Click + Add Learning\n3. Period: YYYY-MM (e.g. 2026-06)\n4. Topic: what you studied\n5. Hours: total hours this month\n6. Save",
         where="Left sidebar → KPI\n→ Self-Log → Learning\n→ + Add Learning",
         fields="• Period: 2026-06 (must match current month)\n• Topic: e.g. 'PMP Certification – Module 4'\n• Hours: total this month\n  ≥ 4 hrs = 5 KPI pts (full score)\n  ≥ 1 hr = 3 KPI pts\n  0 hrs = 0 KPI pts\n• Description: optional detail",
         mistake="Logging learning in the wrong period (e.g. entering 2026-05 in June) — the report filters by period so June's learning shows as 0 for June. Always match the period to the current month.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="M3", timing="MONTHLY",
         action="Log your own Innovation self-log\n1. Left sidebar → KPI → Self-Log → Innovation\n2. Click + Add Innovation\n3. Period: YYYY-MM\n4. Title: name of what you built\n5. Impact: measurable result\n6. Type: AI_IMPLEMENTATION or OTHER\n7. Save",
         where="Left sidebar → KPI\n→ Self-Log → Innovation\n→ + Add Innovation",
         fields="• Title: 'Risk predictor dashboard for sprint planning'\n• Impact: 'Reduced sprint planning time by 30%'\n• Type:\n  AI_IMPLEMENTATION = 5 pts (genuine AI/ML tool implemented)\n  OTHER = 3 pts (automation, scripts, process improvement)\n\nPM innovation examples:\n  AI: AI-based sprint risk prediction tool\n  OTHER: Automated sprint retrospective report template",
         mistake="Skipping innovation because 'I'm a PM, not a developer' — PMs can log process automation, AI planning tools, workflow improvements. Missing = 0 pts for a metric worth 5 pts.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    "ADMIN TRIGGER",
    dict(num="!!", timing="ADMIN TRIGGER",
         action="⚠ IMPACT SUMMARY: What breaks if PM skips each step",
         where="This is a reference row, not an action",
         fields="P1 (no project): team can't log timesheet → all reports = 0\nP3 (no holidays): Capacity shows all days available → wrong planning\nS1 (no sprint): Sprint Reliability = 0 for all team members\nS2 (no est. hrs): Planned vs Actual = 0 planned | Est. Accuracy = 0 pts for team\nS2 (no due date): Delivery Timeliness = 0 pts for team\nS3 (sprint not closed): KPI scores don't finalise for the period\nW2 (leave not approved): Attendance KPI = 0 for team member\nM1 (KPI scores not entered): up to 15 pts per person lost",
         mistake="These are all PM responsibilities — nobody else can fix them.",
         ts="", pva="", cap="", kpi="", row_h=120, skip=True),
]

build_sheet("1. Project Manager", C["pm_t"], C["pm_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "The PM's data entry is the FOUNDATION for the entire team's KPI and reports. Without correct "
    "sprint setup, work items, due dates, and estimated hours, the team's auto-calculated KPI scores "
    "(Sprint Reliability, Delivery Timeliness, Estimation Accuracy) will all show 0 regardless of "
    "actual team performance. The PM also controls the 3 manual KPI scores per team member each month.",
    pm_steps)


# ── SHEET 3: Team Lead ────────────────────────────────────────────────────────
tl_steps = [
    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Verify all items have KPI-critical fields before sprint starts\n1. Board → List View\n2. Filter by: Sprint = current sprint\n3. For each item assigned to your team:\n   • estimatedHours ≠ blank and ≠ 0\n   • dueDate is set within sprint\n   • storyPoints ≠ blank\n   • assigneeId = specific person\n4. Click any item to edit missing fields\n5. Confirm with PM that all items are ready",
         where="Board → List View\n→ Filter: Sprint = [active sprint]\n→ Open each item to verify/edit",
         fields="Fields to verify for each item:\n• Estimated Hours: 4=half day, 8=1 day, 16=2 days (NOT in story points, in HOURS)\n• Due Date: must be within sprint window\n• Story Points: 1–5 (affects Sprint Reliability weighting)\n• Assignee: named person (not 'Unassigned')\n• Sprint: linked to active sprint (not backlog)\n\nIf any field is missing: edit the item and fill it in — you don't need PM to do this",
         mistake="Assuming PM filled everything correctly — in practice, bulk-created items often miss estimated hours. TL verifying is the safety net. Missing hours → Estimation Accuracy = 0 for ALL team members.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=120),

    dict(num="S2", timing="SPRINT/PHASE",
         action="Create your own work items for TL-owned tasks\n1. Board → + New Work Item\n2. Type: TASK\n3. Assignee: yourself (Team Lead)\n4. Sprint: active sprint\n5. Set Estimated Hours (your work estimate)\n6. Set Due Date within sprint\n7. Story Points: 2–4\n8. Save",
         where="Board → + New Work Item",
         fields="TL task examples:\n• 'Code review for Sprint 5 PRs' — Est. 6 hrs, Due: end of sprint\n• 'Architecture review for auth module' — Est. 8 hrs\n• 'Pair programming session with Dev team' — Est. 4 hrs\n• 'Sprint planning facilitation' — Est. 2 hrs\n\nLog timesheet against these items daily.",
         mistake="Creating no work items for yourself as TL — then logging timesheet has no target. Planned vs Actual shows 0 planned for TL, timesheet appears as 'unstructured work'.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=90),

    dict(num="S3", timing="SPRINT/PHASE",
         action="Before sprint closes: verify ALL done items are in QA_DONE\n1. Board → Kanban View\n2. Filter by sprint (active)\n3. Check each column:\n   • IN_REVIEW: is this actually done? → move to TESTING or QA_DONE\n   • TESTING: has QA confirmed? → move to QA_DONE\n   • IN_PROGRESS: is it actually complete? → confirm with developer\n4. Any item that IS complete → move to QA_DONE before sprint closes\n5. Any item NOT complete → leave in current status",
         where="Board → Kanban View\n→ Filter by Sprint\n→ Drag cards or open item to change status",
         fields="Status meanings at sprint close:\n• QA_DONE = DELIVERED — counts in Sprint Reliability\n• Anything else = NOT DELIVERED — does not count\n\nFor IN_REVIEW items: if review is done and code is merged → move to QA_DONE\nFor TESTING items: check with QA if testing passed → if yes, move to QA_DONE\nFor IN_PROGRESS: 90% done but not complete → leave as IN_PROGRESS (do not fake QA_DONE)",
         mistake="Not doing this check — completed work stays in TESTING or IN_REVIEW at sprint close. Sprint Reliability shows lower than actual performance. The team gets penalised for work that WAS done.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=130),

    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log your own timesheet every working day\n1. Left sidebar → Timesheet → + Add Entry\n2. Work Item: select one of your assigned tasks\n3. Date: today\n4. Hours: actual hours (split if you worked on multiple items)\n5. Description: brief note\n6. Save",
         where="Left sidebar → Timesheet\n→ + Add Entry",
         fields="TL timesheet examples:\n• 8 hrs on 'Code review Sprint 5 PRs' (day spent entirely on reviews)\n• 4 hrs on 'Code review' + 4 hrs on 'Architecture design doc' (split day)\n• 3 hrs on 'Pair programming with Dev A' + 5 hrs on 'Auth module design'\n\nKey: split hours across multiple items if you switched tasks.\nNever log 0 hrs — even a meeting-heavy day should have a timesheet entry.",
         mistake="Logging timesheet once per week instead of daily — you forget what you actually worked on, hours are less accurate, and Capacity shows you as 'available' Mon–Thu even when you were busy.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=110),

    dict(num="D2", timing="DAILY",
         action="Check Timesheet Report for team members who haven't logged today\n1. Left sidebar → Reports → Timesheet\n2. Filter by: Project = your project, Period = this month\n3. Look for team members with very low hours vs expected\n4. Send a quick message to remind them to log",
         where="Left sidebar → Reports\n→ Timesheet\n→ Filter: Project + Period",
         fields="What to look for:\n• A developer showing < 30 hrs this month (5th of month or later) = likely not logging daily\n• A designer showing 0 hrs for 3+ consecutive days = not logging at all\n• Expected: each person logs ~8 hrs/day × working days in month\n\nAction: WhatsApp / Slack the person — 'Hey [Name], timesheet for Mon–Wed is missing, can you log?'",
         mistake="Assuming everyone is logging because they're busy — people forget. A single reminder from TL each morning keeps the data clean. Silent data gaps compound quickly.",
         ts="FEEDS", pva="PARTIAL", cap="PARTIAL", kpi="PARTIAL", row_h=100),

    "WEEKLY",
    dict(num="W1", timing="WEEKLY",
         action="Review Planned vs Actual for your team (every Friday)\n1. Reports → Planned vs Actual\n2. Filter by: Project + current month\n3. Look at each team member's variance status\n4. For OVER variance (red): update estimatedHours on remaining items OR flag to PM for reallocation\n5. For UNDER variance (blue): check if they're actually logging timesheet daily",
         where="Left sidebar → Reports\n→ Planned vs Actual\n→ Filter: Project + Period",
         fields="Variance status meanings:\n• ON TRACK (green): actual hours within 20% of planned — no action\n• OVER (red): actual > planned by >20% — re-estimate or reduce scope\n• UNDER (blue): actual < planned by >20%\n  → Either: they're ahead of schedule (good)\n  → Or: they forgot to log timesheet (check)\n\nAction for OVER: open work items → increase estimatedHours on remaining unfinished items\nAction for UNDER: message to confirm they're logging daily",
         mistake="Only looking at total hours, not per-item variance — an employee might be on-track overall but consistently over-estimating easy tasks and under-estimating complex ones. Look at individual item variance.",
         ts="NO", pva="FEEDS", cap="NO", kpi="PARTIAL", row_h=110),

    "SELF-LOG",
    dict(num="L1", timing="SELF-LOG",
         action="Log Learning activities (by last working day of each month)\n1. Left sidebar → KPI → Self-Log → Learning\n2. + Add Learning\n3. Period: 2026-06 (current month)\n4. Topic: what you studied\n5. Hours: total hours this month\n6. Save",
         where="Left sidebar → KPI\n→ Self-Log → Learning\n→ + Add Learning",
         fields="Hours scoring:\n≥ 4 hrs = 5 KPI pts | ≥ 1 hr = 3 pts | 0 hrs = 0 pts\n\nTL learning examples (any of these qualify):\n• Online course: 'AWS Certified Developer' on Udemy — 6 hrs this month\n• Internal workshop: 'Agile Best Practices' session — 2 hrs\n• Book: 'Clean Code' — tracking 4+ hrs of reading\n• Conference: attended DevConf 2026 — counts as 4+ hrs",
         mistake="Logging 0 hrs and skipping — 'I didn't learn anything formal this month.' Informal learning (YouTube tutorials, blog posts, articles) counts if you track the hours. Missing = 0 pts.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="L2", timing="SELF-LOG",
         action="Log Innovation work (by last working day of each month)\n1. Left sidebar → KPI → Self-Log → Innovation\n2. + Add Innovation\n3. Period: YYYY-MM\n4. Title + Impact + Type\n5. Save",
         where="Left sidebar → KPI\n→ Self-Log → Innovation\n→ + Add Innovation",
         fields="Type scoring:\nAI_IMPLEMENTATION = 5 pts | OTHER = 3 pts\n\nTL innovation examples:\nAI (5 pts):\n  • AI-powered PR review suggestion tool integrated into GitHub workflow\n  • Automated sprint capacity predictor using historical data\n\nOTHER (3 pts):\n  • Created reusable NestJS service template for team\n  • Automated PR size checker that flags PRs > 500 lines\n  • Sprint burndown tracker built in Google Sheets",
         mistake="Only logging AI items — if you built automation scripts or process improvements, they qualify as OTHER (3 pts). Don't miss 3 pts because you think only 'real AI' counts.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    dict(num="L3", timing="SELF-LOG",
         action="Submit leave requests before any absence\n1. Left sidebar → Leave Module\n2. + Submit Leave Request\n3. Type, dates, reason\n4. Submit (PM approves within 24 hrs)",
         where="Left sidebar → Leave Module\n→ + Submit Leave Request",
         fields="Leave types:\n• Casual / Sick / Annual / Unpaid\n\nScoring:\n  ≤ 1 day approved leave = 5 pts (Attendance KPI)\n  1–1.5 days = 3 pts\n  > 1.5 days or any unapproved = 0 pts\n\nFor sick leave: submit same morning via mobile browser before 10am",
         mistake="Going absent and NOT submitting a leave request even if you message the PM on WhatsApp — the system doesn't know you had permission. Attendance KPI = 0 regardless of whether PM was verbally OK with it.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=90),
]

build_sheet("2. Team Lead", C["tl_t"], C["tl_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "The TL is the bridge between PM setup and developer execution. They verify data quality "
    "at sprint start, ensure team members update statuses correctly, and catch data gaps early "
    "via weekly Planned vs Actual review. TLs now have access to Timesheet, Planned vs Actual, "
    "and Capacity reports — use them actively to monitor team data health, not just as post-mortem tools.",
    tl_steps)


# ── SHEET 4: Developer ────────────────────────────────────────────────────────
dev_steps = [
    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log timesheet hours for EVERY work item you touched today\n1. Left sidebar → Timesheet\n2. Click + Add Entry\n3. Work Item: select from dropdown (items assigned to you)\n4. Date: TODAY (do not backdate more than 2 days)\n5. Hours: actual hours spent on this specific item\n6. Description: one-line note\n7. Save\n8. Repeat for each additional item worked on\n9. Total daily hours: 6–9 hrs is normal",
         where="Left sidebar → Timesheet\n→ + Add Entry\n\nRepeat step for each\nwork item you touched",
         fields="✅ CORRECT way to log:\nWorked on 3 items today?\n→ Entry 1: 'Build login API' — 4 hrs — 'Wrote POST /auth/login endpoint'\n→ Entry 2: 'Fix auth bug #42' — 3 hrs — 'Traced null pointer in JWT decode'\n→ Entry 3: 'Code review PR #15' — 1 hr — 'Reviewed John's auth changes'\n\n❌ WRONG: Single entry: 'Login API' — 8 hrs\n(The 3 hrs on bug #42 don't show up at all)\n\nDescription tips:\n• 'Built X' = implemented something new\n• 'Fixed X' = bug fix\n• 'Reviewed X' = code review\n• 'Meeting: sprint planning' = meetings",
         mistake="Logging all hours against one item when you worked on several — the other items show 0 actual hours vs their estimated hours. Planned vs Actual report shows massive variance for items you DID work on but didn't log.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=150),

    dict(num="D2", timing="DAILY",
         action="Update work item status to match current progress\n1. Board → find your work item\n2. Drag card to the correct column OR open item → change Status field\n3. Use correct status for each stage\n4. If you're blocked: set BLOCKED immediately + add a comment",
         where="Board → Kanban View\n→ Drag card to column\n\nOR: Board → open item\n→ Status dropdown",
         fields="Status flow you must follow:\nTODO → (start coding) → IN_PROGRESS\nIN_PROGRESS → (raise PR) → IN_REVIEW\nIN_REVIEW → (PR approved/merged) → TESTING\nTESTING → (QA confirms) → QA_DONE\n\nBLOCKED: set this if you cannot continue due to:\n• Waiting for API from another team\n• Missing requirements/design\n• Dependency on another item not ready\n→ Add comment: 'Blocked waiting for Design approval on login screen'\n→ Tell your TL immediately",
         mistake="Leaving status as IN_PROGRESS when you're actually waiting for review — your item looks 'in progress' to PM and TL but no work is happening. Sprint Reliability is unaffected by this mistake but it misleads planning.",
         ts="NO", pva="PARTIAL", cap="PARTIAL", kpi="FEEDS", row_h=120),

    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Move ALL completed items to QA_DONE before sprint closes\n1. Board → Kanban View → filter by this sprint\n2. Look at every item assigned to you\n3. For items you've FINISHED:\n   • Status must be QA_DONE\n   • If it's in IN_REVIEW or TESTING: confirm with reviewer/QA and move to QA_DONE\n4. For items NOT finished: leave in current status\n5. Do this check on the last 2 days of the sprint\n6. Ask TL to confirm before sprint is closed",
         where="Board → Kanban View\n→ Filter by Sprint\n\nFor each finished item:\nDrag to QA_DONE column\nor open item → Status = QA_DONE",
         fields="Sprint Reliability formula:\n= (Story Points in QA_DONE) ÷ (Total Story Points Committed) × 15\n\nDelivery Timeliness formula:\n= (Items QA_DONE on or before Due Date) ÷ (Total Items with Due Date) × 15\n\nThroughput formula:\n= (Items QA_DONE) ÷ (Total Assigned Items) × 10\n\n= 40 KPI points depend on items reaching QA_DONE!\n\nTiming matters: QA_DONE before due date = full Delivery Timeliness pts\nQA_DONE after due date = 0 pts for that item's Delivery Timeliness",
         mistake="Assuming TL or PM will move items to QA_DONE for you — this is YOUR responsibility as the developer. Items in TESTING or IN_REVIEW at sprint close = NOT counted as delivered. Up to 40 KPI pts lost.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=130),

    dict(num="S2", timing="SPRINT/PHASE",
         action="NEVER reopen a completed (QA_DONE) work item for rework\nIf a bug is found on a completed item:\n1. Create a NEW work item (Type: BUG)\n2. Link it to the original in the description: 'Bug found in item #42'\n3. Do NOT change the original item's status back to IN_PROGRESS\n4. Assign the new bug to the responsible developer",
         where="Board → + New Work Item\n→ Type: BUG\n\nFill: Title, Description,\nResponsible User, Severity",
         fields="Why this matters:\n• Reopen count is tracked per item\n• Internal Rework Ratio = total reopens ÷ total completed items\n  0 reopens = 5 pts | ≤10% = 3 pts | >10% = 0 pts\n\nCorrect flow:\n• Item #42 moved to QA_DONE ✅\n• Bug found later\n• Create new BUG item: 'Login screen crashes on empty password'\n  → Responsible User: Dev who wrote the code\n  → Severity: MAJOR\n  → Parent reference: 'Related to item #42'",
         mistake="Moving a QA_DONE item back to IN_PROGRESS for rework — this increments reopenCount. If your team has >10% rework rate, everyone's Internal Rework Ratio = 0 pts.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=110),

    dict(num="S3", timing="SPRINT/PHASE",
         action="When a bug is created against your code, verify Responsible User is set correctly\n1. Open the BUG work item\n2. Check 'Responsible User' field\n3. If blank: set it to the developer who wrote the buggy code\n4. Check Severity is filled (BLOCKER / CRITICAL / MAJOR / MINOR / TRIVIAL)\n5. Check Classification is filled\n6. Save",
         where="Board → open BUG work item\n→ Responsible User field\n→ Severity dropdown\n→ Classification dropdown",
         fields="Severity guide:\n• BLOCKER: system unusable, affects all users\n• CRITICAL: major feature broken, no workaround\n• MAJOR: feature broken but workaround exists\n• MINOR: cosmetic issue, annoying but usable\n• TRIVIAL: typo, label wrong\n\nClassification:\n• NEW_BUG: fresh defect\n• REGRESSION: broke something that was working\n• UI_USABILITY: front-end/UX issue\n• SECURITY: vulnerability\n• PERFORMANCE: slow response\n• DATA: wrong data displayed\n\nDefect Leakage KPI:\n0 bugs = 10 pts | 1 MINOR = 7 pts | 2 MINOR = 4 pts\n1 CRITICAL/BLOCKER against you = 0 pts immediately",
         mistake="Leaving Responsible User blank — the bug has no owner, nobody's Defect Leakage KPI is affected, and the bug may never get fixed. Severity left blank = bug is ignored in the Bugs Report severity chart.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=120),

    "SELF-LOG",
    dict(num="L1", timing="SELF-LOG",
         action="Log Learning activities by last working day of the month\n1. Left sidebar → KPI → Self-Log → Learning\n2. + Add Learning\n3. Period: 2026-06\n4. Topic: what you studied\n5. Hours: total for this month\n6. Save",
         where="Left sidebar → KPI\n→ Self-Log → Learning",
         fields="Hours: ≥ 4 hrs = 5 pts | ≥ 1 hr = 3 pts | 0 = 0 pts\n\nDeveloper learning examples:\n• Udemy: 'React 18 Advanced Patterns' — 6 hrs this month\n• Book: 'Clean Code' by Robert Martin — 4 hrs\n• YouTube: system design videos — track hours in a notepad\n• Internal: attended team knowledge-sharing session — 2 hrs\n• Certification prep: AWS Developer Associate — 5 hrs\n• Reading docs: Next.js 14 migration guide — 1.5 hrs",
         mistake="Not tracking learning hours during the month and then guessing at the end — guessing leads to incorrect hours, which won't be accurate. Use a notepad, Notion, or calendar to track learning hours as you go.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="L2", timing="SELF-LOG",
         action="Log Innovation work by last working day of the month\n1. Left sidebar → KPI → Self-Log → Innovation\n2. + Add Innovation\n3. Period + Title + Impact + Type\n4. Save",
         where="Left sidebar → KPI\n→ Self-Log → Innovation",
         fields="Type scoring:\nAI_IMPLEMENTATION = 5 pts | OTHER = 3 pts\n\nDeveloper examples:\nAI (5 pts):\n  • Integrated GitHub Copilot with custom codebase context → saves 2 hrs/day\n  • Built AI-powered PR description generator\n  • Implemented AI code review bot for team repo\n\nOTHER (3 pts):\n  • Automated database seeding script for staging env\n  • Created custom ESLint plugin for team code standards\n  • Built test data factory for unit tests\n  • Automated deployment health-check script",
         mistake="Thinking 'automation scripts don't count as innovation' — they absolutely count as OTHER (3 pts). Any script/tool/process improvement that saves time counts. Only genuine AI/ML implementations get 5 pts.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    dict(num="L3", timing="SELF-LOG",
         action="Submit leave requests BEFORE any absence\n1. Left sidebar → Leave Module\n2. + Submit Leave Request\n3. Select Type (Casual/Sick/Annual/Unpaid)\n4. Start Date and End Date\n5. Reason\n6. Submit\n\nFor sick leave: submit same morning by 9:30 AM",
         where="Left sidebar → Leave Module\n→ + Submit Leave Request",
         fields="Attendance KPI scoring:\n≤ 1 approved leave day/month = 5 pts\n1–1.5 approved days = 3 pts\n> 1.5 days OR any unapproved absence = 0 pts\n\n'Approved' means: leave request SUBMITTED + PM/TL APPROVED in the system\n\nEven if your PM verbally says 'OK take the day off' — if you don't submit in the system AND PM doesn't approve in the system → it counts as unapproved absence = 0 pts",
         mistake="Telling PM on WhatsApp that you're sick but not submitting a leave request in PMS — verbal approval doesn't count. System treats it as absence. Attendance KPI = 0.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=100),
]

build_sheet("3. Developer", C["dev_t"], C["dev_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "The Developer's two critical actions are: (1) log timesheet DAILY against EVERY work item "
    "worked on, and (2) move items to QA_DONE when finished. These two actions alone account for "
    "up to 65 KPI points out of 100. Everything else (self-logs, leave requests) adds the remaining "
    "points. If both critical actions are done consistently, the KPI will largely take care of itself.",
    dev_steps)


# ── SHEET 5: QA Engineer ──────────────────────────────────────────────────────
qa_steps = [
    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Create QA test task items at sprint start\n1. Board → + New Work Item\n2. Type: TASK\n3. Title: 'QA Testing – [Feature Name]' (e.g. 'QA Testing – Login v2')\n4. Assignee: yourself (QA engineer)\n5. Sprint: active sprint\n6. Estimated Hours: time you expect testing to take\n7. Due Date: within sprint (typically 2–3 days before sprint close)\n8. Story Points: 2–3\n9. Save",
         where="Board → + New Work Item",
         fields="Create one QA task per feature/module being tested:\n• 'QA Testing – Login & Auth Module' — Est: 8 hrs, Due: Jun 20\n• 'QA Testing – Dashboard Filters' — Est: 4 hrs, Due: Jun 22\n• 'Regression Testing – Sprint 5' — Est: 6 hrs, Due: Jun 28\n\nWhy create these explicitly?\n→ Without your own work items, you can't log timesheet\n→ Without timesheet, your contribution is INVISIBLE in all reports\n→ Planned vs Actual shows 0 planned for QA even though you tested all day",
         mistake="Not creating QA work items — logging timesheet requires a work item to log against. Without items: QA shows 0 hours in Timesheet, Planned vs Actual, and Capacity reports. QA contribution is completely invisible.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=120),

    dict(num="S2", timing="SPRINT/PHASE",
         action="When testing begins on a feature item: move it to TESTING status\n1. Board → find the developer's work item (the one you're testing)\n2. Open item → Status = TESTING\n3. This signals to the developer that QA testing has started\n4. Begin your testing",
         where="Board → Kanban View\n→ Open developer's work item\n→ Status → TESTING",
         fields="Status flow for items going through QA:\nDeveloper moves item: IN_REVIEW → (QA picks it up) → TESTING\nQA runs tests:\n• If ALL tests pass → move to QA_DONE\n• If tests fail → move BACK to IN_PROGRESS + log a BUG item\n\nDo NOT skip TESTING status — moving from IN_REVIEW directly to QA_DONE means you've approved without testing.",
         mistake="Moving items from IN_REVIEW straight to QA_DONE without setting TESTING first — this makes it look like you never tested the feature. No accountability for the QA step in the flow.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=100),

    dict(num="S3", timing="SPRINT/PHASE",
         action="Log a BUG work item for every defect found during testing\n1. Board → + New Work Item\n2. Type: BUG\n3. Title: clear, specific (e.g. 'Login: null pointer when email field is empty')\n4. Fill ALL required bug fields (see below)\n5. Move the original item BACK to IN_PROGRESS\n6. Assign bug to the responsible developer\n7. Save\n8. Notify developer",
         where="Board → + New Work Item\n→ Type: BUG\n\nAlso: open original item\n→ Status = IN_PROGRESS",
         fields="MANDATORY bug fields — never leave blank:\n• Title: specific and reproducible (e.g. 'Dashboard: export button missing when no data')\n• Severity:\n  BLOCKER = app crash, cannot proceed\n  CRITICAL = major flow broken\n  MAJOR = feature broken with workaround\n  MINOR = cosmetic/UX issue\n  TRIVIAL = typo or label wrong\n• Classification: UI_USABILITY / PERFORMANCE / SECURITY / DATA / REGRESSION / NEW_BUG\n• Responsible User: developer who wrote the buggy code (NOT yourself)\n• Reporter: yourself\n• Steps to Reproduce: numbered steps so dev can reproduce\n• Expected vs Actual Result\n\nCRITICAL: Set Responsible User = the developer, NOT yourself as QA reporter.",
         mistake="Logging a bug without filling Responsible User, Severity, or Classification — bugs without these fields are invisible in the Bugs Report charts. The responsible developer's Defect Leakage KPI is unaffected. Bugs look like they never happened.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=140),

    dict(num="S4", timing="SPRINT/PHASE",
         action="Move item to QA_DONE when all tests pass\n1. Run full test suite for the feature\n2. Verify all acceptance criteria are met\n3. Check on multiple browsers/devices if relevant\n4. Open work item → Status = QA_DONE\n5. Add a comment: 'QA Passed – tested on Chrome, Firefox. All AC met.'\n6. Notify developer",
         where="Board → open work item\n→ Status → QA_DONE",
         fields="Before setting QA_DONE, verify:\n✅ All acceptance criteria in the item description are met\n✅ Happy path works correctly\n✅ Edge cases tested (empty inputs, long strings, wrong formats)\n✅ No console errors in browser DevTools\n✅ Tested on at least 2 browsers if front-end\n✅ API responses correct (check via DevTools Network tab)\n\nTiming: QA_DONE must be set ON OR BEFORE the item's Due Date for full Delivery Timeliness pts\nQA_DONE after due date = 0 pts for that item's Delivery Timeliness",
         mistake="Setting QA_DONE without fully testing all edge cases — a BLOCKER bug found post-delivery will cause the item to be reopened, increasing Internal Rework Ratio for the developer AND reflecting poorly on QA thoroughness.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=120),

    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log timesheet against your QA work items every day\n1. Left sidebar → Timesheet → + Add Entry\n2. Work Item: select your QA task (e.g. 'QA Testing – Login v2')\n3. Date: today\n4. Hours: actual hours spent testing today\n5. Description: brief note on what you tested\n6. Save\n7. If you tested multiple features, create a separate entry for each",
         where="Left sidebar → Timesheet\n→ + Add Entry",
         fields="QA timesheet examples:\n• 'QA Testing – Login v2' — 5 hrs — 'Tested all 12 test cases, found 2 bugs'\n• 'Regression Testing Sprint 5' — 3 hrs — 'Ran regression suite on dashboard'\n• 'Bug verification – Login bug #43' — 1 hr — 'Verified fix for null pointer bug'\n• 'Test case writing – API endpoints' — 2 hrs — 'Wrote 8 test cases for auth API'\n\nLog DAILY — even on days when all you did was re-test bugs (that's real work).",
         mistake="Not logging timesheet on days spent verifying bug fixes ('it was just an hour') — all those hours add up. QA's actual contribution becomes invisible in the Timesheet and Capacity reports.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=110),

    dict(num="D2", timing="DAILY",
         action="Update bug statuses as developer fixes are verified\n1. Receive notification that developer fixed a bug\n2. Re-test the fix in the latest build\n3. If fix is CORRECT: open BUG item → Status = QA_DONE (or CLOSED)\n4. If fix is INCORRECT: open BUG item → add comment 'Fix not working: [details]' → leave in TESTING",
         where="Board → find BUG work item\n→ Status update\n→ Add comment",
         fields="Bug lifecycle:\nNEW (reported) → ASSIGNED (developer assigned) → IN_PROGRESS (being fixed)\n→ IN_REVIEW (fix ready for QA) → TESTING (QA re-testing)\n→ QA_DONE or CLOSED (fix verified) OR back to IN_PROGRESS (not fixed)\n\nAlways add a comment when closing a bug:\n'Fix verified on Chrome + Firefox + mobile. QA approved Jun 20.'",
         mistake="Closing bugs without actually re-testing the fix — a 'verified' bug may still be open in production. This creates regression issues in the next release and reflects on QA credibility.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    "WEEKLY",
    dict(num="W1", timing="WEEKLY",
         action="Review Bugs Report to ensure all open bugs have correct metadata\n1. Reports → Bugs\n2. Look for bugs missing Severity or Classification\n3. Open each incomplete bug and fill the missing fields\n4. Escalate any BLOCKER or CRITICAL bugs that aren't being fixed fast enough",
         where="Left sidebar → Reports\n→ Bugs\n→ Filter by Project + Period",
         fields="Fields to check for all open bugs:\n✅ Severity: filled with correct level\n✅ Classification: filled\n✅ Responsible User: assigned to a developer\n✅ Status: matches actual state (not stuck in TESTING for 5+ days)\n✅ Reproducibility: ALWAYS / SOMETIMES / RARELY / CANNOT_REPRODUCE\n\nEscalation rule:\nBLOCKER or CRITICAL bug not fixed within 48 hrs → notify PM immediately\nMAJOR bug not fixed within sprint → plan for next sprint (move to backlog with PM approval)",
         mistake="Leaving bugs with missing metadata — the Bugs Report severity and classification charts become inaccurate. PM and TL use this report to prioritise fixes. Incomplete data leads to wrong prioritisation.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    "SELF-LOG",
    dict(num="L1", timing="SELF-LOG",
         action="Log Learning activities by last working day of the month\n1. Left sidebar → KPI → Self-Log → Learning\n2. + Add Learning\n3. Period: YYYY-MM | Topic | Hours\n4. Save",
         where="Left sidebar → KPI\n→ Self-Log → Learning",
         fields="Hours: ≥ 4 hrs = 5 pts | ≥ 1 hr = 3 pts\n\nQA learning examples:\n• ISTQB Foundation course — 5 hrs (Udemy)\n• Playwright/Cypress testing framework tutorial — 4 hrs\n• API testing with Postman/Newman — 3 hrs\n• Performance testing with k6 or JMeter — 4 hrs\n• Selenium WebDriver advanced guide — 3 hrs\n• Bug writing best practices article series — 1.5 hrs\n• Attended internal 'QA in Agile' workshop — 2 hrs",
         mistake="Thinking QA learning must be formal certification — any QA-related course, tutorial, or workshop qualifies. Track hours informally if needed.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="L2", timing="SELF-LOG",
         action="Log Innovation work — test automation, QA tools, process improvements\n1. Left sidebar → KPI → Self-Log → Innovation\n2. + Add Innovation\n3. Period | Title | Impact | Type\n4. Save",
         where="Left sidebar → KPI\n→ Self-Log → Innovation",
         fields="Type: AI_IMPLEMENTATION = 5 pts | OTHER = 3 pts\n\nQA innovation examples:\nAI (5 pts):\n  • AI-based test case generator from user stories\n  • Implemented AI-powered visual regression testing\n  • Built ML-based flaky test detector\n\nOTHER (3 pts):\n  • Automated regression test suite with Playwright (covers 50 test cases)\n  • Built test data seeder script for staging environment\n  • Created Postman collection for all API endpoints (60+ requests)\n  • Set up scheduled nightly regression run in CI/CD\n  • Created bug template that reduced 'invalid bugs' by 40%",
         mistake="QA has huge automation potential — any test automation script qualifies as OTHER (3 pts). Not logging it means missing easy KPI points for real work already done.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=110),

    dict(num="L3", timing="SELF-LOG",
         action="Submit leave requests before any absence",
         where="Left sidebar → Leave Module\n→ + Submit Leave Request",
         fields="Same as all roles:\n• Submit BEFORE absence (same morning for sick leave)\n• PM approves in system within 24 hrs\n• Attendance KPI: ≤1 day approved = 5 pts | 1–1.5 days = 3 pts | >1.5 or unapproved = 0 pts\n\nRemember: verbal approval from PM does NOT count. Must be submitted in PMS.",
         mistake="Taking leave with PM WhatsApp approval but no system submission — Attendance KPI = 0 even if PM said 'yes'.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=80),

    "ADMIN TRIGGER",
    dict(num="!!", timing="ADMIN TRIGGER",
         action="⚠ IMPACT SUMMARY: What breaks if QA skips data entry",
         where="Reference — not an action",
         fields="No QA work items created:\n→ QA can't log timesheet (no item to log against)\n→ QA contribution = 0 in ALL reports (Timesheet, Capacity, Planned vs Actual)\n\nNo timesheet logged:\n→ Capacity shows QA as 'available' every day\n→ Timesheet Report: QA row missing or shows 0\n→ Estimation Accuracy = 0 (no actual hours to compare)\n\nBugs logged without Responsible User/Severity:\n→ Developer's Defect Leakage KPI unaffected (no accountability)\n→ Bugs Report charts inaccurate\n\nNot moving items to QA_DONE:\n→ Sprint Reliability drops for developer (work done but not signed off)\n→ Delivery Timeliness = 0 for items QA holds in TESTING past due date",
         mistake="QA is a critical data entry role — their actions directly determine how developer KPI scores are calculated.",
         ts="", pva="", cap="", kpi="", row_h=120, skip=True),
]

build_sheet("4. QA Engineer", C["qa_t"], C["qa_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "QA's daily actions affect both their OWN KPI scores AND the Developer's Defect Leakage KPI. "
    "Without QA creating work items, their contribution is completely invisible in all reports. "
    "The two most important actions: (1) create QA test task items so you can log timesheet, "
    "and (2) always log bugs with Responsible User, Severity, and Classification filled — "
    "these fields are what make the Bugs Report and Defect Leakage KPI meaningful.",
    qa_steps)


# ── SHEET 6: Designer ─────────────────────────────────────────────────────────
des_steps = [
    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Verify your design tasks have KPI-critical fields filled\n1. Board → List View → filter by your name\n2. Check each item assigned to you:\n   • estimatedHours filled? (6 hrs = wireframe, 12 hrs = full screen, 16 hrs = prototype)\n   • dueDate set within sprint?\n   • storyPoints filled?\n3. If missing: ask PM or edit directly",
         where="Board → List View\n→ Filter: Assignee = You",
         fields="Design task estimated hours guide:\n• Low-fidelity wireframe (1 screen): 4–6 hrs\n• High-fidelity design (1 screen): 8–12 hrs\n• Design system component: 4–8 hrs\n• Prototype with interactions: 12–20 hrs\n• Design review and revisions: 4–6 hrs\n• Full mobile app screen set (5 screens): 30–40 hrs",
         mistake="Accepting 0 or blank Estimated Hours on your design tasks — your Estimation Accuracy KPI = 0 for that item. One missing field on your design tasks = 0 pts for the entire metric.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=100),

    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log timesheet against each design task worked on today\n1. Left sidebar → Timesheet → + Add Entry\n2. Work Item: select design task from dropdown\n3. Date: today\n4. Hours: actual hours spent\n5. Description: what design work you did\n6. Save\n7. Repeat for each additional design task",
         where="Left sidebar → Timesheet\n→ + Add Entry",
         fields="Designer timesheet examples:\n• 'Login Screen UI Design' — 5 hrs — 'Completed mobile-first wireframe, uploaded to Figma'\n• 'Design System – Button Component' — 3 hrs — 'Built 6 button variants with hover states'\n• 'Dashboard Mockup Review' — 2 hrs — 'Incorporated PM feedback on layout'\n• 'Logo exploration for client' — 4 hrs — 'Created 5 concept directions'\n\nLog design research, mood boarding, and feedback incorporation — it's all billable design work.",
         mistake="Not logging hours on 'non-design' design work like feedback sessions, revisions, and handoff prep — these hours are real work that affects your utilisation. Skipping them makes you appear less occupied than you are.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=110),

    dict(num="D2", timing="DAILY",
         action="Update design item status as work progresses\n1. Board → find your design item\n2. Update status to match current stage\n3. If waiting for stakeholder feedback: set BLOCKED with comment\n4. When fully approved: set QA_DONE SAME DAY as approval",
         where="Board → open design item\n→ Status dropdown",
         fields="Designer status flow:\nTODO → (start designing) → IN_PROGRESS\nIN_PROGRESS → (share for feedback) → IN_REVIEW\nIN_REVIEW → (incorporating feedback) → IN_PROGRESS (repeat if needed)\nIN_REVIEW → (PM/client approves final) → QA_DONE ← SET THIS SAME DAY!\n\nBLOCKED: use when:\n• Waiting for content/copy from client\n• Waiting for PM approval that's delayed >2 days\n• Waiting for brand guidelines not yet provided\n\nADD A COMMENT when setting BLOCKED:\n'Blocked: waiting for client logo files from PM. Requested June 18.'",
         mistake="Moving to QA_DONE days or weeks after the design was actually approved — completedAt = the day YOU change status. If approval was June 15 but you changed status June 22, Delivery Timeliness shows as late even though the work was done on time.",
         ts="NO", pva="PARTIAL", cap="PARTIAL", kpi="FEEDS", row_h=120),

    "WEEKLY",
    dict(num="W1", timing="WEEKLY",
         action="Ensure APPROVED designs are moved to QA_DONE on the day of approval\n1. When PM / TL / client approves your design (in meeting, Slack, email)\n2. Open the work item on the board IMMEDIATELY\n3. Set Status = QA_DONE\n4. Add comment: 'Design approved by [name] on [date]. Figma link: [link]'\n5. Move to QA_DONE on the SAME DAY as approval",
         where="Board → open design item\n→ Status = QA_DONE\n→ Add comment with approval details",
         fields="Timing rule:\ncompletedAt = the exact date+time you set QA_DONE in PMS\n\nDelivery Timeliness = (completedAt ≤ dueDate)\n• Approved Jun 15, Due Jun 15 → set QA_DONE Jun 15 → ✅ ON TIME\n• Approved Jun 15, Due Jun 15 → set QA_DONE Jun 20 → ❌ LATE (even though approval was on time)\n\nAlways keep Figma/design tool links in the item description for handoff.",
         mistake="Setting QA_DONE at end of week as a 'cleanup' task instead of on the day of approval — Delivery Timeliness KPI incorrectly shows you as late for work that was actually delivered on time.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=110),

    dict(num="W2", timing="WEEKLY",
         action="Log any UI/UX bugs you find during design review — attribute correctly\n1. Board → + New Work Item → Type: BUG\n2. Title: specific UI issue (e.g. 'Button text truncated on mobile < 375px')\n3. Severity: MINOR (most UI) or MAJOR (broken functionality)\n4. Classification: UI_USABILITY\n5. Responsible User: the DEVELOPER who implemented the UI (NOT yourself)\n6. Reporter: yourself\n7. Add screenshot or Figma reference",
         where="Board → + New Work Item\n→ Type: BUG",
         fields="CRITICAL: Responsible User = the developer, NOT you (the Designer)\n\nReason: Defect Leakage KPI is calculated against the Responsible User.\nIf you're the Reporter = you found the bug (correct)\nIf you're the Responsible User = you CAUSED the bug (incorrect for a designer reporting UI bugs)\n\nDesigner commonly reports:\n• Font doesn't match design spec → Responsible: Frontend Dev\n• Button spacing off by 8px → Responsible: Frontend Dev\n• Mobile layout broken → Responsible: Frontend Dev\n• Wrong colour hex used → Responsible: Frontend Dev",
         mistake="Setting yourself (Designer) as Responsible User for UI implementation bugs — your Defect Leakage KPI drops to 0 for a bug caused by a developer's implementation of your design. This is a common data entry error.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=120),

    "SELF-LOG",
    dict(num="L1", timing="SELF-LOG",
         action="Log Learning activities by last working day of the month",
         where="Left sidebar → KPI\n→ Self-Log → Learning",
         fields="Hours: ≥ 4 hrs = 5 pts | ≥ 1 hr = 3 pts\n\nDesigner learning examples:\n• Figma Advanced Components & Auto Layout course — 5 hrs\n• UX Research methods (Interaction Design Foundation) — 4 hrs\n• Accessibility: WCAG 2.2 guidelines study — 3 hrs\n• Design Tokens and Systems (Specify, Style Dictionary) — 4 hrs\n• Motion design with Framer or Lottie — 3 hrs\n• User psychology: 'Don't Make Me Think' (book) — 4 hrs\n• Design critique session led by senior designer — 2 hrs",
         mistake="Not tracking learning hours during the month. Keep a sticky note or Notion page tracking 'Learning Hours' weekly so you know your total at month end.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="L2", timing="SELF-LOG",
         action="Log Innovation — AI design tools, design systems, workflow improvements",
         where="Left sidebar → KPI\n→ Self-Log → Innovation",
         fields="Type: AI_IMPLEMENTATION = 5 pts | OTHER = 3 pts\n\nDesigner examples:\nAI (5 pts):\n  • Integrated Midjourney/DALL-E into concept creation workflow\n  • Used Galileo AI / Uizard for rapid wireframe generation\n  • Built AI-assisted colour palette generator for brand projects\n  • Implemented AI icon generator reducing search time by 60%\n\nOTHER (3 pts):\n  • Created comprehensive Figma component library (50+ components)\n  • Built design token system integrated with developer CSS variables\n  • Established design handoff checklist reducing revision cycles\n  • Created UI kit reused across 3 projects",
         mistake="Not logging design system work as innovation — building a Figma component library or design token system that the team reuses is legitimate innovation (OTHER = 3 pts).",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=100),

    dict(num="L3", timing="SELF-LOG",
         action="Submit leave requests before any absence",
         where="Left sidebar → Leave Module\n→ + Submit Leave Request",
         fields="Same rules as all roles:\n• Submit BEFORE absence or same morning for sick leave\n• PM approves within 24 hrs in the system\n• Attendance KPI: ≤1 day = 5 pts | 1–1.5 days = 3 pts | >1.5 days or unapproved = 0 pts",
         mistake="Verbal approval from PM doesn't count. Must be submitted in PMS AND approved in PMS.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=60),
]

build_sheet("5. Designer", C["des_t"], C["des_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "Designer's biggest KPI risk: setting QA_DONE status LATE after a design is approved. "
    "The system records completedAt as the moment you change the status — not the moment the "
    "client or PM said 'approved'. Always set QA_DONE on the SAME DAY as approval. "
    "The second biggest risk: attributing UI implementation bugs to yourself instead of the developer.",
    des_steps)


# ── SHEET 7: DevOps ───────────────────────────────────────────────────────────
dvo_steps = [
    "SPRINT/PHASE",
    dict(num="S1", timing="SPRINT/PHASE",
         action="Create work items for ALL DevOps tasks BEFORE starting work\n1. Board → + New Work Item\n2. Type: TASK\n3. Title: specific deliverable (not just 'DevOps work')\n4. Assignee: yourself\n5. Sprint: active sprint\n6. Estimated Hours: realistic time estimate\n7. Due Date: within sprint\n8. Story Points: 2–5 (infra tasks are complex)\n9. Save",
         where="Board → + New Work Item",
         fields="Create ONE item per deliverable — not one item for all DevOps work:\n✅ 'Set up CI/CD pipeline for staging environment' — Est: 12 hrs, Due: Jun 15\n✅ 'Configure SSL certificate on PMS production server' — Est: 4 hrs, Due: Jun 10\n✅ 'Docker containerise backend service' — Est: 8 hrs, Due: Jun 20\n✅ 'Set up monitoring alerts (CPU/Memory)' — Est: 6 hrs, Due: Jun 25\n✅ 'Database backup automation script' — Est: 5 hrs, Due: Jun 18\n\n❌ WRONG: Single item 'DevOps tasks Sprint 5' — too vague, all hours pile into one item",
         mistake="Creating ONE catch-all 'DevOps work' item for the entire sprint — you can't split hours meaningfully, Planned vs Actual shows one huge item instead of individual deliverables, and Sprint Reliability can't show which deliverables were completed on time.",
         ts="NO", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=130),

    dict(num="S2", timing="SPRINT/PHASE",
         action="Move completed DevOps tasks to QA_DONE when verified working\n1. Complete the infrastructure task\n2. Test that it works end-to-end\n3. Board → open the work item\n4. Status = QA_DONE\n5. Add comment with verification evidence\n6. Do this on the SAME DAY you verify it works",
         where="Board → open DevOps work item\n→ Status = QA_DONE\n→ Add comment",
         fields="Definition of DONE for DevOps tasks:\n• CI/CD pipeline: QA_DONE when first pipeline run is GREEN ✅\n• SSL configuration: QA_DONE when HTTPS tested via browser + curl\n• Docker containerisation: QA_DONE when container runs in target env\n• Database backup: QA_DONE when backup file created AND restore test passed\n• Monitoring alerts: QA_DONE when test alert fires and is received\n• Kubernetes deployment: QA_DONE when pods are healthy and service responds\n\nAdd a verification comment: 'SSL configured. Tested https://203.193.165.229 via Chrome and curl. Certificate valid until 2027-06-10.'",
         mistake="Forgetting to move DevOps items to QA_DONE after completion — 'the pipeline works, everyone can see it' — the system doesn't know it's done unless YOU change the status. Sprint Reliability, Delivery Timeliness, Throughput = 0.",
         ts="NO", pva="PARTIAL", cap="NO", kpi="FEEDS", row_h=120),

    "DAILY",
    dict(num="D1", timing="DAILY",
         action="Log timesheet hours against each DevOps work item worked on today\n1. Left sidebar → Timesheet → + Add Entry\n2. Work Item: select specific DevOps task\n3. Date: today\n4. Hours: actual hours\n5. Description: what specifically you worked on\n6. Save\n7. Repeat for each additional item",
         where="Left sidebar → Timesheet\n→ + Add Entry",
         fields="DevOps timesheet examples:\n• 'Set up CI/CD pipeline' — 6 hrs — 'Configured GitHub Actions workflow, fixed npm ci issue'\n• 'SSL certificate setup' — 3 hrs — 'Installed cert via nginx, tested HTTPS on prod server'\n• 'On-call incident response' — 2 hrs — 'Investigated high CPU alert, identified memory leak in auth service'\n• 'Docker containerisation' — 4 hrs — 'Created Dockerfile, optimised image size from 800MB to 120MB'\n\nFor on-call/incident work:\nCreate a 'Incident Response – [description]' work item first, then log time against it.",
         mistake="Not logging time for on-call or incident response because 'it wasn't planned work' — all work counts. Without a work item and timesheet entry, your incident response is invisible. Create an 'Incident Response' task item in advance.",
         ts="FEEDS", pva="FEEDS", cap="FEEDS", kpi="FEEDS", row_h=110),

    dict(num="D2", timing="DAILY",
         action="Update work item status as infrastructure tasks progress\n1. Board → open your DevOps item\n2. Move to IN_PROGRESS when you start\n3. Move to IN_REVIEW when config is ready for review/testing\n4. Move to TESTING when running smoke tests\n5. Move to QA_DONE when fully verified\n6. If blocked by access/approvals: set BLOCKED + comment",
         where="Board → open work item\n→ Status dropdown",
         fields="DevOps status flow:\nTODO → (start infra work) → IN_PROGRESS\nIN_PROGRESS → (config ready, asking for review) → IN_REVIEW\nIN_REVIEW → (running integration tests) → TESTING\nTESTING → (verified working end-to-end) → QA_DONE\n\nBLOCKED scenarios for DevOps:\n• Waiting for cloud account access (e.g. AWS IAM role not created yet)\n• Waiting for DNS changes to propagate\n• Waiting for PM approval to take production down for maintenance\n• Waiting for security review of firewall rule changes\n→ Set BLOCKED + comment explaining who needs to act",
         mistake="Leaving a completed infra task in TESTING for days while you're working on the next item — you forget to move it to QA_DONE, Sprint Reliability drops, and TL/PM think the task isn't done yet.",
         ts="NO", pva="PARTIAL", cap="PARTIAL", kpi="FEEDS", row_h=120),

    "SELF-LOG",
    dict(num="L1", timing="SELF-LOG",
         action="Log Innovation work by last working day of the month\n(BIGGEST KPI opportunity for DevOps)\n1. Left sidebar → KPI → Self-Log → Innovation\n2. + Add Innovation\n3. Period: YYYY-MM\n4. Title + Impact + Type\n5. Save",
         where="Left sidebar → KPI\n→ Self-Log → Innovation",
         fields="Type: AI_IMPLEMENTATION = 5 pts | OTHER = 3 pts\n\nDevOps AI (5 pts — any of these qualifies):\n  • AI-powered anomaly detection for server logs\n  • ML-based predictive auto-scaling\n  • AIOps tool for automated incident classification\n  • AI-assisted zero-downtime deployment system\n  • ML-based build time predictor\n\nDevOps OTHER (3 pts — most DevOps work qualifies!):\n  • Automated SSL certificate renewal script → 'Saves 2 hrs/month manual work'\n  • Multi-stage CI/CD pipeline with automated test gates → 'Reduced deploy failures by 70%'\n  • Infrastructure as Code (Terraform) for staging env → 'New env now spins up in 5 mins vs 2 hrs'\n  • Automated DB backup + restore verification → 'DR time reduced from 4 hrs to 30 mins'\n  • Kubernetes pod health monitoring → 'Zero manual restarts since setup'\n  • Docker image optimisation → 'Reduced build time from 12 mins to 3 mins'\n\n⚡ Almost ALL automation/pipeline work DevOps does counts here. Log it every month.",
         mistake="Not logging any innovation because 'the pipeline is just maintenance' — pipeline automation, IaC, backup scripts, monitoring setup: ALL count as OTHER (3 pts minimum). DevOps loses the easiest 3–5 pts in the KPI if this is skipped.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=160),

    dict(num="L2", timing="SELF-LOG",
         action="Log Learning activities by last working day of the month",
         where="Left sidebar → KPI\n→ Self-Log → Learning",
         fields="Hours: ≥ 4 hrs = 5 pts | ≥ 1 hr = 3 pts\n\nDevOps learning examples:\n• AWS Solutions Architect / CKA (Kubernetes) certification prep — 5 hrs\n• Terraform / Pulumi IaC deep-dive course — 4 hrs\n• GitOps with ArgoCD / Flux — 4 hrs\n• SRE book: 'Site Reliability Engineering' (Google) — 4 hrs\n• Container security: CIS Benchmarks for Docker/K8s — 3 hrs\n• Observability: Prometheus + Grafana setup tutorial — 3 hrs\n• Platform engineering trends (blog series) — 1.5 hrs",
         mistake="Doing a certification course but not logging the hours this month — must be logged in the same period month as when you studied. Log as you go, not retrospectively.",
         ts="NO", pva="NO", cap="NO", kpi="FEEDS", row_h=90),

    dict(num="L3", timing="SELF-LOG",
         action="Submit leave requests before any planned absence",
         where="Left sidebar → Leave Module\n→ + Submit Leave Request",
         fields="Same rules as all roles.\nNote for DevOps:\n• On-call weekend work = NOT a leave day (log as timesheet against 'On-Call' task)\n• Compensatory day off after on-call = submit as Casual Leave\n• Planned maintenance window at 2 AM = still a working day, log timesheet\n\nAttendance KPI: ≤1 day approved = 5 pts | 1–1.5 days = 3 pts | >1.5 or unapproved = 0 pts",
         mistake="Not submitting comp-off leave in the system after weekend on-call — taking comp-off without a leave request = unapproved absence = Attendance KPI = 0.",
         ts="NO", pva="NO", cap="FEEDS", kpi="FEEDS", row_h=80),

    "ADMIN TRIGGER",
    dict(num="!!", timing="ADMIN TRIGGER",
         action="⚠ IMPACT SUMMARY: What breaks if DevOps skips data entry",
         where="Reference — not an action",
         fields="No work items created for DevOps tasks:\n→ No timesheet logging target\n→ DevOps shows 0 in ALL reports (Timesheet, Capacity, Planned vs Actual, Productivity)\n→ Sprint Reliability = 0 (no items to show as delivered)\n\nNo daily timesheet logging:\n→ Estimation Accuracy = 0 (no actual hours to compare vs estimated)\n→ Capacity Report shows DevOps as 'available' every day even during heavy infra weeks\n→ PM has no data to plan future sprints around DevOps capacity\n\nNo innovation self-log:\n→ Automation & Innovation KPI = 0 (DevOps loses their easiest metric)\n→ Missing 3–5 pts for work that was already done\n\nNot moving tasks to QA_DONE:\n→ Sprint Reliability + Delivery Timeliness + Throughput = 0\n→ Up to 40 KPI points lost automatically",
         mistake="The sum of all skips: DevOps can appear to have 0 contribution in all reports despite being one of the busiest engineers on the team.",
         ts="", pva="", cap="", kpi="", row_h=120, skip=True),
]

build_sheet("6. DevOps", C["dvo_t"], C["dvo_b"],
    "Step-by-Step SOP — What to do, when, and exactly how",
    "DevOps has the biggest gap between actual work done and data visible in reports — because "
    "infra work often feels like 'background work' that doesn't need to be logged. It does. "
    "Every pipeline, deployment, and automation task needs a work item + timesheet entry. "
    "The Innovation self-log is the easiest high-value KPI action for DevOps: "
    "almost all automation work qualifies. Never skip it.",
    dvo_steps)


# ── SHEET 8: Quick Checklist ──────────────────────────────────────────────────
ws_ck = wb.create_sheet("7. Quick Checklist")
ws_ck.sheet_view.showGridLines = False
page_hdr(ws_ck, 1,
    "QUICK CHECKLIST — Print this and tick off daily / weekly / monthly",
    C["dark"], size=12)
info(ws_ck, 2,
    "Share this with ALL team members. Any unticked item = a report or KPI score will be inaccurate for that period.",
    "F8FAFC", C["dark"], h=28)

ck_h = ["✓", "When", "Action", "PM", "TL", "Dev", "QA", "Designer", "DevOps", "Report / KPI Impact"]
for ci, h in enumerate(ck_h, 1):
    cl = ws_ck.cell(row=3, column=ci, value=h)
    cl.fill = fill(C["dark"]); cl.font = fn(True, C["white"], 8)
    cl.alignment = al("center", "center"); cl.border = bdr()
ws_ck.row_dimensions[3].height = 22

checklist = [
    ("DAILY", None),
    ("☐", "DAILY", "Log timesheet hours against EVERY work item you touched today",
     "✔","✔","✔","✔","✔","✔","Timesheet · Planned vs Actual · Capacity · KPI Estimation Accuracy"),
    ("☐", "DAILY", "Update work item STATUS to current stage (IN_PROGRESS / IN_REVIEW / TESTING / QA_DONE / BLOCKED)",
     "—","✔","✔","✔","✔","✔","Sprint Reliability · Delivery Timeliness · KPI Dependency & Agile"),
    ("☐", "DAILY", "If BLOCKED: set status = BLOCKED and add comment with who needs to act",
     "—","✔","✔","✔","✔","✔","KPI Dependency & Agile (up to 5 pts)"),

    ("WEEKLY", None),
    ("☐", "WEEKLY", "PM & TL: Approve team timesheet entries (every Friday)",
     "✔","✔","—","—","—","—","Timesheet Report (entries show APPROVED)"),
    ("☐", "WEEKLY", "PM: Approve or reject leave requests within 24 hours",
     "✔","—","—","—","—","—","Capacity Report · KPI Attendance"),
    ("☐", "WEEKLY", "QA: Log bugs with Responsible User, Severity, and Classification filled",
     "—","—","—","✔","—","—","Bugs Report · KPI Defect Leakage for developers"),
    ("☐", "WEEKLY", "Designer: Move approved designs to QA_DONE SAME DAY as approval",
     "—","—","—","—","✔","—","KPI Delivery Timeliness"),
    ("☐", "WEEKLY", "TL: Review Planned vs Actual — fix over/under variance",
     "—","✔","—","—","—","—","KPI Estimation Accuracy for team"),

    ("SPRINT START", None),
    ("☐", "SPRINT START", "PM: Create sprint with Name, Goal, Start Date, End Date — then ACTIVATE on day 1",
     "✔","—","—","—","—","—","KPI Sprint Reliability for all team members"),
    ("☐", "SPRINT START", "PM: Create ALL work items with Estimated Hours, Due Date, Story Points, Assignee filled",
     "✔","—","—","—","—","—","Planned vs Actual · KPI Estimation Accuracy · Delivery Timeliness"),
    ("☐", "SPRINT START", "TL: Verify every team item has estimated hours and due date before sprint starts",
     "—","✔","—","—","—","—","KPI Estimation Accuracy for whole team"),
    ("☐", "SPRINT START", "Dev / QA / Designer / DevOps: Verify your items have estimated hours and due dates",
     "—","—","✔","✔","✔","✔","Your personal Estimation Accuracy and Delivery Timeliness KPI"),
    ("☐", "SPRINT START", "QA: Create QA test task items for each feature being tested this sprint",
     "—","—","—","✔","—","—","QA contribution visible in Timesheet · Planned vs Actual · Capacity"),
    ("☐", "SPRINT START", "DevOps: Create work items for ALL planned infra tasks before starting work",
     "—","—","—","—","—","✔","DevOps contribution visible in all reports"),

    ("SPRINT CLOSE", None),
    ("☐", "SPRINT CLOSE", "Dev / QA / Designer / DevOps: Move ALL finished items to QA_DONE (2 days before close)",
     "—","✔","✔","✔","✔","✔","KPI Sprint Reliability · Delivery Timeliness · Throughput (up to 40 pts)"),
    ("☐", "SPRINT CLOSE", "PM: Close the sprint on the last day",
     "✔","—","—","—","—","—","Finalises Sprint Reliability calculation for the period"),

    ("MONTHLY — By Last Working Day", None),
    ("☐", "MONTHLY", "Log Learning self-log for this month (≥4 hrs = 5 pts, ≥1 hr = 3 pts)",
     "✔","✔","✔","✔","✔","✔","KPI Learning Velocity (5 pts)"),
    ("☐", "MONTHLY", "Log Innovation self-log (AI_IMPLEMENTATION = 5 pts / OTHER = 3 pts)",
     "✔","✔","✔","✔","✔","✔","KPI Automation & Innovation (up to 5 pts)"),
    ("☐", "MONTHLY", "Submit leave requests for ALL absences this month (if not already submitted)",
     "✔","✔","✔","✔","✔","✔","KPI Attendance (5 pts) · Capacity Report"),

    ("BY 5th OF NEXT MONTH — PM / Admin Only", None),
    ("☐", "5TH OF MONTH", "Enter Engineering Hygiene score for EACH team member (0 / 3 / 5)",
     "✔","—","—","—","—","—","KPI Engineering Hygiene (5 pts per person)"),
    ("☐", "5TH OF MONTH", "Enter Reporting & Documentation score for EACH team member (0 / 3 / 5)",
     "✔","—","—","—","—","—","KPI Reporting & Documentation (5 pts per person)"),
    ("☐", "5TH OF MONTH", "Enter Positive Behaviour score for EACH team member (0 / 3 / 5)",
     "✔","—","—","—","—","—","KPI Positive Behaviour (5 pts per person)"),
    ("☐", "5TH OF MONTH", "Verify all KPI scores show on KPI Appraisal page — check for any 0s that shouldn't be 0",
     "✔","—","—","—","—","—","Confirms 100-pt KPI calculation is complete for all team members"),
]

ck_w = [5, 14, 44, 6, 6, 6, 6, 8, 8, 32]
for ci, w in enumerate(ck_w, 1):
    ws_ck.column_dimensions[get_column_letter(ci)].width = w

r = 4
for item in checklist:
    if len(item) == 2:
        _, _, hc = TIMING.get(item[0], (C["slate"], C["white"], C["slate"]))
        sec_hdr(ws_ck, r, item[0], hc); r += 1
    else:
        tick, when, action = item[0], item[1], item[2]
        roles = item[3:9]; impact = item[9]
        rb = C["row_a"] if r % 2 == 0 else C["row_b"]
        tbg, ttc, _ = TIMING.get(when, (rb, C["dark"], C["dark"]))
        cel(ws_ck, r, 1, tick,   C["daily_bg"], True,  C["daily_t"], "center", size=11)
        cel(ws_ck, r, 2, when,   tbg,           True,  ttc,          "center", size=8)
        cel(ws_ck, r, 3, action, rb,            True,  C["dark"],    size=9)
        role_colors = [(C["pm_b"], C["pm_t"]), (C["tl_b"], C["tl_t"]),
                       (C["dev_b"], C["dev_t"]), (C["qa_b"], C["qa_t"]),
                       (C["des_b"], C["des_t"]), (C["dvo_b"], C["dvo_t"])]
        for i, (rv, (rbg, rtc)) in enumerate(zip(roles, role_colors)):
            cel(ws_ck, r, 4+i, rv, rbg, True, rtc, "center", size=9)
        cel(ws_ck, r, 10, impact, C["kpi_b"], False, C["kpi_t"], size=8, italic=True)
        ws_ck.row_dimensions[r].height = 28
        r += 1

ws_ck.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Role-Wise Data Entry Procedure Guide.xlsx")
wb.save(out)
print(f"Saved -> {out}")
