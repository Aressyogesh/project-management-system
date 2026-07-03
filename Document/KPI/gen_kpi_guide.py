"""
Generates: KPI & Reports Data Entry Guide.xlsx
Saved to:  Document/KPI/KPI & Reports Data Entry Guide.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ───────────────────────────────────────────────────────────────────
C = {
    "dark":       "1E293B",
    "blue_h":     "1D4ED8",
    "green_h":    "15803D",
    "amber_h":    "92400E",
    "purple_h":   "5B21B6",
    "rose_h":     "9F1239",
    "auto_bg":    "DCFCE7",   # green-100  → auto-generated
    "auto_txt":   "14532D",
    "manual_bg":  "FEF3C7",   # amber-100  → manual entry
    "manual_txt": "78350F",
    "mix_bg":     "DBEAFE",   # blue-100   → mixed
    "mix_txt":    "1E3A8A",
    "row_a":      "F8FAFC",
    "row_b":      "FFFFFF",
    "cat1":       "EFF6FF",   # Delivery
    "cat2":       "F0FDF4",   # Quality
    "cat3":       "F5F3FF",   # Ownership
    "cat4":       "FFFBEB",   # Growth
    "cat5":       "FFF1F2",   # Behaviour
    "hdr_row":    "334155",
    "warn":       "FEE2E2",
    "border":     "CBD5E1",
    "thick_bdr":  "94A3B8",
    "white":      "FFFFFF",
    "section":    "475569",
}

CAT_COLORS = {
    "Delivery & Execution":              ("1D4ED8", C["cat1"]),
    "Quality & Engineering Excellence":  ("15803D", C["cat2"]),
    "Ownership & Collaboration":         ("5B21B6", C["cat3"]),
    "Growth & Innovation":               ("92400E", C["cat4"]),
    "Behaviour & Reliability":           ("9F1239", C["cat5"]),
}

def fill(h):    return PatternFill("solid", fgColor=h)
def fn(bold=False, color="1E293B", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(c=None):
    c = c or C["border"]
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bdr():
    s = Side(style="medium", color=C["thick_bdr"])
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, r, c, v="", bg=None, bold=False, color="1E293B", ha="left",
         size=10, italic=False, wrap=True, border=True):
    cl = ws.cell(row=r, column=c, value=v)
    if bg: cl.fill = fill(bg)
    cl.font = fn(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = al(ha, wrap=wrap)
    if border: cl.border = bdr()
    return cl

def hdr(ws, r, c1, c2, v, bg, tc="FFFFFF", size=11, row_h=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1, value=v)
    cl.fill = fill(bg)
    cl.font = fn(bold=True, color=tc, size=size)
    cl.alignment = al("center")
    cl.border = thick_bdr()
    if row_h: ws.row_dimensions[r].height = row_h
    return cl

def section_row(ws, r, ncols, label, bg=None):
    bg = bg or C["section"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=f"  {label}")
    cl.fill = fill(bg)
    cl.font = fn(bold=True, color="FFFFFF", size=10)
    cl.alignment = al("left")
    cl.border = thick_bdr()
    ws.row_dimensions[r].height = 18

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, ref="A3"):
    ws.freeze_panes = ref

def badge(v):
    """Return display text for auto/manual badge."""
    return {"AUTO": "✦ AUTO", "MANUAL": "✎ MANUAL", "SELF": "👤 SELF-LOG"}.get(v, v)

# =============================================================================
# SHEET 1  ▸  Quick Reference — All 13 Metrics
# =============================================================================
ws1 = wb.active
ws1.title = "1. KPI Quick Reference"
ws1.sheet_view.showGridLines = False

hdr(ws1, 1, 1, 13,
    "KPI APPRAISAL — QUICK REFERENCE  |  13 Metrics  |  100 Points Total  |  Digital Appraisal System",
    C["dark"], size=13, row_h=34)

col_hdrs = [
    "#", "Metric (exact name from appraisal sheet)",
    "Category", "Max\nPts", "Weightage",
    "How Points\nAre Generated",
    "Data Source\n(What feeds this metric)",
    "Auto ✦ or\nManual ✎",
    "Who Is\nResponsible",
    "Sprint\nProject?",
    "Non-Sprint\nProject?",
    "If Data Is\nMissing",
    "Points\nFormula"
]
for ci, h in enumerate(col_hdrs, 1):
    cell(ws1, 2, ci, h, bg=C["hdr_row"], bold=True, color="FFFFFF", ha="center", size=9)
ws1.row_dimensions[2].height = 32

METRICS = [
    # (id, name, category, pts, weight, how_generated, data_source, auto_manual,
    #  responsible, sprint, non_sprint, if_missing, formula)
    (1,
     "Sprint Reliability\n(Story Points Delivered vs Total Story Points Committed)",
     "Delivery & Execution", 15, "15%",
     "System counts story points on work items moved to QA_DONE in the sprint, divided by total committed story points",
     "Board → Work Items\n• sprintId (item must be in a sprint)\n• storyPoints (set at sprint planning)\n• status moved to QA_DONE",
     "AUTO",
     "Developer moves item to QA_DONE\nPM sets story points",
     "Uses sprint items only",
     "Falls back to ALL assigned items\n(each item = 1 pt if no story points)",
     "Scores 0 if:\n• No sprint assigned\n• No story points set\n• Items never moved to QA_DONE",
     "(Story Pts Delivered ÷ Story Pts Committed) × 15"),

    (2,
     "Delivery Timeliness\n(On-Time Tasks vs Total Assigned Tasks)",
     "Delivery & Execution", 15, "15%",
     "System compares completedAt date against dueDate for every assigned item. Items completed on or before due date count as on-time.",
     "Board → Work Items\n• assigneeId (assigned to employee)\n• dueDate (must be set)\n• completedAt (auto-set when moved to QA_DONE)",
     "AUTO",
     "PM sets due date\nDeveloper moves item to QA_DONE",
     "Same — uses all items with dueDate",
     "Same — uses all items with dueDate",
     "Scores 0 if:\n• No due date set\n• Item never moved to QA_DONE\n• (completedAt never populated)",
     "(On-Time Items ÷ Total Items with dueDate) × 15"),

    (3,
     "Estimation Accuracy\n(Actual Hours vs Total Estimated Hours)",
     "Delivery & Execution", 10, "10%",
     "System compares sum of estimatedHours on assigned items against actual hours logged in Timesheet for the same period.",
     "Board → Work Items\n• estimatedHours (set at creation)\nTimesheet Module\n• hours logged daily against work items",
     "AUTO",
     "PM/Dev sets estimated hours\nEmployee logs timesheet daily",
     "Uses estimated hrs from sprint items",
     "Uses estimated hrs from ALL assigned items",
     "Scores 0 if:\n• estimatedHours not filled\n• No timesheet hours logged",
     "Variance = |Actual − Estimated| ÷ Estimated\n≤15% = 10 pts | 16–30% = 7 pts\n31–50% = 4 pts | >50% = 0 pts"),

    (4,
     "Throughput & Complexity Handling\n(Valid PRs Merged vs Total PRs Generated)",
     "Delivery & Execution", 10, "10%",
     "System measures work item completion rate as a proxy for PR throughput. Items moved to QA_DONE vs total items assigned.",
     "Board → Work Items\n• assigneeId\n• status moved to QA_DONE\n• sprintId (for sprint projects)",
     "AUTO",
     "Developer moves items to QA_DONE",
     "Sprint items completed ÷ total sprint items",
     "ALL assigned items completed ÷ total assigned",
     "Scores 0 if:\n• No items assigned\n• Items never moved to QA_DONE",
     "(Items Completed ÷ Total Items) × 10"),

    (5,
     "Internal Rework Ratio\n(Reopened Tasks vs Total Completed Tasks)",
     "Quality & Engineering Excellence", 5, "5%",
     "System tracks how many times completed items are reopened (moved back from QA_DONE to earlier status). Auto-increments reopenCount.",
     "Board → Work Items\n• reopenCount (auto-incremented by system)\n• status history (QA_DONE → earlier status = reopen)",
     "AUTO",
     "System tracks automatically\n(no manual entry needed)",
     "Uses all assigned items",
     "Uses all assigned items",
     "Default = 5 pts if no items completed\nDegrades only when items are reopened",
     "0% reopens = 5 pts\n≤10% reopens = 3 pts\n>10% reopens = 0 pts"),

    (6,
     "Technical & Functional Defect Leakage\n(Production Bugs / Post-Release Bugs)",
     "Quality & Engineering Excellence", 10, "10%",
     "System counts BUG work items where this employee is the REPORTER. Severity determines the score impact.",
     "Board → Create Bug\n• type = BUG\n• severity (BLOCKER/CRITICAL/MAJOR/MINOR/TRIVIAL)\n• reporterId = this employee",
     "AUTO",
     "QA / Developer creates bug and sets severity",
     "BUG items reported by user in period",
     "Same — no difference",
     "Scores 10 pts if no bugs reported\nSeverity=null bugs are IGNORED",
     "0 bugs = 10 pts\n1 Minor = 7 pts | 2 Minor = 4 pts\n1 Critical or 3+ Minor = 0 pts\nBlocker/Critical present = 0 pts"),

    (7,
     "Engineering Hygiene\n(Best Practices, Security & Linting)",
     "Quality & Engineering Excellence", 5, "5%",
     "Admin observes code quality, security adherence, and linting standards during the month and enters a score manually.",
     "KPI Module\n• Admin enters score in 'Enter Monthly Scores'\n• metricId = engineering_hygiene\n• Score options: 5 / 3 / 0",
     "MANUAL",
     "Admin / Super User enters score\nby 5th of next month",
     "Same",
     "Same",
     "Defaults to 0 if admin hasn't entered\n(not 5 — no entry = 0)",
     "5 = Excellent (best practices, security, linting)\n3 = Minor gaps\n0 = Consistent violations"),

    (8,
     "Dependency & Agile Management\n(Daily Standup Logs / Blocker Communication)",
     "Ownership & Collaboration", 5, "5%",
     "System measures how proactively blockers are flagged. Items stuck in BLOCKED status signal uncommunicated delays.",
     "Board → Work Items\n• status = BLOCKED (when item is blocked)\n• Ratio of BLOCKED items to total assigned items",
     "AUTO",
     "Developer updates status to BLOCKED\nwhen genuinely blocked",
     "BLOCKED items ÷ total assigned items",
     "Same — no difference",
     "Scores 5 if no items assigned\nScore drops with high % of blocked items",
     "0% blocked = 5 pts\n≤15% blocked = 3 pts\n>15% blocked = 0 pts\n(Proactive flagging = fewer prolonged blocks)"),

    (9,
     "Reporting & Documentation\n(Status Reports, Technical Docs, KT Notes)",
     "Ownership & Collaboration", 5, "5%",
     "Admin reviews quality and timeliness of status updates, documentation and knowledge transfer notes during the month.",
     "KPI Module\n• Admin enters score in 'Enter Monthly Scores'\n• metricId = reporting_documentation\n• Score options: 5 / 3 / 0",
     "MANUAL",
     "Admin / Super User enters score\nby 5th of next month",
     "Same",
     "Same",
     "Defaults to 0 if admin hasn't entered",
     "5 = Accurate, no major gaps\n3 = Inconsistent / formatting issues / delays\n0 = Critical misses / incorrect info"),

    (10,
     "Learning Velocity\n(Professional Upskilling Path)",
     "Growth & Innovation", 5, "5%",
     "Employee self-logs learning activities (courses, certifications, articles, internal training) completed in the month.",
     "KPI → Self-Log → Learning\n• period (YYYY-MM)\n• topic (what was learned)\n• hours (time spent)\n• description (optional)",
     "SELF",
     "Employee logs own learning\nby last day of month",
     "Same — learning logs are per employee",
     "Same",
     "Scores 0 if no learning log entries\nfor the period",
     "≥4 hrs logged = 5 pts\n≥1 hr logged = 3 pts\n0 hrs = 0 pts"),

    (11,
     "Automation & Innovation\n(Tangible Improvements / AI Implementation)",
     "Growth & Innovation", 5, "5%",
     "Employee self-logs automation scripts, AI tools implemented, or process improvements delivered during the month.",
     "KPI → Self-Log → Innovation\n• period (YYYY-MM)\n• title (what was built/improved)\n• impact (description)\n• type: AI_IMPLEMENTATION or OTHER",
     "SELF",
     "Employee logs own innovation work\nby last day of month",
     "Same",
     "Same",
     "Scores 0 if no innovation log entries\nfor the period",
     "AI implementation = 5 pts\nDrafted process improvement = 3 pts\nZero inputs = 0 pts"),

    (12,
     "Attendance\n(No Unapproved Absences / Max 1 Leave per Month)",
     "Behaviour & Reliability", 5, "5%",
     "System reads approved leave requests for the period. Unapproved absences or excessive leave days reduce the score.",
     "Leave Module\n• Employee submits LeaveRequest\n• Manager approves it (status = APPROVED)\n• startDate, endDate, type",
     "AUTO",
     "Employee submits leave request\nManager approves within 24 hrs",
     "Same — leave requests per employee",
     "Same",
     "Unapproved absences = 0 pts\nNot submitting leave = treated as unapproved",
     "≤1 approved leave day = 5 pts\n1 to 1.5 approved leave days = 3 pts\n>1.5 approved OR any unapproved = 0 pts"),

    (13,
     "Positive Behaviour & Conduct\n(Professional, Flexible, Cooperative, Punctual)",
     "Behaviour & Reliability", 5, "5%",
     "Admin observes professionalism, teamwork, punctuality and conduct throughout the month and enters a score manually.",
     "KPI Module\n• Admin enters score in 'Enter Monthly Scores'\n• metricId = positive_behaviour\n• Score options: 5 / 3 / 0",
     "MANUAL",
     "Admin / Super User enters score\nby 5th of next month",
     "Same",
     "Same",
     "Defaults to 0 if admin hasn't entered",
     "5 = Professional, flexible, zero late comings\n3 = Professional but <3 late comings (<10 min)\n0 = Unprofessional or >3 late comings (>10 min)"),
]

for ri, m in enumerate(METRICS, 3):
    cat_tc, cat_bg = CAT_COLORS[m[2]]
    row_bg = C["row_a"] if ri % 2 == 0 else C["row_b"]
    am = m[7]
    am_bg = C["auto_bg"] if am == "AUTO" else C["manual_bg"] if am == "MANUAL" else C["mix_bg"]
    am_tc = C["auto_txt"] if am == "AUTO" else C["manual_txt"] if am == "MANUAL" else C["mix_txt"]

    cell(ws1, ri, 1,  m[0],  bg=row_bg, ha="center", bold=True, size=10)
    cell(ws1, ri, 2,  m[1],  bg=cat_bg, bold=True, size=9)
    cell(ws1, ri, 3,  m[2],  bg=cat_bg, color=cat_tc, italic=True, size=8)
    cell(ws1, ri, 4,  m[3],  bg=row_bg, ha="center", bold=True, color=C["blue_h"], size=11)
    cell(ws1, ri, 5,  m[4],  bg=row_bg, ha="center", size=9)
    cell(ws1, ri, 6,  m[5],  bg=row_bg, size=9)
    cell(ws1, ri, 7,  m[6],  bg=row_bg, size=9)
    cell(ws1, ri, 8,  badge(am), bg=am_bg, bold=True, color=am_tc, ha="center", size=9)
    cell(ws1, ri, 9,  m[8],  bg=row_bg, size=9)
    cell(ws1, ri, 10, m[9],  bg=C["cat1"], size=9)
    cell(ws1, ri, 11, m[10], bg=C["cat2"], size=9)
    cell(ws1, ri, 12, m[11], bg=C["warn"], size=9, italic=True)
    cell(ws1, ri, 13, m[12], bg=row_bg, size=9)
    ws1.row_dimensions[ri].height = 70

# Legend below table
lr = len(METRICS) + 4
hdr(ws1, lr, 1, 4, "Legend", C["dark"], row_h=20)
legends = [
    ("✦ AUTO",    C["auto_bg"],   C["auto_txt"],   "Score is calculated automatically from data entered in PMS modules (Board, Timesheet, Leave). No separate KPI entry needed."),
    ("✎ MANUAL",  C["manual_bg"], C["manual_txt"], "Score must be entered manually by Admin in the KPI → Enter Monthly Scores screen by the 5th of the following month."),
    ("👤 SELF-LOG", C["mix_bg"],  C["mix_txt"],    "Score is calculated automatically from data logged by the employee themselves in the KPI → Self-Log section (Learning / Innovation)."),
]
for i, (lbl, bg, tc, desc) in enumerate(legends, lr+1):
    cell(ws1, i, 1, lbl,  bg=bg, bold=True, color=tc, ha="center", size=10)
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    cl = ws1.cell(row=i, column=2, value=desc)
    cl.fill = fill(bg); cl.font = fn(color=tc, size=9)
    cl.alignment = al("left"); cl.border = bdr()
    ws1.row_dimensions[i].height = 20

col_widths(ws1, [4, 34, 26, 6, 9, 36, 36, 13, 22, 28, 28, 32, 36])
freeze(ws1, "A3")

# =============================================================================
# SHEET 2  ▸  Employee Action Guide (What YOU must do)
# =============================================================================
ws2 = wb.create_sheet("2. Employee Action Guide")
ws2.sheet_view.showGridLines = False

hdr(ws2, 1, 1, 9,
    "EMPLOYEE ACTION GUIDE  —  What YOU Must Do Each Month to Maximise Your KPI Score",
    C["blue_h"], size=13, row_h=34)

# Intro box
ws2.merge_cells("A2:I4")
intro = ws2["A2"]
intro.value = (
    "This guide tells every employee exactly what to do, when, and where in the system — "
    "so your KPI score reflects your actual work accurately.\n"
    "  ✦ AUTO = system calculates for you once you do the action   |   "
    "✎ MANUAL = admin enters on your behalf   |   "
    "👤 SELF-LOG = you enter it yourself in KPI → Self-Log"
)
intro.fill = fill(C["cat1"])
intro.font = fn(size=10, color=C["blue_h"])
intro.alignment = al("left", wrap=True)
intro.border = thick_bdr()
ws2.row_dimensions[2].height = 48

h2_cols = ["#", "What to Do", "Where in PMS", "When",
           "KPI Metric(s) Impacted", "Points\nAt Stake", "Type", "Sprint\nProject", "Non-Sprint\nProject"]
for ci, h in enumerate(h2_cols, 1):
    cell(ws2, 5, ci, h, bg=C["blue_h"], bold=True, color="FFFFFF", ha="center", size=9)
ws2.row_dimensions[5].height = 28

actions = [
    # (label, section)
    ("DAILY ACTIONS — Do these every working day", None),
    (1,  "Log your timesheet hours against the work item you worked on today",
         "Timesheet → Log Hours\n(select work item, enter hours, add short description)",
         "Every working day",
         "Estimation Accuracy (3) · Allocation · Planned vs Actual",
         "Up to 10 pts", "✦ AUTO", "Required", "Required (critical for T&M billing)"),

    ("WHEN YOU START WORKING ON A TASK", None),
    (2,  "Make sure your work item has a due date — ask PM if missing",
         "Board → Work Items\n(dueDate field on item)",
         "Before starting work",
         "Delivery Timeliness (2)",
         "Up to 15 pts", "✦ AUTO", "Required", "Required"),
    (3,  "Make sure estimated hours are filled on your work item — ask PM/TL if missing",
         "Board → Work Items\n(estimatedHours field)",
         "Before starting work",
         "Estimation Accuracy (3)",
         "Up to 10 pts", "✦ AUTO", "Required", "Required"),
    (4,  "For sprint projects: make sure your item is added to the active sprint with story points",
         "Board → Work Items\n(sprintId + storyPoints field)",
         "At sprint planning",
         "Sprint Reliability (1) · Throughput (4)",
         "Up to 25 pts", "✦ AUTO", "Required", "Not needed — auto-detects"),

    ("WHEN YOU COMPLETE A TASK", None),
    (5,  "Move work item status to QA_DONE when your work is done",
         "Board → Kanban board\n(drag card to QA_DONE column)",
         "When work is done",
         "Sprint Reliability (1) · Delivery Timeliness (2) · Throughput (4) · Rework Ratio (5)",
         "Up to 45 pts", "✦ AUTO", "Required", "Required"),
    (6,  "Do NOT reopen completed items unnecessarily — every reopen increments your rework count",
         "Board → Work Items\n(avoid moving QA_DONE items back)",
         "Always",
         "Internal Rework Ratio (5)",
         "Up to 5 pts", "✦ AUTO", "Required", "Required"),

    ("WHEN YOU ARE BLOCKED", None),
    (7,  "Immediately update work item status to BLOCKED — do not leave it IN_PROGRESS",
         "Board → Work Items\n(set status = BLOCKED)",
         "As soon as blocked",
         "Dependency & Agile Management (8)",
         "Up to 5 pts", "✦ AUTO", "Required", "Required"),
    (8,  "Add a comment explaining the blocker — keeps your manager informed",
         "Board → Work Item → Comments",
         "Same day as block",
         "Reporting & Documentation (9) — observed by admin",
         "Up to 5 pts", "✎ MANUAL", "Best practice", "Best practice"),

    ("WHEN LOGGING A BUG", None),
    (9,  "Always set severity when creating a BUG work item (never leave it blank)",
         "Board → Create Work Item\n(type = BUG, severity field is mandatory)",
         "At bug creation",
         "Defect Leakage (6)",
         "Up to 10 pts", "✦ AUTO", "Required", "Required"),
    (10, "Set bug classification (NEW_BUG / UI_USABILITY / PERFORMANCE / etc.)",
         "Board → Create Work Item\n(bugClassification field)",
         "At bug creation",
         "Bugs Report accuracy",
         "Report data", "✦ AUTO", "Required", "Required"),

    ("BEFORE / DURING ANY ABSENCE", None),
    (11, "Submit a leave request before every planned absence (sick leave too, ASAP)",
         "Leave Module → Submit Leave Request\n(type, start date, end date, reason)",
         "Before absence / same day",
         "Attendance (12)",
         "Up to 5 pts", "✦ AUTO", "Required", "Required"),

    ("BY END OF EVERY MONTH", None),
    (12, "Log all learning activities completed this month (courses, certifications, reading, training)",
         "KPI → Self-Log → Learning\n(period, topic, hours, description)",
         "By last working day",
         "Learning Velocity (10)",
         "Up to 5 pts", "👤 SELF-LOG", "Required", "Required"),
    (13, "Log any automation scripts, AI tools, or process improvements you delivered",
         "KPI → Self-Log → Innovation\n(period, title, impact, type)",
         "By last working day",
         "Automation & Innovation (11)",
         "Up to 5 pts", "👤 SELF-LOG", "Required", "Required"),

    ("SCORED BY ADMIN (you influence through your work — admin observes and scores)", None),
    (14, "Maintain code quality: follow linting rules, write secure code, follow best practices",
         "Your daily coding practice\n(Admin scores in KPI → Enter Monthly Scores)",
         "Throughout month",
         "Engineering Hygiene (7)",
         "Up to 5 pts", "✎ MANUAL", "Observed", "Observed"),
    (15, "Keep documentation updated: status reports, technical docs, KT notes",
         "Confluence / Docs / Comments\n(Admin scores in KPI → Enter Monthly Scores)",
         "Throughout month",
         "Reporting & Documentation (9)",
         "Up to 5 pts", "✎ MANUAL", "Observed", "Observed"),
    (16, "Be professional, punctual, flexible and cooperative",
         "Your workplace conduct\n(Admin scores in KPI → Enter Monthly Scores)",
         "Throughout month",
         "Positive Behaviour (13)",
         "Up to 5 pts", "✎ MANUAL", "Observed", "Observed"),
]

r = 6
for item in actions:
    if isinstance(item[0], str):
        section_row(ws2, r, 9, item[0], bg=C["blue_h"])
        r += 1
    else:
        row_bg = C["row_a"] if r % 2 == 0 else C["row_b"]
        am = item[7] if len(item) > 7 else ""
        am_bg = C["auto_bg"] if "AUTO" in am else C["manual_bg"] if "MANUAL" in am else C["mix_bg"]
        am_tc = C["auto_txt"] if "AUTO" in am else C["manual_txt"] if "MANUAL" in am else C["mix_txt"]
        cell(ws2, r, 1, item[0],  bg=row_bg, ha="center", bold=True, size=9)
        cell(ws2, r, 2, item[1],  bg=row_bg, size=9, bold=True)
        cell(ws2, r, 3, item[2],  bg=row_bg, size=9)
        cell(ws2, r, 4, item[3],  bg=row_bg, size=9)
        cell(ws2, r, 5, item[4],  bg=row_bg, size=9)
        cell(ws2, r, 6, item[5],  bg=row_bg, ha="center", bold=True, color=C["blue_h"], size=9)
        cell(ws2, r, 7, am,       bg=am_bg,  ha="center", bold=True, color=am_tc, size=9)
        cell(ws2, r, 8, item[7] if len(item)>7 else "", bg=C["cat1"], ha="center", size=9)
        cell(ws2, r, 9, item[8] if len(item)>8 else "", bg=C["cat2"], ha="center", size=9)
        ws2.row_dimensions[r].height = 44
        r += 1

col_widths(ws2, [4, 40, 32, 20, 36, 12, 13, 14, 20])
freeze(ws2, "A6")

# =============================================================================
# SHEET 3  ▸  Sprint vs Non-Sprint — Side-by-Side per Metric
# =============================================================================
ws3 = wb.create_sheet("3. Sprint vs Non-Sprint")
ws3.sheet_view.showGridLines = False

hdr(ws3, 1, 1, 10,
    "SPRINT vs NON-SPRINT — Exact Data Required per Metric  |  What Changes & What Stays the Same",
    C["dark"], size=13, row_h=34)

# column headers — two-row: metric info | sprint data | non-sprint data
ws3.merge_cells("A2:D2")
ws3.cell(row=2, column=1, value="METRIC").fill = fill(C["hdr_row"])
ws3.cell(row=2, column=1).font = fn(bold=True, color="FFFFFF", size=10)
ws3.cell(row=2, column=1).alignment = al("center")

ws3.merge_cells("E2:G2")
ws3.cell(row=2, column=5, value="  SPRINT-BASED PROJECTS  (Scrum / Agile)").fill = fill(C["blue_h"])
ws3.cell(row=2, column=5).font = fn(bold=True, color="FFFFFF", size=10)
ws3.cell(row=2, column=5).alignment = al("left")

ws3.merge_cells("H2:J2")
ws3.cell(row=2, column=8, value="  NON-SPRINT PROJECTS  (Kanban / T&M / Fixed / Dedicated)").fill = fill(C["green_h"])
ws3.cell(row=2, column=8).font = fn(bold=True, color="FFFFFF", size=10)
ws3.cell(row=2, column=8).alignment = al("left")
ws3.row_dimensions[2].height = 22

h3 = ["Metric", "Max Pts", "Type", "Key Rule",
      "Data You Must Enter\n(Sprint)", "Scoring Formula\n(Sprint)", "Score if Data Missing\n(Sprint)",
      "Data You Must Enter\n(Non-Sprint)", "Scoring Formula\n(Non-Sprint)", "Score if Data Missing\n(Non-Sprint)"]
for ci, h in enumerate(h3, 1):
    bg = C["blue_h"] if ci in (5,6,7) else C["green_h"] if ci in (8,9,10) else C["hdr_row"]
    cell(ws3, 3, ci, h, bg=bg, bold=True, color="FFFFFF", ha="center", size=9)
ws3.row_dimensions[3].height = 32

cmp = [
    ("Sprint Reliability", 15, "✦ AUTO",
     "Must move items to QA_DONE; story points amplify score",
     "1. Item added to sprint (sprintId)\n2. Story points set\n3. Status → QA_DONE",
     "(Story Pts QA_DONE ÷ Story Pts Committed) × 15\nHigher story points = more weight",
     "0 pts — no sprint items found",
     "1. Item assigned to you (assigneeId)\n2. Status → QA_DONE\n(No sprint or story points needed)",
     "(Items QA_DONE ÷ Total assigned items) × 15\nEach item counts equally (1 pt each)",
     "0 pts — no assigned items found"),

    ("Delivery Timeliness", 15, "✦ AUTO",
     "IDENTICAL for both — dueDate is critical",
     "1. dueDate set on item\n2. completedAt auto-set when moved to QA_DONE",
     "(Items where completedAt ≤ dueDate) ÷ Total items with dueDate × 15",
     "0 pts — dueDate or QA_DONE missing",
     "SAME — dueDate is even MORE critical\nsince there is no sprint end-date as fallback",
     "IDENTICAL formula",
     "0 pts — dueDate or QA_DONE missing"),

    ("Estimation Accuracy", 10, "✦ AUTO",
     "Requires BOTH estimatedHours AND timesheet hours",
     "1. estimatedHours on sprint items\n2. Log timesheet hours daily",
     "Variance of (estimated vs actual hours)\n≤15% = 10 | 16–30% = 7 | 31–50% = 4 | >50% = 0",
     "0 pts if estimatedHours=0 but hours logged",
     "1. estimatedHours on ALL assigned items\n2. Log timesheet hours daily",
     "Same variance formula\nbut uses ALL items (not just sprint items)",
     "0 pts if estimatedHours=0 but hours logged"),

    ("Throughput & Complexity", 10, "✦ AUTO",
     "Completion rate — move items to QA_DONE",
     "1. Items in sprint\n2. Status → QA_DONE",
     "(Sprint items QA_DONE ÷ Total sprint items) × 10",
     "0 pts — no sprint items",
     "1. Items assigned to you\n2. Status → QA_DONE",
     "(All QA_DONE ÷ All assigned items) × 10",
     "0 pts — no assigned items"),

    ("Internal Rework Ratio", 5, "✦ AUTO",
     "IDENTICAL — avoid reopening completed items",
     "No extra entry — system tracks reopenCount automatically",
     "0% reopens = 5 | ≤10% = 3 | >10% = 0",
     "5 pts (default when no completions yet)",
     "SAME",
     "IDENTICAL",
     "5 pts (default when no completions yet)"),

    ("Defect Leakage", 10, "✦ AUTO",
     "IDENTICAL — always set severity on bugs",
     "Create BUG item with:\n• severity (mandatory)\n• reporterId = you",
     "0 bugs = 10 | 1 minor = 7 | 2 minor = 4\n1 critical or 3+ minor = 0",
     "10 pts (no bugs = perfect score)",
     "SAME",
     "IDENTICAL",
     "10 pts (no bugs = perfect score)"),

    ("Engineering Hygiene", 5, "✎ MANUAL",
     "Admin scores — you influence through daily code quality",
     "Write clean, secure code following standards\n(Admin observes and enters 0/3/5)",
     "5 = Excellent | 3 = Minor gaps | 0 = Violations",
     "0 pts if admin hasn't entered score",
     "SAME — admin observes your code quality",
     "IDENTICAL",
     "0 pts if admin hasn't entered score"),

    ("Dependency & Agile Mgmt", 5, "✦ AUTO",
     "IDENTICAL — set BLOCKED status proactively",
     "When blocked: set status = BLOCKED immediately",
     "0% blocked = 5 | ≤15% = 3 | >15% = 0",
     "5 pts if no items (no data = max score)",
     "SAME",
     "IDENTICAL",
     "5 pts if no items assigned"),

    ("Reporting & Documentation", 5, "✎ MANUAL",
     "Admin scores — keep docs and status updates timely",
     "Update status comments, write docs, send KT notes\n(Admin observes and enters 0/3/5)",
     "5 = Accurate & timely | 3 = Inconsistent | 0 = Critical misses",
     "0 pts if admin hasn't entered score",
     "SAME",
     "IDENTICAL",
     "0 pts if admin hasn't entered score"),

    ("Learning Velocity", 5, "👤 SELF-LOG",
     "IDENTICAL — log learning in Self-Log panel",
     "KPI → Self-Log → Learning\n(topic + hours for this period)",
     "≥4 hrs = 5 | ≥1 hr = 3 | 0 hrs = 0",
     "0 pts if no self-log entries",
     "SAME",
     "IDENTICAL",
     "0 pts if no self-log entries"),

    ("Automation & Innovation", 5, "👤 SELF-LOG",
     "IDENTICAL — log innovation in Self-Log panel",
     "KPI → Self-Log → Innovation\n(title + impact + type for this period)",
     "AI implementation = 5 | Improvement drafted = 3 | None = 0",
     "0 pts if no self-log entries",
     "SAME",
     "IDENTICAL",
     "0 pts if no self-log entries"),

    ("Attendance", 5, "✦ AUTO",
     "IDENTICAL — submit leave requests, manager approves",
     "Leave Module → Submit request\nManager approves (status = APPROVED)",
     "≤1 approved leave = 5 | 1–1.5 = 3 | >1.5 or unapproved = 0",
     "0 pts if absence not submitted as leave request",
     "SAME",
     "IDENTICAL",
     "0 pts if absence not submitted as leave request"),

    ("Positive Behaviour", 5, "✎ MANUAL",
     "Admin scores — be professional and punctual daily",
     "Your daily conduct and punctuality\n(Admin observes and enters 0/3/5)",
     "5 = Professional, zero late | 3 = <3 minor lates | 0 = Unprofessional or >3 lates",
     "0 pts if admin hasn't entered score",
     "SAME",
     "IDENTICAL",
     "0 pts if admin hasn't entered score"),
]

for ri, row in enumerate(cmp, 4):
    cat_tc, cat_bg = CAT_COLORS.get(
        next((m[2] for m in METRICS if m[1].split("\n")[0].startswith(row[0].split(" (")[0][:12])), None)
        or "Delivery & Execution", ("1D4ED8", C["cat1"]))
    rb = C["row_a"] if ri % 2 == 0 else C["row_b"]
    am = row[2]
    am_bg = C["auto_bg"] if "AUTO" in am else C["manual_bg"] if "MANUAL" in am else C["mix_bg"]
    am_tc = C["auto_txt"] if "AUTO" in am else C["manual_txt"] if "MANUAL" in am else C["mix_txt"]

    cell(ws3, ri, 1,  row[0], bg=cat_bg, bold=True, size=9)
    cell(ws3, ri, 2,  row[1], bg=rb, ha="center", bold=True, color=C["blue_h"], size=11)
    cell(ws3, ri, 3,  am,     bg=am_bg, ha="center", bold=True, color=am_tc, size=9)
    cell(ws3, ri, 4,  row[3], bg=rb, size=9, italic=True)
    cell(ws3, ri, 5,  row[4], bg=C["cat1"], size=9)
    cell(ws3, ri, 6,  row[5], bg=C["cat1"], size=9)
    cell(ws3, ri, 7,  row[6], bg=C["warn"], size=9, italic=True)
    cell(ws3, ri, 8,  row[7], bg=C["cat2"], size=9)
    cell(ws3, ri, 9,  row[8], bg=C["cat2"], size=9)
    cell(ws3, ri, 10, row[9], bg=C["warn"], size=9, italic=True)
    ws3.row_dimensions[ri].height = 62

col_widths(ws3, [22, 7, 13, 28, 30, 30, 24, 28, 28, 24])
freeze(ws3, "A4")

# =============================================================================
# SHEET 4  ▸  Auto-Generated Data Map (which action → which score)
# =============================================================================
ws4 = wb.create_sheet("4. Auto-Generated Data Map")
ws4.sheet_view.showGridLines = False

hdr(ws4, 1, 1, 8,
    "AUTO-GENERATED DATA MAP  —  Which Action in PMS Feeds Which KPI Metric & Report",
    C["green_h"], size=13, row_h=34)

ws4.merge_cells("A2:H3")
box = ws4["A2"]
box.value = (
    "Every action below triggers an automatic score update — no separate KPI entry needed.\n"
    "The system reads this data when calculating your KPI for the selected month."
)
box.fill = fill(C["cat2"])
box.font = fn(size=10, color=C["green_h"])
box.alignment = al("left", wrap=True)
box.border = thick_bdr()
ws4.row_dimensions[2].height = 36

h4 = ["Action in PMS", "Module / Where", "DB Field Updated",
      "KPI Metrics Impacted", "Max Pts\nFrom This", "Reports Impacted",
      "Sprint\nProjects", "Non-Sprint\nProjects"]
for ci, h in enumerate(h4, 1):
    cell(ws4, 4, ci, h, bg=C["green_h"], bold=True, color="FFFFFF", ha="center", size=9)
ws4.row_dimensions[4].height = 28

auto_map = [
    ("BOARD — WORK ITEMS", None),
    ("Assign a work item to an employee",
     "Board → Work Items\n(assigneeId field)",
     "work_items.assigneeId",
     "Sprint Reliability · Delivery Timeliness · Estimation Accuracy · Throughput",
     "Up to 50 pts",
     "Productivity · Allocation · Planned vs Actual",
     "✔ Required", "✔ Required"),

    ("Add item to a sprint & set story points",
     "Board → Work Items\n(sprintId + storyPoints)",
     "work_items.sprintId\nwork_items.storyPoints",
     "Sprint Reliability (weighted by pts)\nThroughput",
     "Up to 25 pts",
     "Productivity",
     "✔ Required", "Not needed (auto-detects)"),

    ("Set estimated hours on work item",
     "Board → Work Items\n(estimatedHours field)",
     "work_items.estimatedHours",
     "Estimation Accuracy",
     "Up to 10 pts",
     "Planned vs Actual",
     "✔ Required", "✔ Required"),

    ("Set due date on work item",
     "Board → Work Items\n(dueDate field)",
     "work_items.dueDate",
     "Delivery Timeliness",
     "Up to 15 pts",
     "None (KPI only)",
     "✔ Required", "✔ Required (more critical)"),

    ("Move work item to QA_DONE",
     "Board → Kanban\n(drag to QA_DONE or change status)",
     "work_items.status = QA_DONE\nwork_items.completedAt = now() (auto)",
     "Sprint Reliability · Delivery Timeliness\nThroughput · Rework Ratio",
     "Up to 45 pts",
     "Productivity · Projects",
     "✔ Required", "✔ Required"),

    ("Set work item status to BLOCKED",
     "Board → Work Items\n(status = BLOCKED)",
     "work_items.status = BLOCKED",
     "Dependency & Agile Management",
     "Up to 5 pts",
     "None",
     "✔ Required", "✔ Required"),

    ("Reopen a completed work item (avoid!)",
     "Board → Move item from QA_DONE to earlier status",
     "work_items.reopenCount + 1 (auto)",
     "Internal Rework Ratio (reduces score)",
     "−5 pts risk",
     "None",
     "Tracked", "Tracked"),

    ("BOARD — BUGS", None),
    ("Create a BUG work item with severity set",
     "Board → Create Work Item\n(type=BUG, severity field)",
     "work_items.type = BUG\nwork_items.severity\nwork_items.reporterId",
     "Defect Leakage",
     "Up to 10 pts",
     "Bugs Report",
     "✔ Required", "✔ Required"),

    ("TIMESHEET", None),
    ("Log hours against work item daily",
     "Timesheet → Log Hours\n(workItem, date, hours)",
     "timesheet_entries.hours\ntimesheet_entries.date",
     "Estimation Accuracy",
     "Up to 10 pts",
     "Timesheet · Allocation · Planned vs Actual",
     "✔ Required", "✔ Required"),

    ("Manager approves timesheet entry",
     "Timesheet → Approve\n(status = APPROVED)",
     "timesheet_entries.approvalStatus = APPROVED",
     "None direct (affects report accuracy)",
     "Report data",
     "Timesheet Report (approved hrs only)",
     "Weekly", "Weekly"),

    ("LEAVE", None),
    ("Submit leave request",
     "Leave Module → New Request\n(type, dates, reason)",
     "leave_requests.userId\nleave_requests.startDate / endDate\nleave_requests.status = PENDING",
     "Attendance (pending → need approval)",
     "Prerequisite",
     "Capacity Report",
     "✔ Required", "✔ Required"),

    ("Manager approves leave request",
     "Leave Module → Approve\n(status = APPROVED)",
     "leave_requests.status = APPROVED",
     "Attendance",
     "Up to 5 pts",
     "Capacity Report",
     "✔ Required", "✔ Required"),

    ("SETTINGS (Admin — one-time)", None),
    ("Add public holidays",
     "Settings → Holidays\n(name, date)",
     "holidays.date",
     "Attendance (holidays don't count as leave)",
     "Attendance (12)",
     "Capacity Report",
     "✔ Required", "✔ Required"),

    ("Configure working days",
     "Settings → Portal Config\n(Mon–Fri = working)",
     "portal_config.workingDays",
     "Attendance · Capacity utilisation",
     "Capacity Report",
     "Capacity Report",
     "✔ Required", "✔ Required"),
]

r = 5
for item in auto_map:
    if isinstance(item[0], str) and item[1] is None:
        section_row(ws4, r, 8, item[0], bg=C["green_h"])
        r += 1
    else:
        rb = C["row_a"] if r % 2 == 0 else C["row_b"]
        is_warn = "avoid" in item[0].lower() or "risk" in str(item[4]).lower()
        row_bg = C["warn"] if is_warn else rb
        for ci, val in enumerate(item, 1):
            bg = C["cat1"] if ci == 7 else C["cat2"] if ci == 8 else row_bg
            cell(ws4, r, ci, val, bg=bg, size=9,
                 bold=(ci == 1), italic=(ci == 4 and is_warn))
        ws4.row_dimensions[r].height = 40
        r += 1

col_widths(ws4, [36, 28, 30, 34, 12, 28, 14, 16])
freeze(ws4, "A5")

# =============================================================================
# SHEET 5  ▸  Monthly Score Tracker (fillable template)
# =============================================================================
ws5 = wb.create_sheet("5. Monthly Score Tracker")
ws5.sheet_view.showGridLines = False

hdr(ws5, 1, 1, 9,
    "MONTHLY KPI SCORE TRACKER  —  Fill This Each Month to Monitor Your Progress",
    C["purple_h"], size=13, row_h=34)

ws5.merge_cells("A2:I3")
box2 = ws5["A2"]
box2.value = (
    "How to use: After each month ends, fill columns F–I to see your score for that month.\n"
    "Columns B–E show exactly what data drives each metric — check the system to see if your data was entered correctly."
)
box2.fill = fill(C["cat3"])
box2.font = fn(size=10, color=C["purple_h"])
box2.alignment = al("left", wrap=True)
box2.border = thick_bdr()
ws5.row_dimensions[2].height = 36

h5 = ["Metric", "Type", "Max\nPts",
      "Data That Generates This Score\n(check these in PMS)",
      "Action Required by YOU",
      "Did You\nComplete? (Y/N)",
      "Expected\nScore",
      "Actual Score\n(from KPI page)",
      "Gap /\nNotes"]
for ci, h in enumerate(h5, 1):
    cell(ws5, 4, ci, h, bg=C["purple_h"], bold=True, color="FFFFFF", ha="center", size=9)
ws5.row_dimensions[4].height = 32

tracker_rows = [
    ("Sprint Reliability (15 pts)", "✦ AUTO", 15,
     "• Work items in sprint with story points\n• Items moved to QA_DONE",
     "Add items to sprint → set story points → move to QA_DONE when done"),
    ("Delivery Timeliness (15 pts)", "✦ AUTO", 15,
     "• Due dates on all items\n• Items moved to QA_DONE before due date",
     "Ensure every item has a due date → complete and move to QA_DONE on time"),
    ("Estimation Accuracy (10 pts)", "✦ AUTO", 10,
     "• estimatedHours on items\n• Timesheet hours logged daily",
     "Log timesheet every day → ensure items have estimated hours"),
    ("Throughput & Complexity (10 pts)", "✦ AUTO", 10,
     "• Number of items assigned\n• Items moved to QA_DONE",
     "Move completed items to QA_DONE"),
    ("Internal Rework Ratio (5 pts)", "✦ AUTO", 5,
     "• How many times completed items are reopened\n• System auto-tracks",
     "Avoid reopening completed items unnecessarily"),
    ("Defect Leakage (10 pts)", "✦ AUTO", 10,
     "• BUG items you reported this month\n• Severity of those bugs",
     "When creating bugs: always set severity → fewer/lower bugs = better score"),
    ("Engineering Hygiene (5 pts)", "✎ MANUAL", 5,
     "• Admin observes your code quality\n• Admin enters score by 5th of next month",
     "Write clean, secure, well-linted code throughout the month"),
    ("Dependency & Agile Mgmt (5 pts)", "✦ AUTO", 5,
     "• Items in BLOCKED status\n• Ratio of blocked to total items",
     "Set BLOCKED status immediately when blocked → resolve quickly"),
    ("Reporting & Documentation (5 pts)", "✎ MANUAL", 5,
     "• Admin observes your reporting quality\n• Admin enters score by 5th of next month",
     "Keep status updated, write clear docs, send KT notes on time"),
    ("Learning Velocity (5 pts)", "👤 SELF-LOG", 5,
     "• Your self-log entries in KPI → Learning\n• Total hours logged for the month",
     "Log learning activities by end of month (≥4 hrs = full 5 pts)"),
    ("Automation & Innovation (5 pts)", "👤 SELF-LOG", 5,
     "• Your self-log entries in KPI → Innovation\n• Type: AI_IMPLEMENTATION or other",
     "Log any automation/AI work or process improvements by end of month"),
    ("Attendance (5 pts)", "✦ AUTO", 5,
     "• Approved leave requests for this month\n• Unapproved absences",
     "Submit leave request for every absence → ensure manager approves it"),
    ("Positive Behaviour (5 pts)", "✎ MANUAL", 5,
     "• Admin observes conduct and punctuality\n• Admin enters score by 5th of next month",
     "Be professional, punctual and cooperative throughout the month"),
]

for ri, row in enumerate(tracker_rows, 5):
    cat = None
    for cat_name, (tc, bg) in CAT_COLORS.items():
        names_in_cat = {
            "Delivery & Execution": ["Sprint Reliability", "Delivery Timeliness", "Estimation Accuracy", "Throughput"],
            "Quality & Engineering Excellence": ["Internal Rework", "Defect Leakage", "Engineering Hygiene"],
            "Ownership & Collaboration": ["Dependency", "Reporting"],
            "Growth & Innovation": ["Learning", "Automation"],
            "Behaviour & Reliability": ["Attendance", "Positive Behaviour"],
        }
        if any(n in row[0] for n in names_in_cat.get(cat_name, [])):
            cat = (tc, bg)
            break
    cat_tc, cat_bg = cat or ("1D4ED8", C["cat1"])
    rb = C["row_a"] if ri % 2 == 0 else C["row_b"]
    am = row[1]
    am_bg = C["auto_bg"] if "AUTO" in am else C["manual_bg"] if "MANUAL" in am else C["mix_bg"]
    am_tc = C["auto_txt"] if "AUTO" in am else C["manual_txt"] if "MANUAL" in am else C["mix_txt"]

    cell(ws5, ri, 1, row[0], bg=cat_bg, bold=True, size=9, color=cat_tc)
    cell(ws5, ri, 2, am,     bg=am_bg,  bold=True, color=am_tc, ha="center", size=9)
    cell(ws5, ri, 3, row[2], bg=rb, ha="center", bold=True, color=C["blue_h"], size=11)
    cell(ws5, ri, 4, row[3], bg=rb, size=9)
    cell(ws5, ri, 5, row[4], bg=rb, size=9)
    # Fillable columns
    cell(ws5, ri, 6, "", bg="FFFDE7", ha="center", size=10)
    cell(ws5, ri, 7, "", bg="FFFDE7", ha="center", size=10)
    cell(ws5, ri, 8, "", bg="E8F5E9", ha="center", size=10)
    cell(ws5, ri, 9, "", bg="FFF3E0", ha="center", size=9)
    ws5.row_dimensions[ri].height = 48

# Total row
tr = len(tracker_rows) + 5
ws5.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
cell(ws5, tr, 1,  "TOTAL", bg=C["dark"], bold=True, color="FFFFFF", ha="center", size=11)
cell(ws5, tr, 3,  100, bg=C["dark"], bold=True, color="FFFFFF", ha="center", size=12)
cell(ws5, tr, 4,  "", bg=C["dark"])
cell(ws5, tr, 5,  "", bg=C["dark"])
cell(ws5, tr, 6,  "", bg="FFFDE7", ha="center")
cell(ws5, tr, 7,  "", bg="FFFDE7", ha="center")
cell(ws5, tr, 8,  "", bg="E8F5E9", ha="center", bold=True, size=12, color=C["green_h"])
cell(ws5, tr, 9,  "", bg="FFF3E0")
ws5.row_dimensions[tr].height = 28

# Grade reference
gr = tr + 2
hdr(ws5, gr, 1, 4, "Grade Thresholds", C["dark"], row_h=20)
grades = [("A — Excellent", "90 – 100 pts", C["auto_bg"], C["auto_txt"]),
          ("B — Good",      "75 – 89 pts",  C["mix_bg"],  C["mix_txt"]),
          ("C — Average",   "60 – 74 pts",  C["manual_bg"],C["manual_txt"]),
          ("D — Poor",      "Below 60 pts", C["warn"],    C["rose_h"])]
for i, (g, r_str, bg, tc) in enumerate(grades, gr+1):
    cell(ws5, i, 1, g,     bg=bg, bold=True, color=tc, size=10)
    cell(ws5, i, 2, r_str, bg=bg, color=tc, size=10)
    cell(ws5, i, 3, "",    bg=bg)
    cell(ws5, i, 4, "",    bg=bg)
    ws5.row_dimensions[i].height = 18

col_widths(ws5, [28, 13, 7, 38, 40, 14, 14, 16, 20])
freeze(ws5, "A5")

# =============================================================================
# SHEET 6  ▸  Admin & Manager Monthly Checklist
# =============================================================================
ws6 = wb.create_sheet("6. Admin & Manager Checklist")
ws6.sheet_view.showGridLines = False

hdr(ws6, 1, 1, 7,
    "ADMIN & MANAGER MONTHLY CHECKLIST  —  Actions Required to Enable Accurate KPI & Reports",
    C["amber_h"], size=13, row_h=34)

h6 = ["#", "When", "Role", "Action", "Where in PMS", "KPI / Report Impact", "Done?"]
for ci, h in enumerate(h6, 1):
    cell(ws6, 2, ci, h, bg=C["amber_h"], bold=True, color="FFFFFF", ha="center", size=9)
ws6.row_dimensions[2].height = 24

checklist = [
    ("BEFORE THE MONTH STARTS", None),
    (1, "Before month / sprint starts", "PM / TL",
     "Create sprint — set name, goal, start date, end date",
     "Board → Sprint Manager", "Sprint Reliability · Throughput", "☐"),
    (2, "Sprint planning session", "PM / TL",
     "Add all work items to sprint — set story points on each",
     "Board → Work Items (sprintId + storyPoints)", "Sprint Reliability (weighted)", "☐"),
    (3, "Sprint planning session", "PM / TL",
     "Set estimated hours on every work item",
     "Board → Work Items (estimatedHours)", "Estimation Accuracy · Planned vs Actual", "☐"),
    (4, "Sprint planning session", "PM / TL",
     "Set due date on every work item",
     "Board → Work Items (dueDate)", "Delivery Timeliness", "☐"),
    (5, "Sprint planning session", "PM / TL",
     "Assign every work item to a team member",
     "Board → Work Items (assigneeId)", "All 4 delivery metrics", "☐"),
    (6, "Sprint start day", "PM / TL",
     "Activate the sprint",
     "Board → Sprint Manager → Activate", "Sprint Reliability", "☐"),

    ("DURING THE MONTH — WEEKLY", None),
    (7, "Every Friday", "PM / TL / Admin",
     "Review and approve team timesheet entries for the week",
     "Timesheet → Approval queue", "Timesheet Report accuracy", "☐"),
    (8, "Within 24 hrs of request", "PM / Admin",
     "Approve or reject employee leave requests",
     "Leave Module → Pending approvals", "Attendance (12) · Capacity Report", "☐"),
    (9, "Ongoing", "PM / TL",
     "Ensure blocked items are marked BLOCKED (not left as IN_PROGRESS)",
     "Board → Work Items", "Dependency & Agile Management (8)", "☐"),

    ("END OF MONTH", None),
    (10, "Last day of month", "PM / TL",
     "Close the sprint",
     "Board → Sprint Manager → Close", "Sprint Reliability finalised", "☐"),
    (11, "Last day of month", "PM / TL",
     "Ensure all completed items are in QA_DONE status",
     "Board → Review all columns", "Sprint Rel · Delivery · Throughput · Rework", "☐"),

    ("BY 5th OF NEXT MONTH — MANUAL KPI SCORES (Admin enters for each employee)", None),
    (12, "By 5th of next month", "Admin / Super User",
     "Score Engineering Hygiene for every employee\n(5 = excellent practices | 3 = minor gaps | 0 = violations)",
     "KPI Page → Enter Monthly Scores\n(metricId = engineering_hygiene)", "Engineering Hygiene (7) — 5 pts per employee", "☐"),
    (13, "By 5th of next month", "Admin / Super User",
     "Score Reporting & Documentation for every employee\n(5 = accurate, timely | 3 = inconsistent | 0 = critical misses)",
     "KPI Page → Enter Monthly Scores\n(metricId = reporting_documentation)", "Reporting & Docs (9) — 5 pts per employee", "☐"),
    (14, "By 5th of next month", "Admin / Super User",
     "Score Positive Behaviour for every employee\n(5 = professional, zero late | 3 = <3 minor lates | 0 = unprofessional)",
     "KPI Page → Enter Monthly Scores\n(metricId = positive_behaviour)", "Positive Behaviour (13) — 5 pts per employee", "☐"),
    (15, "By 5th of next month", "Admin / Super User",
     "Open KPI page — verify all 13 metric scores are populated for each employee",
     "KPI Page → Select employee → Check all metrics", "Confirms full 100-pt calculation", "☐"),

    ("ONCE A YEAR / AS NEEDED", None),
    (16, "Start of year", "Admin",
     "Add all public holidays for the year",
     "Settings → Holidays", "Attendance (12) · Capacity Report", "☐"),
    (17, "On new hire", "Admin",
     "Create user — assign department and shift",
     "Users Module", "All KPI metrics for that user", "☐"),
    (18, "On new project", "PM / Admin",
     "Create project — set type, dates, add all members with roles",
     "Projects Module", "All project-filtered reports", "☐"),
    (19, "One-time", "Admin",
     "Configure working days and business hours in Portal Config",
     "Settings → Portal Config", "Capacity Report · Utilisation %", "☐"),
]

r = 3
for item in checklist:
    if isinstance(item[0], str):
        section_row(ws6, r, 7, item[0], bg=C["amber_h"])
        r += 1
    else:
        rb = C["row_a"] if r % 2 == 0 else C["row_b"]
        cell(ws6, r, 1, item[0], bg=rb, ha="center", bold=True, size=9)
        cell(ws6, r, 2, item[1], bg=C["manual_bg"], bold=True, size=9, color=C["amber_h"])
        cell(ws6, r, 3, item[2], bg=rb, size=9)
        cell(ws6, r, 4, item[3], bg=rb, size=9, bold=True)
        cell(ws6, r, 5, item[4], bg=rb, size=9)
        cell(ws6, r, 6, item[5], bg=rb, size=9)
        cell(ws6, r, 7, item[6], bg=C["auto_bg"], ha="center", bold=True, size=12)
        ws6.row_dimensions[r].height = 40
        r += 1

col_widths(ws6, [4, 22, 16, 44, 36, 32, 8])
freeze(ws6, "A3")

# =============================================================================
# SHEET 7  ▸  Reports — What Feeds Each Tab
# =============================================================================
ws7 = wb.create_sheet("7. Reports Data Map")
ws7.sheet_view.showGridLines = False

hdr(ws7, 1, 1, 8,
    "REPORTS DATA MAP  —  What Data Each Report Tab Needs to Show Accurate Numbers",
    C["rose_h"], size=13, row_h=34)

h7 = ["Report Tab", "What It Shows", "Critical Data Required",
      "Who Enters", "How Often", "Sprint\nProject",
      "Non-Sprint\nProject", "Score if Data Missing"]
for ci, h in enumerate(h7, 1):
    cell(ws7, 2, ci, h, bg=C["rose_h"], bold=True, color="FFFFFF", ha="center", size=9)
ws7.row_dimensions[2].height = 28

reports = [
    ("Productivity",
     "Tasks completed, hours logged, on-time % per employee",
     "• work_items: assigneeId, status=QA_DONE, completedAt\n• timesheet_entries: hours, date\n• dueDate on items (for on-time %)",
     "Developer (board + timesheet)\nPM sets due dates",
     "Daily timesheet\nOngoing board updates",
     "Uses sprint items for completion rate",
     "Uses ALL assigned items",
     "Empty rows if no timesheet logged\nOn-time % = 0 if no due dates"),

    ("Projects",
     "Task count, completion %, team size per active project",
     "• projects: status=ACTIVE\n• work_items: status, createdAt in period\n• project members added",
     "PM (project setup + member add)\nDev (board status updates)",
     "Per board update",
     "Same",
     "Same",
     "Archived/On-Hold projects not shown\nWork items not in period = not counted"),

    ("Bugs",
     "Bug count by severity and classification",
     "• work_items: type=BUG, severity (mandatory!)\n• bugClassification\n• createdAt in period",
     "QA / Developer (bug creation)",
     "Per bug logged",
     "Same",
     "Same",
     "severity=null bugs excluded from charts\nEmpty report if no bugs logged"),

    ("Allocation",
     "Tasks allocated, hours logged, utilisation % per employee",
     "• work_items: assigneeId, startDate, dueDate\n• timesheet_entries: hours in period\n• Max 176 hrs/month assumed",
     "PM (assign items + set dates)\nDev (log timesheet)",
     "Per item + daily timesheet",
     "Same",
     "Same — even more critical",
     "0% utilisation if no timesheet hours\n0 tasks if no items assigned"),

    ("Timesheet",
     "Hours logged per employee per project",
     "• timesheet_entries: userId, hours, date, approvalStatus\n• project members (for project filter)",
     "Employee (log hours)\nManager (approve entries)",
     "Daily log\nWeekly approval",
     "Same",
     "Critical for T&M billing",
     "Report is empty if no hours logged\nUnapproved entries shown but flagged"),

    ("Capacity",
     "Daily availability grid per employee (holiday/leave/occupied/available)",
     "• holidays: configured in Settings\n• leave_requests: APPROVED by manager\n• timesheet_entries: daily hours\n• work_items: startDate/dueDate (assigned items)\n• portal_config: working days",
     "Admin (holidays + config)\nEmployee (leave request)\nManager (approve leave)\nDev (log timesheet)",
     "Yearly for holidays\nPer absence for leave\nDaily for timesheet",
     "Same",
     "Same",
     "Holidays missing = shows as 'available'\nLeave not approved = shows as 'available'\nNo timesheet = shows as 'available'"),

    ("Planned vs Actual",
     "Estimated vs actual hours per employee, variance %",
     "• work_items.estimatedHours (MANDATORY — must not be null)\n• timesheet_entries.hours for same period\n• assigneeId on items",
     "PM/Dev (set estimated hours)\nDev (log timesheet)",
     "Per item creation\nDaily timesheet",
     "Uses sprint items estimated hrs",
     "Uses ALL assigned items estimated hrs",
     "plannedHours = 0 if estimatedHours not set\nAll employees show as 'under-planned'"),
]

rep_bgs = [C["cat1"], C["cat2"], C["cat3"], C["cat4"], C["cat5"], C["warn"], C["mix_bg"]]

for ri, row in enumerate(reports, 3):
    rb = rep_bgs[(ri-3) % len(rep_bgs)]
    alt = C["row_a"] if ri % 2 == 0 else C["row_b"]
    cell(ws7, ri, 1, row[0], bg=rb, bold=True, size=10)
    cell(ws7, ri, 2, row[1], bg=alt, size=9, italic=True)
    cell(ws7, ri, 3, row[2], bg=alt, size=9)
    cell(ws7, ri, 4, row[3], bg=alt, size=9)
    cell(ws7, ri, 5, row[4], bg=alt, size=9)
    cell(ws7, ri, 6, row[5], bg=C["cat1"], size=9)
    cell(ws7, ri, 7, row[6], bg=C["cat2"], size=9)
    cell(ws7, ri, 8, row[7], bg=C["warn"], size=9, italic=True)
    ws7.row_dimensions[ri].height = 60

col_widths(ws7, [18, 28, 42, 26, 18, 22, 22, 34])
freeze(ws7, "A3")

# =============================================================================
# Save
# =============================================================================
import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "KPI & Reports Data Entry Guide.xlsx")
wb.save(out)
print(f"Saved -> {out}")
