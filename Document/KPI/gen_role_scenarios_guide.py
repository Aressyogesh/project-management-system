"""
Generates: Role-Based KPI & Reports Scenarios Guide.xlsx
Saved to:  Document/KPI/

Organised BY ROLE with concrete test scenarios so you can add
real data and immediately see the effect in KPI scores and Reports.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── Palette ───────────────────────────────────────────────────────────────────
C = {
    "dark":       "1E293B",
    "slate":      "334155",
    "blue":       "1D4ED8",
    "green":      "15803D",
    "amber":      "92400E",
    "purple":     "5B21B6",
    "rose":       "9F1239",
    "teal":       "0F766E",
    "indigo":     "3730A3",
    "row_a":      "F8FAFC",
    "row_b":      "FFFFFF",
    "hdr":        "1E293B",
    "grade_a_bg": "D1FAE5",   "grade_a_txt": "064E3B",
    "grade_b_bg": "DBEAFE",   "grade_b_txt": "1E3A8A",
    "grade_c_bg": "FEF3C7",   "grade_c_txt": "78350F",
    "grade_d_bg": "FEE2E2",   "grade_d_txt": "7F1D1D",
    "setup_bg":   "F0FDF4",   "setup_txt":   "14532D",
    "action_bg":  "EFF6FF",   "action_txt":  "1E3A8A",
    "result_bg":  "FEFCE8",   "result_txt":  "713F12",
    "report_bg":  "F5F3FF",   "report_txt":  "4C1D95",
    "warn":       "FEE2E2",
    "border":     "CBD5E1",
    "thick_bdr":  "94A3B8",
    "white":      "FFFFFF",
    "auto_bg":    "DCFCE7",   "auto_txt":   "14532D",
    "manual_bg":  "FEF3C7",   "manual_txt": "78350F",
    "self_bg":    "DBEAFE",   "self_txt":   "1E3A8A",
}

ROLE_COLORS = {
    "PROJECT_MANAGER": ("5B21B6", "F5F3FF"),
    "TEAM_LEAD":       ("0F766E", "F0FDFA"),
    "DEVELOPER":       ("1D4ED8", "EFF6FF"),
    "QA":              ("15803D", "F0FDF4"),
    "DESIGNER":        ("9F1239", "FFF1F2"),
    "DEVOPS":          ("92400E", "FFFBEB"),
}

def fill(h):    return PatternFill("solid", fgColor=h)
def fn(bold=False, color="1E293B", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(c=None):
    c = c or C["border"]
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bdr():
    s = Side(style="medium", color=C["thick_bdr"])
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, r, c, v="", bg=None, bold=False, color="1E293B", ha="left",
         size=10, italic=False, wrap=True, border=True):
    cl = ws.cell(row=r, column=c, value=v)
    if bg: cl.fill = fill(bg)
    cl.font = fn(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = al(ha, wrap=wrap)
    if border: cl.border = bdr()
    return cl

def page_hdr(ws, r, ncols, title, bg, tc="FFFFFF", size=13, h=36):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=title)
    cl.fill = fill(bg); cl.font = fn(bold=True, color=tc, size=size)
    cl.alignment = al("center"); cl.border = thick_bdr()
    ws.row_dimensions[r].height = h

def section(ws, r, ncols, label, bg=None, tc="FFFFFF"):
    bg = bg or C["slate"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(row=r, column=1, value=f"  {label}")
    cl.fill = fill(bg); cl.font = fn(bold=True, color=tc, size=10)
    cl.alignment = al("left"); cl.border = thick_bdr()
    ws.row_dimensions[r].height = 20

def info_box(ws, r, c1, c2, text, bg, tc, size=10, h=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1, value=text)
    cl.fill = fill(bg); cl.font = fn(color=tc, size=size)
    cl.alignment = al("left", wrap=True); cl.border = thick_bdr()
    if h: ws.row_dimensions[r].height = h

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, ref="A3"):
    ws.freeze_panes = ref

def score_badge(pts, max_pts):
    pct = (pts / max_pts) * 100 if max_pts else 0
    if pct >= 90:   return C["grade_a_bg"], C["grade_a_txt"], f"{pts}/{max_pts} pts — Grade A target"
    if pct >= 75:   return C["grade_b_bg"], C["grade_b_txt"], f"{pts}/{max_pts} pts — Grade B target"
    if pct >= 60:   return C["grade_c_bg"], C["grade_c_txt"], f"{pts}/{max_pts} pts — Grade C target"
    return C["grade_d_bg"], C["grade_d_txt"], f"{pts}/{max_pts} pts — Grade D"

# =============================================================================
# SHEET 1  ▸  Role Matrix — Overview of all roles
# =============================================================================
ws1 = wb.active
ws1.title = "1. Role Matrix"
ws1.sheet_view.showGridLines = False

page_hdr(ws1, 1, 9,
    "ROLE MATRIX — Who Does What, Which KPI Metrics Apply, and What Reports Show",
    C["dark"], size=13)

info_box(ws1, 2, 1, 9,
    "This sheet shows every project role in the PMS, their responsibilities, which KPI metrics "
    "apply to them, and which report tabs reflect their data. Use the role-specific scenario "
    "sheets (tabs 3–7) to see exactly what data to enter for each role.",
    "EFF6FF", C["blue"], h=40)

col_hdrs = ["Project Role", "System Role\n(typical)", "Primary\nResponsibility",
            "Key Actions in PMS",
            "KPI Metrics That Apply\n(and scoring type)",
            "Report Tabs\nThat Show Their Data",
            "Who Can View\nTheir KPI?",
            "Auto-Scored\nMetrics (9)",
            "Manual-Scored\nMetrics (3)"]
for ci, h in enumerate(col_hdrs, 1):
    cell(ws1, 3, ci, h, bg=C["hdr"], bold=True, color="FFFFFF", ha="center", size=9)
ws1.row_dimensions[3].height = 32

roles = [
    ("PROJECT MANAGER", "ADMIN or\nEMPLOYEE",
     "Plan sprints, assign work,\nmanage timelines,\napprove leave & timesheets",
     "• Create project & add members\n• Create sprint, set sprint goal\n• Create work items, assign, set due dates & story points\n• Approve leave requests\n• Approve timesheets\n• Enter monthly KPI scores for team",
     "All 13 metrics apply to PM individually.\nAs role: enables team metrics by entering\ndue dates, story points, estimated hours.\n• Sprint Reliability ✦\n• Delivery Timeliness ✦\n• Estimation Accuracy ✦\n• Throughput ✦\n• Rework Ratio ✦\n• Defect Leakage ✦\n• Engineering Hygiene ✎\n• Dependency & Agile ✦\n• Reporting & Docs ✎\n• Learning Velocity 👤\n• Innovation 👤\n• Attendance ✦\n• Positive Behaviour ✎",
     "All 7 report tabs\n(can filter by project)",
     "Super User\nAdmin",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),

    ("TEAM LEAD", "ADMIN or\nEMPLOYEE",
     "Technical oversight,\ncode reviews,\nmentoring team",
     "• Review & merge PRs\n• Set estimated hours on items\n• Verify completed items before QA_DONE\n• Log timesheet daily\n• Assist in sprint planning\n• Log learning & innovation",
     "All 13 metrics apply — same as Developer\nbut Admin can also view team KPIs.\n(Same list as PM above)",
     "All 7 report tabs",
     "Super User\nAdmin",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),

    ("DEVELOPER", "EMPLOYEE",
     "Build features,\nfix bugs,\nlog time daily",
     "• Move items from TODO → IN_PROGRESS → QA_DONE\n• Log timesheet hours daily against each work item\n• Set status to BLOCKED immediately if blocked\n• Log learning & innovation self-logs monthly\n• Submit leave requests before any absence",
     "All 13 metrics apply.\nSprint Reliability and Throughput depend\non moving items to QA_DONE.\n(Same list as PM above)",
     "Productivity, Allocation,\nTimesheet,\nPlanned vs Actual",
     "Super User\nAdmin\nTeam Lead\n(own record visible to self)",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),

    ("QA", "EMPLOYEE",
     "Test features,\nreport bugs,\nvalidate fixes",
     "• Test work items assigned to them\n• Create BUG work items with SEVERITY set (mandatory)\n• Set bugClassification on every bug\n• Move bugs to QA_DONE when verified/closed\n• Log timesheet daily\n• Log learning self-logs",
     "All 13 metrics apply.\nDefect Leakage is based on bugs where\nQA is set as 'Responsible User' (not reporter).\nQA creating bugs against DEVELOPERS\ndoes NOT penalise the QA's own score.\n(Same list as PM above)",
     "Productivity, Bugs Report,\nAllocation, Timesheet",
     "Super User\nAdmin\n(own record visible to self)",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),

    ("DESIGNER", "EMPLOYEE",
     "UI/UX design,\nprototypes,\ndesign handoffs",
     "• Move design tasks to QA_DONE when approved\n• Log timesheet against design work items\n• Log learning (design courses, tools)\n• Submit leave requests",
     "All 13 metrics apply.\nSprint Reliability applies if added to sprint.\nDefect Leakage: design bugs where\nDesigner is 'Responsible User'.\n(Same list as PM above)",
     "Productivity, Allocation,\nTimesheet",
     "Super User\nAdmin\n(own record visible to self)",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),

    ("DEVOPS", "EMPLOYEE",
     "CI/CD pipelines,\ndeployments,\ninfrastructure",
     "• Move DevOps tasks to QA_DONE\n• Log timesheet for infra/pipeline work\n• Log innovation (new automation, AI tools)\n• Submit leave requests",
     "All 13 metrics apply.\nAutomation & Innovation (11) is especially\nrelevant for DevOps — pipeline scripts,\nAI tools, deployment improvements.\n(Same list as PM above)",
     "Productivity, Allocation,\nTimesheet",
     "Super User\nAdmin\n(own record visible to self)",
     "Sprint, Delivery, Estimation,\nThroughput, Rework,\nDefect, Dependency, Attendance",
     "Engineering Hygiene,\nReporting & Docs,\nPositive Behaviour"),
]

ROLE_BG_MAP = {
    "PROJECT MANAGER": ("5B21B6", "F5F3FF"),
    "TEAM LEAD":       ("0F766E", "ECFDF5"),
    "DEVELOPER":       ("1D4ED8", "EFF6FF"),
    "QA":              ("15803D", "F0FDF4"),
    "DESIGNER":        ("9F1239", "FFF1F2"),
    "DEVOPS":          ("92400E", "FFFBEB"),
}

for ri, role in enumerate(roles, 4):
    tc, bg = ROLE_BG_MAP.get(role[0], (C["dark"], C["row_a"]))
    rb = C["row_a"] if ri % 2 == 0 else C["row_b"]
    cell(ws1, ri, 1, role[0],  bg=bg, bold=True, color=tc, size=10, ha="center")
    cell(ws1, ri, 2, role[1],  bg=rb, size=9, ha="center")
    cell(ws1, ri, 3, role[2],  bg=rb, size=9)
    cell(ws1, ri, 4, role[3],  bg=rb, size=9)
    cell(ws1, ri, 5, role[4],  bg=C["action_bg"], size=9, color=C["action_txt"])
    cell(ws1, ri, 6, role[5],  bg=C["report_bg"], size=9, color=C["report_txt"])
    cell(ws1, ri, 7, role[6],  bg=rb, size=9, ha="center")
    cell(ws1, ri, 8, role[7],  bg=C["auto_bg"], size=9, color=C["auto_txt"])
    cell(ws1, ri, 9, role[8],  bg=C["manual_bg"], size=9, color=C["manual_txt"])
    ws1.row_dimensions[ri].height = 100

# Note about defect leakage
nr = len(roles) + 5
info_box(ws1, nr, 1, 9,
    "⚠  IMPORTANT — Defect Leakage (Metric 6): Bugs are counted against the user set as "
    "'Responsible User' on the BUG work item — NOT the QA who created/reported it. "
    "This means: a Developer whose code had a bug gets the defect leakage score deduction. "
    "The QA who found and logged the bug is NOT penalised. "
    "Always set 'Responsible User' correctly when creating a BUG.",
    C["warn"], C["rose"], h=52)

col_widths(ws1, [20, 13, 22, 34, 42, 24, 18, 34, 24])
freeze(ws1, "A4")

# =============================================================================
# SHEET 2  ▸  Super Admin / Admin View — How to see all KPIs & Reports
# =============================================================================
ws2 = wb.create_sheet("2. Super Admin & Admin View")
ws2.sheet_view.showGridLines = False

page_hdr(ws2, 1, 7,
    "SUPER ADMIN & ADMIN VIEW — How to Access All KPIs, Reports and Enter Manual Scores",
    C["purple"], size=13)

info_box(ws2, 2, 1, 7,
    "Super User and Admin can view KPIs for every employee and access all 7 report tabs. "
    "They are also responsible for entering the 3 manual KPI scores each month. "
    "This sheet covers the exact navigation path for each task.",
    "F5F3FF", C["purple"], h=40)

# Column headers
cols2 = ["Task", "Where to Navigate",
         "What You'll See / Filter Options",
         "Filters Available", "Frequency", "Role Required", "Notes"]
for ci, h in enumerate(cols2, 1):
    cell(ws2, 3, ci, h, bg=C["purple"], bold=True, color="FFFFFF", ha="center", size=9)
ws2.row_dimensions[3].height = 24

tasks = [
    ("VIEW ALL EMPLOYEES' KPI SCORES", None),
    ("See the KPI Appraisal dashboard for all employees",
     "Left menu → KPI Appraisal",
     "Table shows all active employees with:\n• Total score (out of 100)\n• Grade (A / B / C / D)\n• Score breakdown per category\n• Team average score card at top",
     "• Period (month selector — default = current month)\n• Department filter\n• Employee search box",
     "Monthly review",
     "SUPER_USER or ADMIN",
     "Employee sees only their own KPI.\nAdmin/Super sees ALL employees."),

    ("View individual employee's full KPI breakdown",
     "KPI Appraisal → click employee row or use Employee dropdown",
     "Detailed breakdown of all 13 metrics:\n• Score bar for each metric\n• Points earned vs maximum\n• Radar chart of 5 categories",
     "• Select employee from dropdown\n• Select month/period",
     "On demand",
     "SUPER_USER or ADMIN",
     "Click the row to expand, or use 'Employee' dropdown at top-right to select a specific person."),

    ("ENTER MANUAL KPI SCORES (3 metrics per employee each month)", None),
    ("Enter Engineering Hygiene score for all employees",
     "KPI Appraisal → click 'Enter Monthly Scores' button (top right)",
     "A table with every employee and 3 score dropdowns:\n• Engineering Hygiene (0/3/5)\n• Reporting & Documentation (0/3/5)\n• Positive Behaviour (0/3/5)\nEnter all scores and click 'Save All Scores'",
     "• Period auto-set to selected month\n• All employees listed",
     "By 5th of every month\n(for previous month)",
     "SUPER_USER or ADMIN",
     "Scoring guide:\n5 = Excellent\n3 = Minor gaps / issues\n0 = Violations / critical misses\n\nDefault is 0 if not entered!"),

    ("Verify manual scores were saved",
     "KPI Appraisal → select employee → check individual breakdown",
     "Engineering Hygiene, Reporting & Docs, and Positive\nBehaviour should each show the entered value (0, 3 or 5)",
     "N/A",
     "After entering scores",
     "SUPER_USER or ADMIN",
     "If a metric shows 0 when you expected 5, the score was not saved — re-enter it."),

    ("VIEW ALL REPORTS", None),
    ("Productivity Report — who is completing tasks",
     "Left menu → Reports → Productivity tab",
     "One row per employee showing:\n• Tasks completed in period\n• Hours logged\n• On-time delivery %\n• Productivity score",
     "• Period (month)\n• Project filter",
     "Monthly",
     "SUPER_USER or ADMIN",
     "Empty rows = no timesheet hours logged OR no tasks completed. Check that employees are logging timesheet daily."),

    ("Projects Report — overall project health",
     "Left menu → Reports → Projects tab",
     "One row per ACTIVE project showing:\n• Total work items\n• Completed items\n• Team size\n• Completion %",
     "• Period\n• Project filter",
     "Weekly / Monthly",
     "SUPER_USER or ADMIN",
     "Archived projects are NOT shown. Only ACTIVE projects appear."),

    ("Bugs Report — bug severity and classification breakdown",
     "Left menu → Reports → Bugs tab",
     "Bar charts and tables showing:\n• Bug count by severity (BLOCKER/CRITICAL/MAJOR/MINOR/TRIVIAL)\n• Bug count by classification (NEW_BUG/UI_USABILITY/PERFORMANCE etc.)\n• CSV export available",
     "• Period\n• Project filter",
     "Weekly / Per release",
     "SUPER_USER or ADMIN",
     "⚠ Bugs without severity set are EXCLUDED from charts. Ensure QA always sets severity when creating bugs."),

    ("Allocation Report — who is over/under-allocated",
     "Left menu → Reports → Allocation tab",
     "One row per employee showing:\n• Tasks allocated\n• Hours logged\n• Utilisation % (vs 176 hrs/month max)",
     "• Period\n• Project filter",
     "Monthly",
     "SUPER_USER or ADMIN",
     "Utilisation over 100% = overloaded. Under 30% = underutilised. Check timesheet logging and task assignments."),

    ("Timesheet Report — hours logged and approval status",
     "Left menu → Reports → Timesheet tab",
     "One row per employee showing:\n• Total hours logged\n• Project they logged against\n• Approval status",
     "• Period\n• Project filter",
     "Weekly (approve)\nMonthly (review)",
     "SUPER_USER or ADMIN",
     "Unapproved entries appear separately. Approve weekly via Timesheet → Approval queue."),

    ("Capacity Report — availability calendar",
     "Left menu → Reports → Capacity tab",
     "Monthly calendar grid for every employee:\n• Green = available\n• Orange = partial (some hours)\n• Red = occupied (≥8 hrs)\n• Blue = on leave\n• Grey = holiday / weekend",
     "• Period (month)",
     "Monthly planning",
     "SUPER_USER or ADMIN",
     "If holidays are not configured in Settings → Holidays, those days show as available. Configure holidays at start of year."),

    ("Planned vs Actual — are estimates accurate?",
     "Left menu → Reports → Planned vs Actual tab",
     "One row per employee showing:\n• Planned hours (sum of estimatedHours)\n• Actual hours (sum of timesheet)\n• Variance % and status (over/under/on-track)",
     "• Period\n• Project filter",
     "Monthly",
     "SUPER_USER or ADMIN",
     "Only employees with estimatedHours > 0 OR actual hours > 0 appear. Empty report = no estimated hours set on work items."),

    ("KPI DASHBOARD FILTERS — How to drill down", None),
    ("Filter KPI by department",
     "KPI Appraisal → Department dropdown (top of table)",
     "Shows only employees in the selected department",
     "Department selector",
     "On demand",
     "SUPER_USER or ADMIN",
     "Useful when managing multiple teams — see only your team's KPI scores."),

    ("Filter KPI by period (historical months)",
     "KPI Appraisal → Period selector (top right)",
     "Shows KPI scores calculated for that specific month.\nOnly data created/updated IN that month is included.",
     "Up to 12 months back",
     "Historical review",
     "SUPER_USER or ADMIN",
     "KPIs are recalculated on each page load — historical data reflects the work items created/completed in that month."),

    ("Overview Dashboard — scoped project stats for Admin/PM", None),
    ("View project-scoped overview (PM / TL with project memberships)",
     "Left menu → Overview / Dashboard",
     "Stat cards show:\n• Active Projects (count of YOUR projects)\n• Total Tasks (across YOUR projects)\n• Team Members (distinct members in YOUR projects)\n• Completed Tasks (in YOUR projects)\n\nActivity chart and announcements also scoped to your projects.",
     "No filter needed — auto-scoped\nby project membership",
     "Daily",
     "Any user with project memberships",
     "SUPER_USER with NO project memberships\nsees global system-wide stats.\nEMPLOYEE without PM/TL role sees personal\ntask stats (My Projects, My Tasks)."),
]

r = 4
for item in tasks:
    if isinstance(item[0], str) and item[1] is None:
        section(ws2, r, 7, item[0], bg=C["purple"])
        r += 1
    else:
        rb = C["row_a"] if r % 2 == 0 else C["row_b"]
        cell(ws2, r, 1, item[0],  bg=rb, bold=True, size=9)
        cell(ws2, r, 2, item[1],  bg=C["action_bg"], size=9, color=C["action_txt"])
        cell(ws2, r, 3, item[2],  bg=rb, size=9)
        cell(ws2, r, 4, item[3],  bg=C["setup_bg"], size=9, color=C["setup_txt"])
        cell(ws2, r, 5, item[4],  bg=rb, size=9)
        cell(ws2, r, 6, item[5],  bg=C["manual_bg"], size=9, color=C["manual_txt"], ha="center")
        cell(ws2, r, 7, item[6],  bg=C["warn"] if "⚠" in (item[6] or "") else rb, size=9, italic=True)
        ws2.row_dimensions[r].height = 52
        r += 1

col_widths(ws2, [34, 30, 40, 26, 16, 18, 36])
freeze(ws2, "A4")

# =============================================================================
# Helper — build a role scenario sheet
# =============================================================================
def build_role_sheet(wb, tab_title, role_name, role_color, scenarios):
    """
    scenarios: list of dicts with keys:
      title, grade_target, setup, actions, kpi_breakdown, report_effect, notes
    """
    ws = wb.create_sheet(tab_title)
    ws.sheet_view.showGridLines = False
    tc, bg = ROLE_BG_MAP.get(role_name, (C["dark"], C["row_a"]))

    page_hdr(ws, 1, 7,
        f"{role_name}  —  Role-Based KPI & Report Scenarios  |  Detailed Data Entry Guide",
        tc, size=12)

    intro = (
        f"Each scenario below shows the EXACT data to enter in PMS to achieve the target grade "
        f"for a {role_name.title().replace('_', ' ')}. Follow the SETUP steps first (done by Admin/PM), "
        f"then the USER ACTIONS steps. The KPI Score Breakdown shows how every point is calculated."
    )
    info_box(ws, 2, 1, 7, intro, bg, tc, h=44)

    r = 3
    for si, sc in enumerate(scenarios, 1):
        gbg_map = {
            "A": (C["grade_a_bg"], C["grade_a_txt"]),
            "B": (C["grade_b_bg"], C["grade_b_txt"]),
            "C": (C["grade_c_bg"], C["grade_c_txt"]),
            "D": (C["grade_d_bg"], C["grade_d_txt"]),
        }
        gbg, gtc = gbg_map.get(sc["grade_target"], (C["row_a"], C["dark"]))

        # Scenario header
        section(ws, r, 7,
            f"SCENARIO {si}  ▸  {sc['title']}  —  TARGET GRADE {sc['grade_target']}  ({sc['score_range']})",
            bg=gbg, tc=gtc)
        r += 1

        # Column headers for this scenario
        for ci, h in enumerate(["Step", "Category", "Action / Data to Enter",
                                  "Where in PMS", "Field / Value", "KPI / Report Impact", "Points"], 1):
            cell(ws, r, ci, h, bg=tc, bold=True, color="FFFFFF", ha="center", size=9)
        ws.row_dimensions[r].height = 22
        r += 1

        for step in sc["steps"]:
            rb = C["row_a"] if r % 2 == 0 else C["row_b"]
            cat_bg = {
                "SETUP":    C["setup_bg"],
                "ACTION":   C["action_bg"],
                "RESULT":   C["result_bg"],
                "REPORT":   C["report_bg"],
                "ADMIN":    C["manual_bg"],
            }.get(step.get("cat", ""), rb)
            cat_tc = {
                "SETUP":    C["setup_txt"],
                "ACTION":   C["action_txt"],
                "RESULT":   C["result_txt"],
                "REPORT":   C["report_txt"],
                "ADMIN":    C["manual_txt"],
            }.get(step.get("cat", ""), C["dark"])

            cell(ws, r, 1, step.get("step", ""),   bg=cat_bg, ha="center", bold=True, size=9, color=cat_tc)
            cell(ws, r, 2, step.get("cat", ""),    bg=cat_bg, bold=True, size=9, color=cat_tc, ha="center")
            cell(ws, r, 3, step.get("action", ""), bg=rb, size=9, bold=step.get("bold", False))
            cell(ws, r, 4, step.get("where", ""),  bg=rb, size=9)
            cell(ws, r, 5, step.get("value", ""),  bg=cat_bg, size=9, color=cat_tc)
            cell(ws, r, 6, step.get("impact", ""), bg=rb, size=9, italic=True)
            pts = step.get("pts", "")
            pts_bg = gbg if pts and pts not in ("—", "") else rb
            cell(ws, r, 7, pts, bg=pts_bg, bold=True, ha="center", size=10, color=gtc)
            ws.row_dimensions[r].height = 40
            r += 1

        # Score summary
        cell(ws, r, 1, "TOTAL", bg=tc, bold=True, color="FFFFFF", ha="center", size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell(ws, r, 2,
            f"Expected Total Score: {sc['total_score']} / 100  →  Grade {sc['grade_target']}  "
            f"({sc['score_range']})",
            bg=gbg, bold=True, color=gtc, ha="center", size=11)
        cell(ws, r, 7, f"{sc['total_score']}/100", bg=gbg, bold=True, color=gtc, ha="center", size=12)
        ws.row_dimensions[r].height = 28
        r += 2

    col_widths(ws, [6, 10, 42, 28, 28, 30, 10])
    freeze(ws, "A3")
    return ws

# =============================================================================
# SHEET 3  ▸  Developer Scenarios
# =============================================================================
dev_scenarios = [
    {
        "title": "Developer who completes all sprint tasks on time and logs learning",
        "grade_target": "A",
        "score_range": "90 – 100 pts",
        "total_score": 100,
        "steps": [
            # Setup
            {"step": "S1", "cat": "SETUP",  "action": "Create a sprint in the project",
             "where": "Board → Sprint Manager → New Sprint",
             "value": "Name: 'June Sprint'\nStart: Jun 1 | End: Jun 30\nGoal: Ship v2.0 login feature",
             "impact": "Enables Sprint Reliability metric", "pts": "—"},
            {"step": "S2", "cat": "SETUP",  "action": "Create 10 TASK work items and assign to Developer",
             "where": "Board → New Work Item (×10)",
             "value": "Type: TASK\nAssignee: Developer Alice\nStory Points: 2 each\nEstimated Hours: 8 each\nDue Date: Jun 25\nSprint: June Sprint",
             "impact": "Feeds Sprint Reliability, Delivery Timeliness, Estimation Accuracy, Throughput", "pts": "—"},
            {"step": "S3", "cat": "SETUP",  "action": "Activate the sprint",
             "where": "Board → Sprint Manager → Activate",
             "value": "Status → Active",
             "impact": "Sprint items now count for Sprint Reliability", "pts": "—"},
            # Actions
            {"step": "1",  "cat": "ACTION", "action": "Move each of the 10 tasks to QA_DONE by Jun 25 (before due date)",
             "where": "Board → Kanban → drag each card to QA_DONE",
             "value": "All 10 items:\nstatus = QA_DONE\ncompletedAt = auto-set (e.g. Jun 20)\n(must be BEFORE dueDate Jun 25)",
             "impact": "Sprint Reliability · Delivery Timeliness · Throughput", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 8 hours per day in timesheet (20 working days × 8 hrs = 160 hrs total)",
             "where": "Timesheet → Log Hours",
             "value": "Work Item: select the task\nDate: each working day\nHours: 8\n(10 tasks × 8 estimated hrs = 80 hrs estimated;\nactual logged ≈ 80 hrs → 0% variance)",
             "impact": "Estimation Accuracy", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Do NOT set any item to BLOCKED status",
             "where": "Board (leave items as IN_PROGRESS or QA_DONE)",
             "value": "blockedCount = 0\ntotalItems = 10\nblocked ratio = 0%",
             "impact": "Dependency & Agile Management", "pts": "5"},
            {"step": "4",  "cat": "ACTION", "action": "Do NOT reopen any completed (QA_DONE) item",
             "where": "Board",
             "value": "reopenCount = 0 on all items",
             "impact": "Internal Rework Ratio", "pts": "5"},
            {"step": "5",  "cat": "ACTION", "action": "Ensure no BUG with YOU as 'Responsible User' is logged",
             "where": "Board → Bugs (check responsibleUserId field)",
             "value": "No BUG work items with Alice as Responsible User",
             "impact": "Defect Leakage", "pts": "10"},
            {"step": "6",  "cat": "ACTION", "action": "Log learning activities (e.g. React course): ≥4 hours",
             "where": "KPI → Self-Log → Learning",
             "value": "Period: 2026-06\nTopic: Advanced React Patterns\nHours: 5\nDescription: Completed Udemy course",
             "impact": "Learning Velocity", "pts": "5"},
            {"step": "7",  "cat": "ACTION", "action": "Log an AI tool or automation you built this month",
             "where": "KPI → Self-Log → Innovation",
             "value": "Period: 2026-06\nTitle: AI Code Review Helper\nImpact: Reduced PR review time by 40%\nType: AI_IMPLEMENTATION",
             "impact": "Automation & Innovation", "pts": "5"},
            {"step": "8",  "cat": "ACTION", "action": "Submit a leave request for any planned absence (1 day max for Grade A)",
             "where": "Leave Module → Submit Request",
             "value": "Type: Casual Leave\nStart: Jun 10 | End: Jun 10\n(Manager approves → status = APPROVED)",
             "impact": "Attendance (1 approved day ≤ 1 → 5 pts)", "pts": "5"},
            # Admin
            {"step": "A1", "cat": "ADMIN",  "action": "Admin enters Engineering Hygiene score: 5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Employee: Alice\nPeriod: 2026-06\nEngineering Hygiene: 5",
             "impact": "Engineering Hygiene metric", "pts": "5"},
            {"step": "A2", "cat": "ADMIN",  "action": "Admin enters Reporting & Documentation: 5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Reporting & Documentation: 5",
             "impact": "Reporting & Docs metric", "pts": "5"},
            {"step": "A3", "cat": "ADMIN",  "action": "Admin enters Positive Behaviour: 5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Positive Behaviour: 5",
             "impact": "Positive Behaviour metric", "pts": "5"},
            # Report
            {"step": "R1", "cat": "REPORT", "action": "Check KPI Appraisal page",
             "where": "KPI Appraisal → select Alice → Period: Jun 2026",
             "value": "Expected: 100/100 — Grade A",
             "impact": "All 13 metrics visible in breakdown", "pts": "100"},
            {"step": "R2", "cat": "REPORT", "action": "Check Productivity Report",
             "where": "Reports → Productivity → Period: Jun 2026",
             "value": "Alice shows:\n• tasksDone: 10\n• hoursLogged: ~80\n• onTimePct: 100%\n• score: high",
             "impact": "Productivity Report", "pts": "—"},
        ]
    },
    {
        "title": "Developer with partial sprint completion, 2 tasks reopened",
        "grade_target": "B",
        "score_range": "75 – 89 pts",
        "total_score": 79,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "Create 10 TASK work items in sprint, assign to Developer Bob",
             "where": "Board → New Work Item (×10)",
             "value": "Type: TASK | Assignee: Bob\nStory Points: 2 each | Estimated Hours: 8 each\nDue Date: Jun 25 | Sprint: June Sprint",
             "impact": "Baseline for all delivery metrics", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Complete 8 out of 10 sprint items (move to QA_DONE before due date)",
             "where": "Board → Kanban → QA_DONE",
             "value": "8 items QA_DONE by Jun 25\n2 items remain IN_PROGRESS",
             "impact": "Sprint Reliability: (16sp ÷ 20sp) × 15 = 12 pts\nThroughput: (8÷10) × 10 = 8 pts\nDelivery Timeliness: (8÷8) × 15 = 15 pts (only items with completedAt count)", "pts": "12+15+8"},
            {"step": "2",  "cat": "ACTION", "action": "Log 8 hrs/day in timesheet (actual 70 hrs vs estimated 80 hrs)",
             "where": "Timesheet → Log Hours",
             "value": "Total hours logged: 70\nTotal estimated: 80\nVariance: |70-80|/80 = 12.5% (≤15%)",
             "impact": "Estimation Accuracy: ≤15% variance → 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Reopen 2 completed items (move back from QA_DONE to IN_PROGRESS)",
             "where": "Board → drag card from QA_DONE back to IN_PROGRESS",
             "value": "2 items: reopenCount increments to 1\ntotalReopens = 2\ntotalCompleted = 8",
             "impact": "Rework Ratio: 2÷8 = 25% (>10%) → 0 pts", "pts": "0"},
            {"step": "4",  "cat": "ACTION", "action": "Log 2 hours of learning only",
             "where": "KPI → Self-Log → Learning",
             "value": "Period: 2026-06\nHours: 2\n(≥1 hr but <4 hrs)",
             "impact": "Learning Velocity: ≥1 hr → 3 pts", "pts": "3"},
            {"step": "5",  "cat": "ACTION", "action": "Log innovation: process improvement (not AI)",
             "where": "KPI → Self-Log → Innovation",
             "value": "Type: OTHER (not AI_IMPLEMENTATION)\nTitle: Improved CI pipeline script",
             "impact": "Automation & Innovation: no AI → 3 pts", "pts": "3"},
            {"step": "6",  "cat": "ACTION", "action": "No blocked items, no bugs against Bob, 1 approved leave day",
             "where": "Board / Leave Module",
             "value": "blockedCount = 0 → Dependency = 5 pts\nNo bugs with Bob as Responsible User → Defect = 10 pts\n1 approved leave day → Attendance = 5 pts",
             "impact": "Dependency + Defect Leakage + Attendance", "pts": "5+10+5"},
            {"step": "A1", "cat": "ADMIN",  "action": "Admin enters manual scores: Hygiene=5, Reporting=3, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Engineering Hygiene: 5\nReporting & Docs: 3 (some missed updates)\nPositive Behaviour: 5",
             "impact": "3 manual metrics", "pts": "5+3+5"},
            {"step": "R1", "cat": "REPORT", "action": "Check KPI Appraisal → Bob",
             "where": "KPI Appraisal → select Bob",
             "value": "Expected: 79/100 — Grade B",
             "impact": "Score breakdown visible in radar chart", "pts": "79"},
        ]
    },
    {
        "title": "Newly added Developer — no tasks assigned yet (explains Grade D default)",
        "grade_target": "D",
        "score_range": "Below 60 pts",
        "total_score": 35,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "Add Developer Carol to the project (no tasks assigned yet)",
             "where": "Project → Members → Add Member",
             "value": "User: Carol\nProjectRole: DEVELOPER\nNo work items assigned",
             "impact": "Carol appears in KPI list", "pts": "—"},
            {"step": "!", "cat": "RESULT",
             "action": "With NO tasks assigned and NO manual scores entered, Carol's auto score is 35/100 (Grade D).\nThis is expected system behaviour — it does NOT mean Carol is a poor performer.",
             "where": "KPI Appraisal → Carol",
             "value": "Sprint Reliability: 0 (no sprint items)\nDelivery Timeliness: 0 (no items with dueDate)\nEstimation Accuracy: 10 (both 0 = 10)\nThroughput: 0 (no items)\nRework Ratio: 5 (default for no completions)\nDefect Leakage: 10 (no bugs)\nEngineering Hygiene: 0 (not entered)\nDependency & Agile: 5 (no items = 5)\nReporting & Docs: 0 (not entered)\nLearning Velocity: 0 (no self-log)\nInnovation: 0 (no self-log)\nAttendance: 5 (no leaves)\nPositive Behaviour: 0 (not entered)\n─────────────────────────\nTotal = 35 / 100",
             "impact": "The KPI page shows 'N/A — Insufficient Data' badge for new members with no activity",
             "pts": "35"},
            {"step": "F1", "cat": "ACTION", "action": "To fix: assign tasks to Carol and have admin enter manual scores",
             "where": "Board → assign work items to Carol\nKPI → Enter Monthly Scores",
             "value": "Once Carol has at least 1 task assigned\nand manual scores entered,\nKPI will show a meaningful grade.",
             "impact": "Score will increase from 35 to a fair assessment", "pts": "—"},
        ]
    },
]
build_role_sheet(wb, "3. Developer Scenarios", "DEVELOPER", "1D4ED8", dev_scenarios)

# =============================================================================
# SHEET 4  ▸  QA Scenarios
# =============================================================================
qa_scenarios = [
    {
        "title": "QA who tests all items, reports bugs correctly, logs learning",
        "grade_target": "A",
        "score_range": "90 – 100 pts",
        "total_score": 97,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "Create 8 TASK/BUG-type work items and assign to QA Priya",
             "where": "Board → New Work Item (×8)",
             "value": "Type: TASK\nAssignee: QA Priya\nEstimated Hours: 4 each\nDue Date: Jun 20\nStory Points: 1 each",
             "impact": "Baseline for delivery metrics", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Move all 8 assigned items to QA_DONE before due date",
             "where": "Board → Kanban → QA_DONE",
             "value": "All 8 items: status = QA_DONE\ncompletedAt = before Jun 20",
             "impact": "Sprint Rel: 8÷8 × 15 = 15 pts\nDelivery: 8÷8 × 15 = 15 pts\nThroughput: 8÷8 × 10 = 10 pts", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 4 hrs/day in timesheet (actual = 32 hrs, estimated = 32 hrs)",
             "where": "Timesheet → Log Hours",
             "value": "Total hours: 32\nEstimated: 32\nVariance: 0% → ≤15%",
             "impact": "Estimation Accuracy: 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Create 2 MINOR bugs found during testing — set Responsible User = Developer (NOT QA)",
             "where": "Board → New Work Item",
             "value": "Type: BUG\nSeverity: MINOR (each)\nReporter: Priya (QA)\nResponsible User: Developer Alice (the dev whose code had the bug)\nClassification: NEW_BUG",
             "impact": "⚠ Priya is the REPORTER but Alice is RESPONSIBLE.\nPriya's Defect Leakage = 10 pts (no bugs against HER).\nAlice's Defect Leakage = 4 pts (2 MINOR bugs against her).", "pts": "10"},
            {"step": "4",  "cat": "ACTION", "action": "Do NOT have any bugs where Priya is set as Responsible User",
             "where": "Board (verify responsibleUserId on all bugs)",
             "value": "Priya has 0 bugs where she is Responsible User",
             "impact": "Defect Leakage: 0 bugs against Priya → 10 pts", "pts": "10"},
            {"step": "5",  "cat": "ACTION", "action": "Log 5 hours of learning (testing tools, ISTQB preparation)",
             "where": "KPI → Self-Log → Learning",
             "value": "Period: 2026-06\nTopic: ISTQB Foundation Level Prep\nHours: 5",
             "impact": "Learning Velocity: ≥4 hrs → 5 pts", "pts": "5"},
            {"step": "6",  "cat": "ACTION", "action": "Log AI tool for test automation",
             "where": "KPI → Self-Log → Innovation",
             "value": "Type: AI_IMPLEMENTATION\nTitle: AI-Powered Test Case Generator\nImpact: Reduced test writing time by 50%",
             "impact": "Automation & Innovation: AI type → 5 pts", "pts": "5"},
            {"step": "7",  "cat": "ACTION", "action": "No blocked items, 1 approved leave day, no reopens",
             "where": "Board / Leave Module",
             "value": "blockedCount = 0 → 5 pts\nAttendance: 1 day → 5 pts\nReopens = 0 → 5 pts",
             "impact": "Dependency + Attendance + Rework", "pts": "5+5+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters: Hygiene=5, Reporting=5, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "All three manual metrics: 5 each",
             "impact": "Manual metrics", "pts": "5+5+5"},
            {"step": "R1", "cat": "REPORT", "action": "Check Bugs Report for the project",
             "where": "Reports → Bugs → Period: Jun 2026",
             "value": "Shows 2 MINOR bugs in the chart\nClassification: NEW_BUG × 2\n(CSV export also shows classification)",
             "impact": "Bugs Report", "pts": "—"},
            {"step": "R2", "cat": "REPORT", "action": "Check Developer Alice's KPI after Priya reports bugs",
             "where": "KPI Appraisal → select Alice",
             "value": "Alice's Defect Leakage: 2 MINOR bugs\ntotalSevere = 2 × 1 = 2 → 4 pts (not 10)",
             "impact": "Confirms Responsible User logic is correct", "pts": "—"},
        ]
    },
    {
        "title": "QA who reports critical bugs (penalises the responsible developer)",
        "grade_target": "B",
        "score_range": "75 – 89 pts",
        "total_score": 78,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "Create 6 TASK items assigned to QA Raj",
             "where": "Board → New Work Item (×6)",
             "value": "Assignee: QA Raj | Story Points: 1 each\nEstimated Hours: 6 each | Due Date: Jun 28",
             "impact": "Baseline setup", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Complete 5 of 6 items before due date, 1 item missed",
             "where": "Board → QA_DONE",
             "value": "5 items QA_DONE before Jun 28\n1 item completed Jun 30 (AFTER due date)",
             "impact": "Sprint Rel: 5÷6 × 15 = 12.5 pts\nDelivery: 5÷6 × 15 = 12.5 pts\nThroughput: 5÷6 × 10 = 8.3 pts", "pts": "12+12+8"},
            {"step": "2",  "cat": "ACTION", "action": "Log 30 hrs (estimated 36 hrs → 16.6% over budget)",
             "where": "Timesheet → Log Hours",
             "value": "Total actual: 30 hrs\nTotal estimated: 36 hrs\nVariance: |30-36|/36 = 16.6% (16–30%)",
             "impact": "Estimation Accuracy: 16–30% variance → 7 pts", "pts": "7"},
            {"step": "3",  "cat": "ACTION", "action": "Create 1 CRITICAL bug with Developer Bob as Responsible User",
             "where": "Board → New Work Item",
             "value": "Type: BUG | Severity: CRITICAL\nResponsible User: Developer Bob\nReporter: QA Raj",
             "impact": "Raj's Defect Leakage: 0 bugs against Raj → 10 pts\nBob's Defect Leakage: 1 CRITICAL → 0 pts\n(criticalOrBlocker > 0 = 0 pts for Bob)", "pts": "10"},
            {"step": "4",  "cat": "ACTION", "action": "Log learning: 2 hrs | No innovation logged",
             "where": "KPI → Self-Log → Learning",
             "value": "Hours: 2 (≥1 but <4 hrs)\nNo innovation log",
             "impact": "Learning: 3 pts | Innovation: 0 pts", "pts": "3+0"},
            {"step": "5",  "cat": "ACTION", "action": "1 item in BLOCKED status, 0 reopens, 1 approved leave day",
             "where": "Board / Leave Module",
             "value": "blockedCount = 1 | total = 6\nblocked ratio = 1/6 = 16.7% (>10%)\nReopens = 0 → Rework = 5 pts\nAttendance = 5 pts",
             "impact": "Dependency: >10% → 0 pts\nRework: 5 pts | Attendance: 5 pts", "pts": "0+5+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters: Hygiene=5, Reporting=3, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Engineering Hygiene: 5\nReporting & Docs: 3\nPositive Behaviour: 5",
             "impact": "Manual metrics", "pts": "5+3+5"},
        ]
    },
]
build_role_sheet(wb, "4. QA Scenarios", "QA", "15803D", qa_scenarios)

# =============================================================================
# SHEET 5  ▸  Team Lead Scenarios
# =============================================================================
tl_scenarios = [
    {
        "title": "Team Lead who sets up sprint, codes, reviews, logs learning (Grade A)",
        "grade_target": "A",
        "score_range": "90 – 100 pts",
        "total_score": 95,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "PM creates 8 TASK items and assigns 4 to Team Lead Suresh",
             "where": "Board → New Work Item (×4)",
             "value": "Assignee: Suresh (TEAM_LEAD)\nStory Points: 3 each\nEstimated Hours: 10 each\nDue Date: Jun 22",
             "impact": "TL also gets individual KPI — not just team oversight", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Move all 4 TL tasks to QA_DONE before Jun 22",
             "where": "Board → QA_DONE",
             "value": "All 4 items QA_DONE by Jun 20",
             "impact": "Sprint Rel: 12÷12 × 15 = 15 pts\nDelivery: 4÷4 × 15 = 15 pts\nThroughput: 4÷4 × 10 = 10 pts", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 10 hrs per item in timesheet (40 hrs total = estimated)",
             "where": "Timesheet → Log Hours",
             "value": "Actual: 40 hrs | Estimated: 40 hrs\nVariance: 0%",
             "impact": "Estimation Accuracy: 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "No items in BLOCKED, no reopens",
             "where": "Board",
             "value": "blockedCount = 0 → Dependency = 5 pts\nReopens = 0 → Rework = 5 pts",
             "impact": "Dependency + Rework", "pts": "5+5"},
            {"step": "4",  "cat": "ACTION", "action": "Log 6 hrs of learning (architecture articles + internal tech talk)",
             "where": "KPI → Self-Log → Learning",
             "value": "Period: 2026-06\nTopic: System Design Architecture\nHours: 6",
             "impact": "Learning Velocity: ≥4 hrs → 5 pts", "pts": "5"},
            {"step": "5",  "cat": "ACTION", "action": "Log an AI tool or automation",
             "where": "KPI → Self-Log → Innovation",
             "value": "Type: AI_IMPLEMENTATION\nTitle: AI-based code linting assistant\nImpact: Cut code review time by 30%",
             "impact": "Innovation: AI type → 5 pts", "pts": "5"},
            {"step": "6",  "cat": "ACTION", "action": "Ensure no bugs have Suresh as Responsible User, 1 approved leave",
             "where": "Board / Leave Module",
             "value": "No CRITICAL/MAJOR bugs against Suresh\nAttendance: 1 approved day → 5 pts\nDefect Leakage: 10 pts",
             "impact": "Defect Leakage + Attendance", "pts": "10+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters: Hygiene=5, Reporting=5, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "All three: 5 pts each",
             "impact": "Manual metrics", "pts": "5+5+5"},
            {"step": "R1", "cat": "REPORT", "action": "Team Lead also visible in Overview Dashboard (as management user)",
             "where": "Overview Dashboard",
             "value": "Stat cards show:\nActive Projects: (their projects)\nTeam Members: (members in their projects)\nTotal Tasks: (tasks in their projects)",
             "impact": "TL with EMPLOYEE system role now gets management view\nif they hold TEAM_LEAD project role", "pts": "—"},
        ]
    },
]
build_role_sheet(wb, "5. Team Lead Scenarios", "TEAM_LEAD", "0F766E", tl_scenarios)

# =============================================================================
# SHEET 6  ▸  Project Manager Scenarios
# =============================================================================
pm_scenarios = [
    {
        "title": "PM who sets up sprint correctly so team KPI auto-calculates",
        "grade_target": "A",
        "score_range": "Focus on enabling team — own KPI also tracked",
        "total_score": 90,
        "steps": [
            {"step": "!", "cat": "RESULT",
             "action": "The PM's most important KPI contribution is SETTING UP the correct data so the TEAM's auto-metrics work.\n"
                       "Without due dates, story points, and estimated hours — the team gets 0 on Sprint Reliability, Delivery Timeliness, and Estimation Accuracy.",
             "where": "Board → Work Items",
             "value": "Missing due date → Delivery Timeliness = 0 for whole team\nMissing story points → Sprint Reliability uses count=1 (less accurate)\nMissing estimated hours → Estimation Accuracy = 0",
             "impact": "Critical — PM data entry quality determines team KPI accuracy", "pts": "—"},
            {"step": "S1", "cat": "SETUP", "action": "Create Sprint with correct dates",
             "where": "Board → Sprint Manager → New Sprint",
             "value": "Name: 'Sprint 1'\nStart: Jun 1\nEnd: Jun 30\nGoal: Feature X launch",
             "impact": "Enables Sprint Reliability for all sprint members", "pts": "—"},
            {"step": "S2", "cat": "SETUP", "action": "Create all work items for the sprint with COMPLETE metadata",
             "where": "Board → New Work Item",
             "value": "For EVERY work item:\n• assigneeId: set to a team member\n• sprintId: this sprint\n• storyPoints: 1–5 (based on complexity)\n• estimatedHours: realistic estimate (e.g., 8 for a day's work)\n• dueDate: set within sprint duration\n• type: TASK / STORY / SUB_TASK",
             "impact": "Enables all 4 delivery metrics for each assignee", "pts": "—"},
            {"step": "S3", "cat": "SETUP", "action": "Activate the sprint at start of month",
             "where": "Board → Sprint Manager → Activate",
             "value": "Status → Active",
             "impact": "Sprint items now qualify for Sprint Reliability calculation", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "PM completes their own assigned tasks (4 items)",
             "where": "Board → QA_DONE",
             "value": "4 items QA_DONE before due date\n(PM also has 4 tasks assigned in sprint)",
             "impact": "PM's own Sprint Rel/Delivery/Throughput\n15+15+10 if all done", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 8 hrs/day in timesheet for PM's own tasks",
             "where": "Timesheet → Log Hours",
             "value": "Actual ≈ Estimated (within 15%)",
             "impact": "PM's Estimation Accuracy: 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Approve team's weekly timesheet entries",
             "where": "Timesheet → Approval Queue",
             "value": "Review and approve each Friday\n(Keeps Timesheet Report accurate)",
             "impact": "Timesheet Report shows Approved status", "pts": "—"},
            {"step": "4",  "cat": "ACTION", "action": "Approve team leave requests within 24 hours",
             "where": "Leave Module → Pending",
             "value": "Status → APPROVED\n(Enables team Attendance KPI)",
             "impact": "Team Attendance metric — leave not approved = 0 pts for employee", "pts": "—"},
            {"step": "5",  "cat": "ACTION", "action": "Close sprint at end of month",
             "where": "Board → Sprint Manager → Close",
             "value": "Status → Closed\nAll remaining items move to backlog",
             "impact": "Finalises Sprint Reliability for the month", "pts": "—"},
            {"step": "6",  "cat": "ACTION", "action": "PM logs learning and innovation",
             "where": "KPI → Self-Log → Learning / Innovation",
             "value": "Learning: 5 hrs (project management training)\nInnovation: AI_IMPLEMENTATION (project risk predictor tool)",
             "impact": "Learning: 5 pts | Innovation: 5 pts", "pts": "5+5"},
            {"step": "7",  "cat": "ACTION", "action": "No blocked items, no bugs against PM, 1 approved leave",
             "where": "Board / Leave Module",
             "value": "Dependency: 5 pts | Rework: 5 pts\nDefect: 10 pts | Attendance: 5 pts",
             "impact": "4 auto-scored metrics", "pts": "5+5+10+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters manual KPI scores for PM",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "Engineering Hygiene: 5\nReporting & Docs: 5 (PM keeps excellent docs)\nPositive Behaviour: 5",
             "impact": "Manual metrics", "pts": "5+5+5"},
            {"step": "A2", "cat": "ADMIN", "action": "PM enters manual KPI scores for ALL team members (by 5th of next month)",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "For each team member, enter:\n• Engineering Hygiene: 0/3/5\n• Reporting & Documentation: 0/3/5\n• Positive Behaviour: 0/3/5",
             "impact": "Without these entries, each team member loses up to 15 pts\n(3 metrics × 5 pts each = 15 pts of manual scoring)", "pts": "—"},
            {"step": "R1", "cat": "REPORT", "action": "Check Overview Dashboard as PM",
             "where": "Overview Dashboard",
             "value": "Active Projects: count of PM's projects\nTeam Members: members in those projects\nTotal Tasks: tasks across those projects\n(NOT global system stats)",
             "impact": "Scoped dashboard for PM/TL roles", "pts": "—"},
            {"step": "R2", "cat": "REPORT", "action": "Check Projects Report",
             "where": "Reports → Projects",
             "value": "Your project shows:\n• All tasks created this month\n• Completion %\n• Team size",
             "impact": "Projects Report", "pts": "—"},
        ]
    },
]
build_role_sheet(wb, "6. Project Manager Scenarios", "PROJECT_MANAGER", "5B21B6", pm_scenarios)

# =============================================================================
# SHEET 7  ▸  Designer & DevOps
# =============================================================================
other_scenarios = [
    {
        "title": "Designer — completes UI tasks, reports UI bugs, logs design learning",
        "grade_target": "A",
        "score_range": "90 – 100 pts",
        "total_score": 95,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "PM creates 6 design tasks and assigns to Designer Deepa",
             "where": "Board → New Work Item (×6)",
             "value": "Type: TASK\nAssignee: Deepa\nStory Points: 2 each\nEstimated Hours: 6 each\nDue Date: Jun 22",
             "impact": "Baseline for delivery metrics", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Complete all 6 design tasks (QA_DONE) before Jun 22",
             "where": "Board → QA_DONE",
             "value": "All 6 items QA_DONE by Jun 20",
             "impact": "Sprint Rel: 15 pts | Delivery: 15 pts | Throughput: 10 pts", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 6 hrs/task in timesheet (actual 36 hrs = estimated 36 hrs)",
             "where": "Timesheet → Log Hours",
             "value": "Variance: 0%",
             "impact": "Estimation Accuracy: 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Create 1 UI_USABILITY bug (set DEVELOPER as Responsible User)",
             "where": "Board → New Work Item",
             "value": "Type: BUG | Severity: MINOR\nClassification: UI_USABILITY\nResponsible User: Developer (NOT Deepa)",
             "impact": "Deepa's Defect Leakage: 10 pts (no bugs against her)\nThis bug appears in Bugs Report under UI_USABILITY", "pts": "10"},
            {"step": "4",  "cat": "ACTION", "action": "Log learning: 5 hrs (Figma Advanced, design systems)",
             "where": "KPI → Self-Log → Learning",
             "value": "Hours: 5 → ≥4 hrs",
             "impact": "Learning Velocity: 5 pts", "pts": "5"},
            {"step": "5",  "cat": "ACTION", "action": "Log innovation: AI design tool implementation",
             "where": "KPI → Self-Log → Innovation",
             "value": "Type: AI_IMPLEMENTATION\nTitle: AI-assisted UI mockup generator",
             "impact": "Innovation: 5 pts", "pts": "5"},
            {"step": "6",  "cat": "ACTION", "action": "No blocked items, no reopens, 1 approved leave",
             "where": "Board / Leave Module",
             "value": "Dependency: 5 pts | Rework: 5 pts | Attendance: 5 pts",
             "impact": "3 auto metrics", "pts": "5+5+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters: Hygiene=5, Reporting=5, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "All three: 5 pts each",
             "impact": "Manual metrics", "pts": "5+5+5"},
        ]
    },
    {
        "title": "DevOps — pipeline tasks, AI automation, logging innovation",
        "grade_target": "A",
        "score_range": "90 – 100 pts",
        "total_score": 95,
        "steps": [
            {"step": "S1", "cat": "SETUP", "action": "PM creates 5 DevOps tasks (CI/CD, infra) and assigns to DevOps Jay",
             "where": "Board → New Work Item (×5)",
             "value": "Type: TASK\nAssignee: Jay (DEVOPS)\nStory Points: 3 each\nEstimated Hours: 10 each\nDue Date: Jun 28",
             "impact": "Baseline", "pts": "—"},
            {"step": "1",  "cat": "ACTION", "action": "Complete all 5 DevOps tasks (QA_DONE) before Jun 28",
             "where": "Board → QA_DONE",
             "value": "All 5 completed on time",
             "impact": "Sprint Rel: 15 pts | Delivery: 15 pts | Throughput: 10 pts", "pts": "15+15+10"},
            {"step": "2",  "cat": "ACTION", "action": "Log 10 hrs/task in timesheet (50 hrs actual vs 50 hrs estimated)",
             "where": "Timesheet → Log Hours",
             "value": "Variance: 0%",
             "impact": "Estimation Accuracy: 10 pts", "pts": "10"},
            {"step": "3",  "cat": "ACTION", "action": "Log innovation: new AI deployment pipeline",
             "where": "KPI → Self-Log → Innovation",
             "value": "Type: AI_IMPLEMENTATION\nTitle: AI-assisted zero-downtime deployment\nImpact: Deployment time reduced from 20 to 3 minutes",
             "impact": "Innovation: AI type → 5 pts\n(DevOps has biggest opportunity here!)", "pts": "5"},
            {"step": "4",  "cat": "ACTION", "action": "Log 4 hrs of learning (Kubernetes, cloud certifications)",
             "where": "KPI → Self-Log → Learning",
             "value": "Hours: 4 (meets ≥4 threshold exactly)",
             "impact": "Learning Velocity: 5 pts", "pts": "5"},
            {"step": "5",  "cat": "ACTION", "action": "No blocked items, no bugs against Jay, 1 approved leave",
             "where": "Board / Leave Module",
             "value": "Dependency: 5 | Rework: 5 | Defect: 10 | Attendance: 5",
             "impact": "4 auto metrics", "pts": "5+5+10+5"},
            {"step": "A1", "cat": "ADMIN", "action": "Admin enters: Hygiene=5, Reporting=5, Behaviour=5",
             "where": "KPI Appraisal → Enter Monthly Scores",
             "value": "All three: 5 pts each",
             "impact": "Manual metrics", "pts": "5+5+5"},
        ]
    },
]
build_role_sheet(wb, "7. Designer & DevOps", "DESIGNER", "9F1239", other_scenarios)

# =============================================================================
# SHEET 8  ▸  KPI Formula Reference (all 13 formulas with worked examples)
# =============================================================================
ws8 = wb.create_sheet("8. KPI Formula Reference")
ws8.sheet_view.showGridLines = False

page_hdr(ws8, 1, 8,
    "KPI FORMULA REFERENCE — All 13 Metrics: Exact Code Logic, Thresholds & Worked Examples",
    C["indigo"], size=13)

info_box(ws8, 2, 1, 8,
    "These formulas are implemented exactly as shown in backend/src/analytics/analytics.service.ts. "
    "The 'Period' is always a calendar month (YYYY-MM). Data must be CREATED within the period dates to be counted. "
    "Sprint-based projects use sprint items; non-sprint projects fall back to ALL assigned items.",
    "EFF6FF", C["indigo"], h=44)

col_hdrs8 = ["#", "Metric Name", "Category", "Max\nPts",
             "Exact Formula (from source code)",
             "Score Thresholds",
             "Worked Example (Grade A)",
             "Worked Example (Grade D / 0)"]
for ci, h in enumerate(col_hdrs8, 1):
    cell(ws8, 3, ci, h, bg=C["indigo"], bold=True, color="FFFFFF", ha="center", size=9)
ws8.row_dimensions[3].height = 28

formulas = [
    (1, "Sprint Reliability", "Delivery & Execution", 15,
     "isSprintBased = sprintItems.length > 0\n"
     "deliveryBase = isSprintBased ? sprintItems : allAssignedItems\n"
     "committed = sum(storyPoints ?? 1) for all deliveryBase items\n"
     "delivered = sum(storyPoints ?? 1) for deliveryBase items where status=QA_DONE\n"
     "score = committed=0 ? 0 : min((delivered/committed) × 15, 15)",
     "0% completion = 0 pts\n100% completion = 15 pts\n(proportional)",
     "10 sprint items × 2sp each\n8 QA_DONE × 2sp = 16sp delivered\n(16/20) × 15 = 12 pts",
     "committed=0 (no sprint items,\nno assigned items) → 0 pts"),

    (2, "Delivery Timeliness", "Delivery & Execution", 15,
     "itemsWithDue = assignedItems where dueDate≠null AND completedAt≠null\n"
     "onTime = itemsWithDue where completedAt ≤ dueDate\n"
     "score = itemsWithDue=0 ? 0 : min((onTime/itemsWithDue) × 15, 15)",
     "0 on-time = 0 pts\n100% on-time = 15 pts\n(proportional)",
     "8 items with dueDate, all 8 QA_DONE\nbefore dueDate → (8/8) × 15 = 15 pts",
     "No items have dueDate set →\nitemsWithDue=0 → 0 pts"),

    (3, "Estimation Accuracy", "Delivery & Execution", 10,
     "totalEstimated = sum(estimatedHours ?? 0) for deliveryBase\n"
     "totalActual = sum(timesheet.hours) for deliveryBase item IDs\n"
     "  (or all timesheet in period if no deliveryBase items)\n"
     "if estimated=0: return actual=0 ? 10 : 0\n"
     "variance = |actual-estimated| / estimated\n"
     "≤0.15→10 | ≤0.30→7 | ≤0.50→4 | else→0",
     "0% variance = 10 pts\n16–30% = 7 pts\n31–50% = 4 pts\n>50% = 0 pts",
     "Estimated: 80 hrs | Actual: 82 hrs\nVariance: 2/80 = 2.5% (≤15%)\n→ 10 pts",
     "Estimated: 80 hrs | Actual: 200 hrs\nVariance: 120/80 = 150% (>50%)\n→ 0 pts"),

    (4, "Throughput & Complexity", "Delivery & Execution", 10,
     "deliveryBase = sprintItems if sprint; else allAssignedItems\n"
     "total = deliveryBase.length\n"
     "completed = deliveryBase where status=QA_DONE\n"
     "score = total=0 ? 0 : min((completed/total) × 10, 10)",
     "0% done = 0 pts\n100% done = 10 pts\n(proportional)",
     "8 sprint items, all 8 QA_DONE\n(8/8) × 10 = 10 pts",
     "0 sprint items and 0 assigned\nitems in period → 0 pts"),

    (5, "Internal Rework Ratio", "Quality & Engineering Exc.", 5,
     "totalReopens = sum(reopenCount) for allAssignedItems\n"
     "totalCompleted = allAssignedItems where status=QA_DONE\n"
     "if totalCompleted=0: return 5  ← DEFAULT full score\n"
     "ratio = totalReopens / totalCompleted\n"
     "ratio=0→5 | ratio≤0.10→3 | else→0",
     "0% reopen = 5 pts\n≤10% reopen = 3 pts\n>10% reopen = 0 pts",
     "10 completed, 0 reopened\nratio = 0 → 5 pts",
     "10 completed, 2 reopened\nratio = 0.20 (>10%) → 0 pts"),

    (6, "Technical Defect Leakage", "Quality & Engineering Exc.", 10,
     "bugs = workItems where\n  type=BUG AND responsibleUserId=this_user\n  AND createdAt in period\n"
     "criticalOrBlocker = count(SHOW_STOPPER/CRITICAL/BLOCKER bugs)\n"
     "if criticalOrBlocker > 0: return 0\n"
     "minors = count(MINOR/TRIVIAL)\n"
     "majors = count(MAJOR)\n"
     "totalSevere = majors×2 + minors\n"
     "0→10 | 1→7 | 2→4 | ≥3→0\n"
     "⚠ Uses responsibleUserId NOT reporterId",
     "0 bugs = 10 pts\n1 MINOR = 7 pts\n2 MINOR = 4 pts\n1 CRITICAL or 3+ MINOR = 0 pts",
     "0 bugs with this user as\nResponsible → 10 pts",
     "1 CRITICAL bug with this\nuser as Responsible → 0 pts"),

    (7, "Engineering Hygiene", "Quality & Engineering Exc.", 5,
     "MANUAL — Admin enters in KPI → Enter Monthly Scores\n"
     "metricId = 'engineering_hygiene'\n"
     "options: 5 / 3 / 0\n"
     "DEFAULT = 0 if not entered",
     "5 = Excellent practices\n3 = Minor gaps\n0 = Violations or\n    not entered by admin",
     "Admin enters 5 (excellent code\nquality, linting, security) → 5 pts",
     "Admin has not entered score\nyet → 0 pts (default)"),

    (8, "Dependency & Agile Mgmt", "Ownership & Collaboration", 5,
     "blockedCount = allAssignedItems where status=BLOCKED\n"
     "totalItems = allAssignedItems.length\n"
     "if totalItems=0: return 5  ← DEFAULT full score\n"
     "ratio = blockedCount / totalItems\n"
     "ratio=0→5 | ratio≤0.10→3 | else→0",
     "0% blocked = 5 pts\n≤10% blocked = 3 pts\n>10% blocked = 0 pts",
     "10 items, 0 blocked\nratio = 0 → 5 pts",
     "10 items, 2 blocked\nratio = 0.20 (>10%) → 0 pts"),

    (9, "Reporting & Documentation", "Ownership & Collaboration", 5,
     "MANUAL — Admin enters in KPI → Enter Monthly Scores\n"
     "metricId = 'reporting_documentation'\n"
     "options: 5 / 3 / 0\n"
     "DEFAULT = 0 if not entered",
     "5 = Accurate, no gaps\n3 = Inconsistent / delays\n0 = Critical misses or\n    not entered by admin",
     "Admin enters 5 (all status updates\non time, clear documentation) → 5 pts",
     "Admin has not entered score\nyet → 0 pts (default)"),

    (10, "Learning Velocity", "Growth & Innovation", 5,
     "learningLogs = SelfLog.learningLog where\n  userId=this AND period=YYYY-MM\n"
     "totalHours = sum(hours) across those logs\n"
     "totalHours≥4→5 | totalHours≥1→3 | else→0",
     "≥4 hours = 5 pts\n≥1 hour = 3 pts\n0 hours = 0 pts",
     "5 hrs logged in Self-Log →\n≥4 hrs → 5 pts",
     "No self-log learning entries\nfor period → 0 pts"),

    (11, "Automation & Innovation", "Growth & Innovation", 5,
     "innovationLogs = SelfLog.innovationLog where\n  userId=this AND period=YYYY-MM\n"
     "if innovationLogs.length=0: return 0\n"
     "hasAI = any log where type=AI_IMPLEMENTATION\n"
     "hasAI→5 | else→3",
     "AI implementation = 5 pts\nOther improvement = 3 pts\nNo entries = 0 pts",
     "1 innovation log with\ntype=AI_IMPLEMENTATION → 5 pts",
     "No innovation logs for\nperiod → 0 pts"),

    (12, "Attendance", "Behaviour & Reliability", 5,
     "leaveRequests = LeaveRequest where\n  userId=this AND status in [APPROVED,REJECTED]\n  AND period overlaps\n"
     "hasRejected = any(status=REJECTED)\n"
     "if hasRejected: return 0\n"
     "approvedDays = sum of overlap days for APPROVED leaves\n"
     "approvedDays>1.5→0 | approvedDays>1→3 | else→5",
     "≤1 day approved leave = 5 pts\n1 to 1.5 days = 3 pts\n>1.5 days OR any rejected = 0 pts",
     "1 approved leave day in month\n→ ≤1 → 5 pts",
     "Manager rejected leave request\n(status=REJECTED) → 0 pts immediately"),

    (13, "Positive Behaviour", "Behaviour & Reliability", 5,
     "MANUAL — Admin enters in KPI → Enter Monthly Scores\n"
     "metricId = 'positive_behaviour'\n"
     "options: 5 / 3 / 0\n"
     "DEFAULT = 0 if not entered",
     "5 = Professional, punctual,\n    zero late comings\n3 = <3 minor late comings\n0 = >3 lates or misconduct",
     "Admin enters 5 (perfect\nconduct month) → 5 pts",
     "Admin has not entered score\nyet → 0 pts (default)"),
]

CAT_COLORS_LOCAL = {
    "Delivery & Execution":       ("1D4ED8", "EFF6FF"),
    "Quality & Engineering Exc.": ("15803D", "F0FDF4"),
    "Ownership & Collaboration":  ("5B21B6", "F5F3FF"),
    "Growth & Innovation":        ("92400E", "FFFBEB"),
    "Behaviour & Reliability":    ("9F1239", "FFF1F2"),
}

for ri, row in enumerate(formulas, 4):
    cat_tc, cat_bg = CAT_COLORS_LOCAL.get(row[2], (C["dark"], C["row_a"]))
    rb = C["row_a"] if ri % 2 == 0 else C["row_b"]
    cell(ws8, ri, 1, row[0],  bg=cat_bg, bold=True, ha="center", size=11, color=cat_tc)
    cell(ws8, ri, 2, row[1],  bg=cat_bg, bold=True, size=10, color=cat_tc)
    cell(ws8, ri, 3, row[2],  bg=cat_bg, italic=True, size=8, color=cat_tc)
    cell(ws8, ri, 4, row[3],  bg=rb, bold=True, ha="center", size=12, color=C["blue"])
    cell(ws8, ri, 5, row[4],  bg=rb, size=8)
    cell(ws8, ri, 6, row[5],  bg=C["action_bg"], size=9, color=C["action_txt"])
    cell(ws8, ri, 7, row[6],  bg=C["grade_a_bg"], size=9, color=C["grade_a_txt"])
    cell(ws8, ri, 8, row[7],  bg=C["grade_d_bg"], size=9, color=C["grade_d_txt"])
    ws8.row_dimensions[ri].height = 80

# Total score / Grade table
tr = len(formulas) + 5
section(ws8, tr, 8, "GRADE THRESHOLDS  &  SCORE BREAKDOWN BY CATEGORY", C["indigo"])
tr += 1

grade_rows = [
    ("Grade A — Excellent", "90 – 100 pts", C["grade_a_bg"], C["grade_a_txt"]),
    ("Grade B — Good",      "75 – 89 pts",  C["grade_b_bg"], C["grade_b_txt"]),
    ("Grade C — Average",   "60 – 74 pts",  C["grade_c_bg"], C["grade_c_txt"]),
    ("Grade D — Poor",      "Below 60 pts", C["grade_d_bg"], C["grade_d_txt"]),
]
cat_totals = [
    ("Delivery & Execution", "50 pts", "Sprint(15) + Delivery(15) + Estimation(10) + Throughput(10)", "1D4ED8", "EFF6FF"),
    ("Quality & Engineering Excellence", "20 pts", "Rework(5) + Defect Leakage(10) + Engineering Hygiene(5)", "15803D", "F0FDF4"),
    ("Ownership & Collaboration", "10 pts", "Dependency & Agile(5) + Reporting & Docs(5)", "5B21B6", "F5F3FF"),
    ("Growth & Innovation", "10 pts", "Learning Velocity(5) + Automation & Innovation(5)", "92400E", "FFFBEB"),
    ("Behaviour & Reliability", "10 pts", "Attendance(5) + Positive Behaviour(5)", "9F1239", "FFF1F2"),
]
for g, rng, gbg, gtc in grade_rows:
    cell(ws8, tr, 1, g,   bg=gbg, bold=True, color=gtc, size=11)
    ws8.merge_cells(start_row=tr, start_column=2, end_row=tr, end_column=4)
    cell(ws8, tr, 2, rng, bg=gbg, bold=True, color=gtc, ha="center", size=11)
    ws8.row_dimensions[tr].height = 22
    tr += 1

tr += 1
for cat, pts, breakdown, ctc, cbg in cat_totals:
    ws8.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    cell(ws8, tr, 1, cat,       bg=cbg, bold=True, color=ctc, size=10)
    cell(ws8, tr, 3, pts,       bg=cbg, bold=True, color=ctc, ha="center", size=11)
    ws8.merge_cells(start_row=tr, start_column=4, end_row=tr, end_column=8)
    cell(ws8, tr, 4, breakdown, bg=cbg, color=ctc, size=9)
    ws8.row_dimensions[tr].height = 22
    tr += 1

col_widths(ws8, [5, 22, 22, 7, 42, 30, 34, 34])
freeze(ws8, "A4")

# =============================================================================
# Save
# =============================================================================
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Role-Based KPI & Reports Scenarios Guide.xlsx")
wb.save(out)
print(f"Saved -> {out}")
